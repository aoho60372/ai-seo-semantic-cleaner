"""Single entry point for the model-directed SEO workflow."""

from __future__ import annotations

import argparse
import csv
import contextlib
import json
import logging
import os
import re
import shutil
import sys
import time
import traceback
import uuid
from pathlib import Path

# Hermes launches this module directly, without the PowerShell wrapper.  Keep
# the process environment and terminal streams deterministic before importing
# pandas, transformers, or any other library that may inspect them at import
# time.  seo.ps1 remains a convenience launcher for manual Windows use.
os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
os.environ.setdefault("HF_HUB_OFFLINE", "1")
os.environ.setdefault("HF_HUB_DISABLE_PROGRESS_BARS", "1")
os.environ.setdefault("HF_HUB_DISABLE_SYMLINKS_WARNING", "1")
os.environ.setdefault("TQDM_DISABLE", "1")
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")
os.environ.setdefault("TRANSFORMERS_NO_ADVISORY_WARNINGS", "1")
for _stream in (sys.stdout, sys.stderr):
    _reconfigure = getattr(_stream, "reconfigure", None)
    if callable(_reconfigure):
        _reconfigure(encoding="utf-8", errors="strict")

import pandas as pd
from openpyxl import load_workbook
from seo_knowledge import (
    forget_topic,
    knowledge_status,
    save_topic_profile,
    store_examples,
    topic_key,
)


PROJECT_DIR = Path(__file__).resolve().parent
JOBS_DIR = PROJECT_DIR / "jobs"
OUTPUT_DIR = PROJECT_DIR / "outputs"
# Keep light workflow commands independent from seo_pipeline.  The agent sends
# only these canonical English labels, but accepting common aliases preserves
# compatibility with previously labelled workbooks.
LABEL_ALIASES = {
    "commercial": "commercial",
    "transactional": "commercial",
    "\u043a\u043e\u043c\u043c\u0435\u0440\u0447\u0435\u0441\u043a\u0438\u0439": "commercial",
    "\u0438\u043d\u0444\u043e\u0440\u043c\u0430\u0446\u0438\u043e\u043d\u043d\u044b\u0439": "informational",
    "\u0438\u043d\u0444\u043e\u0440\u043c\u0430\u0446\u0438\u043e\u043d\u043d\u044b\u0435": "informational",
    "informational": "informational",
    "information": "informational",
    "\u043c\u0443\u0441\u043e\u0440": "garbage",
    "garbage": "garbage",
    "irrelevant": "garbage",
}
REQUIRED_RESULT_SHEETS = {
    "Commercial clusters",
    "Informational clusters",
    "Garbage",
    "Human review",
}
ALLOWED_ROOT_PYTHON_FILES = {
    "seo_io.py",
    "seo_embeddings.py",
    "seo_knowledge.py",
    "seo_pipeline.py",
    "seo_prepare.py",
    "seo_workflow.py",
}
VALID_LABELS = {"commercial", "informational", "garbage"}
MIN_REVIEWED_LABELS = 150
MIN_EXAMPLES_PER_CLASS = 10
MIN_LARGE_CORE_EXAMPLES_PER_INTENT = 30
MIN_SEEDS_PER_CLASS = 5
HIGH_PRIORITY_REVIEW_ROWS = 50
GARBAGE_SHARE_WARNING = 0.35
STANDARD_LABEL_TARGET = 300
HIGH_GARBAGE_LABEL_TARGET = 500
INITIAL_LABEL_BATCH_SIZE = 50
ACTIVE_REVIEW_BATCH_SIZE = 50
RELEVANCE_AUDIT_BATCH_SIZE = 50
MIN_GARBAGE_CALIBRATION_LABELS = 30
MIN_RELEVANCE_AUDIT_BATCHES = 3
MAX_RELEVANCE_AUDIT_BATCHES = 6
INTENT_AUDIT_BATCH_SIZE = 50
MIN_INTENT_AUDIT_BATCHES = 2
POLICY_CONFLICT_AUDIT_BATCH_SIZE = 50
MIN_POLICY_CONFLICT_AUDIT_BATCHES = 1
INTENT_FAMILY_REVIEW_BATCH_SIZE = 20
FAMILY_COVERAGE_REVIEW_BATCH_SIZE = 50
SMALL_CORE_DIRECT_LIMIT = 500
MEDIUM_CORE_LIMIT = 5_000
DEFAULT_LARGE_THRESHOLD = 100_000
# Hermes invokes commands through a Bash-compatible transport on Windows.  A
# direct call to the project interpreter avoids PowerShell's script-approval
# category while keeping the maintained workflow as the only Python entry
# point.  Forward slashes are accepted by both Git Bash and Windows Python.
HERMES_COMMAND = r'"./.venv/Scripts/python.exe" "./seo_workflow.py"'
RUN_LOCK_NAME = ".seo_run.lock.json"
MIN_PRODUCTIVE_MARKER_STEM_LENGTH = 4

INTENT_LABELING_CONTRACT = (
    "Use commercial for a query whose practical goal is to find, choose, access, apply to, contact, "
    "book, order, buy, rent, hire, or otherwise convert through a concrete offer or result. Commercial "
    "does not require a payment verb and includes both sides of a marketplace. Thus a bare product/service "
    "offer, a vacancy search or application, a provider search, and an employer hiring query are commercial. "
    "Use informational only when the goal is knowledge: explanation, duties, instructions, diagnostics, "
    "reference facts, requirements discussed as reference, reviews, comparison, or general education without "
    "seeking a concrete offer/result. Question words never decide the class by themselves. Use garbage only "
    "outside the stated topic/business scope. Classify the whole phrase by its user goal, not by one noun."
)


def resolve_input(value: str) -> Path:
    """Resolve an input without loading the ML pipeline."""
    candidate = Path(value)
    if candidate.is_file():
        return candidate.resolve()
    candidate = PROJECT_DIR / "files" / Path(value).name
    if candidate.is_file():
        return candidate.resolve()
    raise FileNotFoundError(f"Input not found: {value}. Expected it in {PROJECT_DIR / 'files'}.")


def state_file(job_dir: Path) -> Path:
    return job_dir / "state.json"


def read_state(job_dir: Path) -> dict[str, object]:
    path = state_file(job_dir)
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def write_state(job_dir: Path, stage: str, **updates: object) -> dict[str, object]:
    state = read_state(job_dir)
    state.update(updates)
    state["stage"] = stage
    state["updated_at"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    temporary = state_file(job_dir).with_suffix(".tmp")
    temporary.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(state_file(job_dir))
    return state


def ensure_workflow_job_id(job_dir: Path) -> str:
    """Return the immutable ASCII ID that links a result workbook to its job."""
    config_path = job_dir / "job_config.json"
    try:
        config = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Cannot read job_config.json for workflow identity: {exc}") from exc
    current = str(config.get("workflow_job_id", "")).strip()
    if not re.fullmatch(r"seo-[a-f0-9]{16}", current):
        current = f"seo-{uuid.uuid4().hex[:16]}"
        config["workflow_job_id"] = current
        temporary = config_path.with_suffix(".tmp")
        temporary.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(config_path)
    state = read_state(job_dir)
    if state.get("workflow_job_id") != current:
        write_state(job_dir, str(state.get("stage", "unknown")), workflow_job_id=current)
    return current


def find_job_by_workflow_id(workflow_job_id: str) -> Path:
    matches: list[Path] = []
    for candidate in JOBS_DIR.iterdir() if JOBS_DIR.is_dir() else []:
        config_path = candidate / "job_config.json"
        if not candidate.is_dir() or not config_path.is_file():
            continue
        try:
            config = json.loads(config_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if str(config.get("workflow_job_id", "")) == workflow_job_id:
            matches.append(candidate)
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise ValueError(f"No SEO job matches Workflow Job ID {workflow_job_id}.")
    raise ValueError(f"Workflow Job ID {workflow_job_id} is ambiguous; no import was made.")


def compact_status(status: dict[str, object]) -> dict[str, object]:
    quality = status.get("label_quality", {})
    state = status.get("state", {})
    return {
        "job_directory": status.get("job_directory"),
        "stage": state.get("stage", "unknown") if isinstance(state, dict) else "unknown",
        "labels_reviewed": quality.get("valid_labeled_rows", 0) if isinstance(quality, dict) else 0,
        "labeled_counts": status.get("labeled_counts", {}),
        "ready_for_supervised_run": status.get("ready_for_supervised_run", False),
        "blocking_errors": list(status.get("blocking_errors", []))[:3],
        "warnings": list(status.get("warnings", []))[:3],
    }


def next_action(job_dir: Path) -> dict[str, object]:
    """Return the sole safe workflow decision point for a Bionic agent."""
    status = job_status(job_dir)
    state = read_state(job_dir)
    stage = str(state.get("stage", "unknown"))
    lock_path = job_dir / RUN_LOCK_NAME
    active_lock = read_run_lock(lock_path) if lock_path.is_file() else {}
    if active_lock and lock_process_is_active(active_lock):
        # A run may outlive the terminal tool's foreground timeout.  Returning
        # a command here would make an agent start a duplicate process or burn
        # its context on polling an already-running job.
        return {
            "status": "running",
            "stage": "run",
            "pid": active_lock.get("pid"),
            "started_at": active_lock.get("started_at"),
            "instruction": "A workflow run is already active. Do not run, retry, wait, poll, or call next. Return immediately and rely on the terminal completion notification.",
        }
    if stage == "cluster_relevance_review":
        audit_status = cluster_relevance_status(job_dir)
        if int(audit_status["remaining_clusters"]) > 0:
            return {
                "status": "continue",
                "stage": stage,
                "action": "label_cluster_relevance_batch",
                "remaining_clusters": audit_status["remaining_clusters"],
                "remaining_decisions": audit_status["remaining_decisions"],
                "command": f'{HERMES_COMMAND} cluster-review --job "{job_dir.name}" --limit 30 --quiet',
                "after_cluster_labels": f'{HERMES_COMMAND} apply-cluster-labels-inline --job "{job_dir.name}" --labels "RC0001|relevant;RC0001-R|garbage;RC0001-B|relevant" --quiet',
            }
        return {
            "status": "continue",
            "stage": "cluster_relevance_finalize",
            "action": "apply_cluster_relevance_decisions",
            "command": f'{HERMES_COMMAND} apply-cluster-decisions --job "{job_dir.name}" --quiet',
        }
    inspection = status.get("inspection", {})
    quality = status.get("label_quality", {})
    reviewed = int(quality.get("valid_labeled_rows", 0)) if isinstance(quality, dict) else 0
    total = int(quality.get("sample_rows", 0)) if isinstance(quality, dict) else 0
    source_total = int(inspection.get("unique_normalized_phrases", total) or total)
    garbage_share = float(quality.get("garbage_share", 0.0)) if isinstance(quality, dict) else 0.0
    labeled_counts = status.get("labeled_counts", {})
    garbage_count = int(labeled_counts.get("garbage", 0)) if isinstance(labeled_counts, dict) else 0
    errors = list(status.get("blocking_errors", []))
    root_files = list(status.get("unexpected_python_files", []))
    if root_files:
        return {
            "status": "blocked",
            "stage": stage,
            "reason": "Unexpected root Python files block the workflow.",
            "files": root_files,
            "instruction": "Report this block to the user. Do not rename, run, or bypass these files.",
        }
    if stage == "quality_review":
        return {
            "status": "stop",
            "stage": stage,
            "reason": "A run already completed. Do not search for another output or rerun cleaning.",
            "last_run": state.get("last_run_path") or state.get("last_large_run"),
            "instruction": "Report the stored result path and request user review or a reviewed workbook for feedback.",
        }
    if stage in {"knowledge_saved", "learned"}:
        return {
            "status": "stop",
            "stage": stage,
            "reason": "Reviewed labels and topic knowledge were saved. Do not rerun cleaning unless the user explicitly requests recalculation.",
            "knowledge_topic_key": state.get("knowledge_topic_key"),
        }
    if stage == "blocked":
        return {
            "status": "blocked",
            "stage": stage,
            "reason": state.get("run_error") or state.get("blocked_reason") or "The workflow recorded a blocking failure.",
            "instruction": "Report this block once. Do not retry the same command.",
        }
    if stage == "finalized":
        return {
            "status": "stop",
            "stage": stage,
            "final": state.get("final"),
            "instruction": "Report the final path. No further processing is required.",
        }
    if not inspection:
        return {
            "status": "blocked",
            "stage": stage,
            "reason": "Job has no valid inspection.json.",
            "instruction": "Report the block; do not create a replacement script.",
        }
    initial_target = min(MIN_REVIEWED_LABELS, total)
    if reviewed < initial_target:
        return {
            "status": "continue",
            "stage": "initial_labeling",
            "action": "label_initial_batch",
            "labels_reviewed": reviewed,
            "target_labels": initial_target,
            "command": f'{HERMES_COMMAND} sample --job "{job_dir.name}" --offset 0 --limit {min(INITIAL_LABEL_BATCH_SIZE, initial_target - reviewed)} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    # A compact source is best classified directly by the language model.  It
    # is cheaper and more accurate to label every phrase than to extrapolate a
    # classifier from a sample and then audit that extrapolation.
    if source_total <= SMALL_CORE_DIRECT_LIMIT and reviewed < total:
        return {
            "status": "continue",
            "stage": "direct_labeling",
            "action": "label_all_small_core_rows",
            "labels_reviewed": reviewed,
            "target_labels": total,
            "command": f'{HERMES_COMMAND} sample --job "{job_dir.name}" --offset 0 --limit {min(INITIAL_LABEL_BATCH_SIZE, total - reviewed)} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    label_dependent_seed_prefixes = (
        "topic.relevant_seeds",
        "intent.commercial_seeds",
        "intent.informational_seeds",
    )
    non_label_errors = [
        error
        for error in errors
        if not error.startswith("Only ")
        and "Class " not in error
        and "priority rows" not in error
        and not error.startswith(label_dependent_seed_prefixes)
    ]
    if non_label_errors:
        return {
            "status": "continue",
            "stage": "configuration",
            "action": "configure_job",
            "job_config": str((job_dir / "job_config.json").resolve()),
            "blocking_errors": non_label_errors[:5],
            "instruction": "Update only the requested topic-specific fields in job_config.json, then call next again.",
        }
    if source_total <= SMALL_CORE_DIRECT_LIMIT:
        if not status.get("ready_for_supervised_run", False):
            return {
                "status": "blocked",
                "stage": stage,
                "reason": "The fully labeled small core still fails readiness checks.",
                "blocking_errors": errors[:5],
            }
        return {
            "status": "continue",
            "stage": "run",
            "action": "run_auto_direct_labels",
            "command": f'{HERMES_COMMAND} run-auto "{Path(str(inspection.get("input_file", ""))).name}" --job "{job_dir.name}" --quiet',
        }
    current_counts = current_sample_intent_counts(load_label_sheet(job_dir))
    underrepresented_intents = [
        label
        for label in ("commercial", "informational")
        if current_counts[label] < MIN_LARGE_CORE_EXAMPLES_PER_INTENT
    ]
    if underrepresented_intents and reviewed < total:
        return {
            "status": "continue",
            "stage": "intent_calibration",
            "action": "label_missing_intent_class",
            "labels_reviewed": reviewed,
            "current_intent_counts": current_counts,
            "required_per_intent": MIN_LARGE_CORE_EXAMPLES_PER_INTENT,
            "missing_or_weak_intents": underrepresented_intents,
            "labeling_contract": INTENT_LABELING_CONTRACT,
            "command": f'{HERMES_COMMAND} sample --job "{job_dir.name}" --offset 0 --limit {min(INITIAL_LABEL_BATCH_SIZE, total - reviewed)} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    if underrepresented_intents:
        return {
            "status": "blocked",
            "stage": "intent_calibration",
            "reason": "The representative sample is exhausted but a safe two-intent classifier cannot be trained.",
            "current_intent_counts": current_counts,
            "required_per_intent": MIN_LARGE_CORE_EXAMPLES_PER_INTENT,
            "missing_or_weak_intents": underrepresented_intents,
            "instruction": "Report the class-calibration failure. Do not run or invent labels; the topic scope or labeling taxonomy needs correction.",
        }
    priority_unreviewed = int(quality.get("high_priority_unreviewed", 0)) if isinstance(quality, dict) else 0
    if priority_unreviewed:
        return {
            "status": "continue",
            "stage": "priority_review",
            "action": "label_priority_batch",
            "labels_reviewed": reviewed,
            "remaining_priority_rows": priority_unreviewed,
            "command": f'{HERMES_COMMAND} priority-review --job "{job_dir.name}" --limit {min(ACTIVE_REVIEW_BATCH_SIZE, priority_unreviewed)} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    family_status = intent_family_status(job_dir)
    if int(family_status["remaining"]) > 0:
        return {
            "status": "continue",
            "stage": "intent_family_audit",
            "action": "label_intent_family_batch",
            "families_labeled": family_status["labeled"],
            "families_remaining": family_status["remaining"],
            "command": f'{HERMES_COMMAND} family-review --job "{job_dir.name}" --limit {INTENT_FAMILY_REVIEW_BATCH_SIZE} --quiet',
            "after_family_labels": f'{HERMES_COMMAND} apply-family-labels-inline --job "{job_dir.name}" --labels "IF0001|informational;IF0002|commercial;IF0003|neutral" --quiet',
        }
    family_coverage = family_coverage_status(load_label_sheet(job_dir))
    if int(family_coverage["remaining"]) > 0:
        return {
            "status": "continue",
            "stage": "intent_family_coverage",
            "action": "label_unrepresented_family_examples",
            "coverage_rows_labeled": family_coverage["labeled"],
            "coverage_rows_remaining": family_coverage["remaining"],
            "command": f'{HERMES_COMMAND} family-coverage-review --job "{job_dir.name}" --limit {min(FAMILY_COVERAGE_REVIEW_BATCH_SIZE, int(family_coverage["remaining"]))} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    config_path = job_dir / "job_config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    policy = config.get("intent_policy", {}) if isinstance(config, dict) else {}
    relevance_policy = config.get("relevance", {}) if isinstance(config, dict) else {}
    commercial_prototypes = policy.get("commercial_prototypes", []) if isinstance(policy, dict) else []
    implicit_commercial_prototypes = policy.get("implicit_commercial_prototypes", []) if isinstance(policy, dict) else []
    informational_prototypes = policy.get("informational_prototypes", []) if isinstance(policy, dict) else []
    informational_signals = configured_values(config, "intent", "informational_markers")
    commercial_signals = configured_values(config, "intent", "commercial_markers")
    weak_question_signals = configured_values(config, "intent", "weak_question_markers")
    relevant_prototypes = relevance_policy.get("relevant_prototypes", []) if isinstance(relevance_policy, dict) else []
    garbage_prototypes = relevance_policy.get("garbage_prototypes", []) if isinstance(relevance_policy, dict) else []
    if (
        len(commercial_prototypes) < 3
        or len(implicit_commercial_prototypes) < 5
        or len(informational_prototypes) < 5
        or len(informational_signals) < 12
        or len(commercial_signals) < 5
        or len(weak_question_signals) < 3
        or len(relevant_prototypes) < 5
        or len(garbage_prototypes) < 5
    ):
        return {
            "status": "continue",
            "stage": "intent_policy",
            "action": "define_intent_policy",
            "labels_reviewed": reviewed,
            "command": f'{HERMES_COMMAND} policy-context --job "{job_dir.name}" --limit 8 --quiet',
            "after_policy": f'{HERMES_COMMAND} apply-policy-inline --job "{job_dir.name}" --commercial "explicit commercial prototype 1; explicit commercial prototype 2; explicit commercial prototype 3" --implicit-commercial "bare transactional structure 1; bare transactional structure 2; bare transactional structure 3; bare transactional structure 4; bare transactional structure 5" --informational "prototype 1; prototype 2; prototype 3; prototype 4; prototype 5" --commercial-signals "strong commercial signal 1; strong commercial signal 2; strong commercial signal 3; strong commercial signal 4; strong commercial signal 5" --informational-signals "strong information signal 1; strong information signal 2; strong information signal 3; strong information signal 4; strong information signal 5; strong information signal 6; strong information signal 7; strong information signal 8; strong information signal 9; strong information signal 10; strong information signal 11; strong information signal 12" --weak-question-signals "weak question 1; weak question 2; weak question 3" --relevant "relevant prototype 1; relevant prototype 2; relevant prototype 3; relevant prototype 4; relevant prototype 5" --garbage "hard negative 1; hard negative 2; hard negative 3; hard negative 4; hard negative 5" --quiet',
        }
    if bool(state.get("signal_policy_refinement_needed")) and int(state.get("signal_policy_attempts", 0) or 0) < 2:
        return {
            "status": "continue",
            "stage": "intent_policy",
            "action": "refine_intent_signals",
            "signal_coverage": state.get("signal_policy_coverage", {}),
            "command": f'{HERMES_COMMAND} policy-context --job "{job_dir.name}" --limit 8 --quiet',
            "after_policy": f'{HERMES_COMMAND} apply-policy-inline --job "{job_dir.name}" --commercial "explicit commercial prototypes" --implicit-commercial "implicit commercial structures" --informational "informational prototypes" --commercial-signals "8-20 strong commercial signals" --informational-signals "12-30 strong informational signals" --weak-question-signals "3-10 weak question signals" --relevant "relevant prototypes" --garbage "hard negative prototypes" --quiet',
        }
    audit_limit = (
        POLICY_CONFLICT_AUDIT_BATCH_SIZE
        if source_total > MEDIUM_CORE_LIMIT
        else min(POLICY_CONFLICT_AUDIT_BATCH_SIZE, max(10, source_total // 20))
    )
    relevance_audit_batches = int(state.get("relevance_audit_batches", 0) or 0)
    relevance_yields = [
        int(value) for value in (state.get("relevance_audit_garbage_yields", []) or [])
    ]
    relevance_plateau = (
        relevance_audit_batches >= MIN_RELEVANCE_AUDIT_BATCHES
        and len(relevance_yields) >= 2
        and max(relevance_yields[-2:]) <= 1
    )
    needs_relevance_audit = (
        relevance_audit_batches < MIN_RELEVANCE_AUDIT_BATCHES
        or (
            garbage_count < MIN_GARBAGE_CALIBRATION_LABELS
            and relevance_audit_batches < MAX_RELEVANCE_AUDIT_BATCHES
            and not relevance_plateau
        )
    )
    if needs_relevance_audit and reviewed < total:
        return {
            "status": "continue",
            "stage": "relevance_audit",
            "action": "label_out_of_topic_candidates",
            "labels_reviewed": reviewed,
            "garbage_labels": garbage_count,
            "recent_garbage_yields": relevance_yields[-2:],
            "audit_batch": relevance_audit_batches + 1,
            "command": f'{HERMES_COMMAND} relevance-review --job "{job_dir.name}" --limit {audit_limit} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    intent_audit_batches = int(state.get("intent_audit_batches", 0) or 0)
    if intent_audit_batches < MIN_INTENT_AUDIT_BATCHES and reviewed < total:
        return {
            "status": "continue",
            "stage": "intent_audit",
            "action": "label_commercial_informational_boundary",
            "labels_reviewed": reviewed,
            "audit_batch": intent_audit_batches + 1,
            "command": f'{HERMES_COMMAND} intent-review --job "{job_dir.name}" --limit {audit_limit} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    policy_conflict_batches = int(state.get("policy_conflict_audit_batches", 0) or 0)
    if policy_conflict_batches < MIN_POLICY_CONFLICT_AUDIT_BATCHES and reviewed < total:
        return {
            "status": "continue",
            "stage": "policy_conflict_audit",
            "action": "label_signal_classifier_conflicts",
            "labels_reviewed": reviewed,
            "audit_batch": policy_conflict_batches + 1,
            "command": f'{HERMES_COMMAND} policy-conflict-review --job "{job_dir.name}" --limit {audit_limit} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    normal_target = min(STANDARD_LABEL_TARGET, total)
    adaptive_target = (
        min(HIGH_GARBAGE_LABEL_TARGET, total)
        if garbage_share > 0.15 or garbage_count >= MIN_GARBAGE_CALIBRATION_LABELS
        else normal_target
    )
    if reviewed < adaptive_target:
        return {
            "status": "continue",
            "stage": "active_review",
            "action": (
                "label_high_garbage_boundary_batch"
                if adaptive_target > normal_target
                else "label_uncertain_batch"
            ),
            "labels_reviewed": reviewed,
            "target_labels": adaptive_target,
            "command": f'{HERMES_COMMAND} review --job "{job_dir.name}" --limit {min(ACTIVE_REVIEW_BATCH_SIZE, adaptive_target - reviewed)} --quiet',
            "after_labels": f'{HERMES_COMMAND} apply-labels-inline --job "{job_dir.name}" --labels "SampleID|label|confidence;..." --quiet',
        }
    if not status.get("ready_for_supervised_run", False):
        return {
            "status": "blocked",
            "stage": stage,
            "reason": "Job readiness checks still fail.",
            "blocking_errors": errors[:5],
        }
    # The first topic policy is intentionally available before the audit loops,
    # but those loops may add hundreds of more representative labels.  Give the
    # model one bounded final refresh after the target is reached.  The apply
    # command compares the candidate with the existing policy on all reviewed
    # examples and retains the old policy if coverage or false positives regress.
    final_policy_checked = int(state.get("final_policy_checked_at_labels", 0) or 0)
    if (
        source_total > SMALL_CORE_DIRECT_LIMIT
        and int(state.get("signal_policy_attempts", 0) or 0) > 0
        and final_policy_checked < reviewed
    ):
        if not bool(state.get("final_policy_refinement_requested")):
            write_state(
                job_dir,
                "intent_policy",
                final_policy_refinement_requested=True,
                final_policy_refinement_labels=reviewed,
            )
        return {
            "status": "continue",
            "stage": "intent_policy",
            "action": "refine_final_intent_policy",
            "labels_reviewed": reviewed,
            "signal_coverage": intent_signal_coverage(
                load_label_sheet(job_dir),
                commercial_signals,
                informational_signals,
                weak_question_signals,
            ),
            "command": f'{HERMES_COMMAND} policy-context --job "{job_dir.name}" --limit 8 --quiet',
            "after_policy": f'{HERMES_COMMAND} apply-policy-inline --job "{job_dir.name}" --commercial "explicit commercial prototypes" --implicit-commercial "implicit commercial structures" --informational "informational prototypes" --commercial-signals "8-20 strong commercial signals" --informational-signals "12-30 strong informational signals" --weak-question-signals "3-10 weak question signals" --relevant "relevant prototypes" --garbage "hard negative prototypes" --quiet',
        }
    return {
        "status": "continue",
        "stage": "run",
        "action": "run_auto",
        "command": f'{HERMES_COMMAND} run-auto "{Path(str(inspection.get("input_file", ""))).name}" --job "{job_dir.name}" --quiet',
    }


def recorded_next_action(job_dir: Path) -> dict[str, object]:
    """Persist the terminal state of `next` for Hermes' SEO autopilot guard."""
    action = next_action(job_dir)
    status = str(action.get("status", "blocked"))
    state = read_state(job_dir)
    persisted_stage = "run" if status == "running" else str(state.get("stage", action.get("stage", "unknown")))
    write_state(
        job_dir,
        persisted_stage,
        autopilot_last_status=status,
    )
    return action


def print_result(value: dict[str, object], quiet: bool = False) -> None:
    print(json.dumps(value, ensure_ascii=False, indent=None if quiet else 2))


def quiet_library_logs() -> None:
    for name in ("huggingface_hub", "transformers", "sentence_transformers", "urllib3"):
        logging.getLogger(name).setLevel(logging.ERROR)


def resolve_job(value: str) -> Path:
    candidate = Path(value)
    if candidate.is_dir():
        return candidate.resolve()
    candidate = JOBS_DIR / value
    if candidate.is_dir():
        return candidate.resolve()
    raise FileNotFoundError(f"Job directory not found: {value}")


def find_job_for_input(input_value: str) -> Path | None:
    wanted = Path(input_value).name.lower()
    matches: list[tuple[float, Path]] = []
    for candidate in (JOBS_DIR.iterdir() if JOBS_DIR.is_dir() else []):
        inspection = candidate / "inspection.json"
        if not candidate.is_dir() or not inspection.is_file():
            continue
        try:
            data = json.loads(inspection.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if Path(str(data.get("input_file", ""))).name.lower() == wanted:
            matches.append((inspection.stat().st_mtime, candidate))
    return max(matches, default=(0.0, None), key=lambda item: item[0])[1]


def normalized_column_lookup(frame: pd.DataFrame) -> dict[str, str]:
    return {
        re.sub(r"\s+", " ", str(column).strip().lower()): column
        for column in frame.columns
    }


def configured_values(config: dict[str, object], section: str, key: str) -> list[str]:
    raw_section = config.get(section, {})
    if not isinstance(raw_section, dict):
        return []
    values = raw_section.get(key, [])
    if not isinstance(values, list):
        return []
    return [str(value).strip() for value in values if str(value).strip()]


def normalize_signal(value: object) -> str:
    text = re.sub(r"[^a-zа-я0-9*]+", " ", str(value).lower().replace("ё", "е"))
    return re.sub(r"\s+", " ", text).strip()


def marker_safety_reason(marker: str) -> str | None:
    """Reject only structurally unsafe signals; never encode topic vocabulary."""
    normalized = normalize_signal(marker)
    if not normalized:
        return "empty"
    for token in normalized.split():
        if "*" in token and len(token.replace("*", "")) < MIN_PRODUCTIVE_MARKER_STEM_LENGTH:
            return "wildcard_stem_too_short"
    return None


def signal_mask(phrases: pd.Series, signals: list[str]) -> pd.Series:
    normalized = [normalize_signal(value) for value in signals if normalize_signal(value)]
    if not normalized:
        return pd.Series(False, index=phrases.index)
    parts = [re.escape(value).replace(r"\*", r"[a-zа-я0-9]*") for value in normalized]
    expression = rf"(?:^|\s)(?:{'|'.join(parts)})(?:$|\s)"
    normalized_phrases = (
        phrases.fillna("").astype(str).str.lower().str.replace("ё", "е", regex=False)
    )
    return normalized_phrases.str.contains(expression, case=False, regex=True, na=False)


def family_overlaps_signals(pattern: str, signals: list[str]) -> bool:
    """Return true when a one-token family is already defined as a weak signal."""
    family_stem = normalize_signal(pattern).rstrip("*")
    if not family_stem:
        return False
    for signal in signals:
        for token in normalize_signal(signal).split():
            signal_stem = token.rstrip("*")
            if signal_stem and (
                family_stem.startswith(signal_stem)
                or signal_stem.startswith(family_stem)
            ):
                return True
    return False


def intent_family_paths(job_dir: Path) -> tuple[Path, Path]:
    return (
        job_dir / "intent_family_candidates.json",
        job_dir / "intent_family_labels.json",
    )


def intent_family_status(job_dir: Path) -> dict[str, object]:
    candidates_path, labels_path = intent_family_paths(job_dir)
    if not candidates_path.is_file():
        return {
            "total": 0,
            "labeled": 0,
            "remaining": 0,
            "pending": [],
            "labels": {},
        }
    payload = json.loads(candidates_path.read_text(encoding="utf-8"))
    families = payload.get("families", []) if isinstance(payload, dict) else []
    labels: dict[str, str] = {}
    if labels_path.is_file():
        loaded = json.loads(labels_path.read_text(encoding="utf-8"))
        if isinstance(loaded, dict) and isinstance(loaded.get("labels"), dict):
            labels = {
                str(key): str(value).strip().lower()
                for key, value in loaded["labels"].items()
            }
    pending = [
        record
        for record in families
        if isinstance(record, dict) and str(record.get("id", "")) not in labels
    ]
    return {
        "total": len(families),
        "labeled": len(families) - len(pending),
        "remaining": len(pending),
        "pending": pending,
        "labels": labels,
    }


def intent_family_batch(job_dir: Path, limit: int) -> dict[str, object]:
    if limit < 1 or limit > INTENT_FAMILY_REVIEW_BATCH_SIZE:
        raise ValueError(
            f"limit must be from 1 to {INTENT_FAMILY_REVIEW_BATCH_SIZE}."
        )
    status = intent_family_status(job_dir)
    rows = [
        {
            "id": str(record.get("id", "")),
            "family": str(record.get("pattern", "")),
            "kind": str(record.get("kind", "lexical")),
            "occurrences": int(record.get("occurrences", 0) or 0),
            "share": float(record.get("share", 0.0) or 0.0),
            "examples": [str(value) for value in record.get("examples", [])][:8],
        }
        for record in list(status["pending"])[:limit]
    ]
    return {
        "stage": "intent_family_audit",
        "topic": read_state(job_dir).get("topic", ""),
        "allowed_labels": ["commercial", "informational", "neutral"],
        "labeling_contract": INTENT_LABELING_CONTRACT,
        "instruction": (
            "Judge whether the lexical family or two-token structural family independently determines intent "
            "across the shown examples. "
            "Use commercial for an inherently offer/result-seeking or conversion family, including implicit "
            "marketplace demand without a payment verb; use informational for an "
            "inherently informational family such as reviews, opinions, experience, pros/cons, instructions, "
            "diagnostics, reference facts, specifications, diagrams, comparisons, or explanations, and neutral "
            "for brands, products, professions, topic nouns, weak question words, or mixed families. A neutral "
            "decision creates "
            "no classification override. The examples intentionally cover different contexts: do not infer a "
            "decisive label from only one subset. Judge the family for the user's requested binary output, not a "
            "generic marketing-funnel taxonomy. A listing/search family for products, services, vacancies, "
            "providers, rentals, admissions, or similar concrete offers is commercial; an explanation of the "
            "same entity is informational."
        ),
        "rows": rows,
        "remaining_after_batch": int(status["remaining"]) - len(rows),
    }


def apply_intent_family_labels_inline(job_dir: Path, value: str) -> dict[str, object]:
    candidates_path, labels_path = intent_family_paths(job_dir)
    if not candidates_path.is_file():
        raise FileNotFoundError("intent_family_candidates.json is missing.")
    payload = json.loads(candidates_path.read_text(encoding="utf-8"))
    families = payload.get("families", []) if isinstance(payload, dict) else []
    family_by_id = {
        str(record.get("id")): record
        for record in families
        if isinstance(record, dict) and record.get("id")
    }
    existing: dict[str, str] = {}
    if labels_path.is_file():
        loaded = json.loads(labels_path.read_text(encoding="utf-8"))
        if isinstance(loaded, dict) and isinstance(loaded.get("labels"), dict):
            existing = {
                str(key): str(label).strip().lower()
                for key, label in loaded["labels"].items()
            }
    applied = 0
    for item in value.split(";"):
        if not item.strip():
            continue
        fields = [field.strip() for field in item.split("|")]
        if len(fields) != 2:
            raise ValueError("Each family decision must be ID|label.")
        family_id, label = fields[0], fields[1].lower()
        if family_id not in family_by_id:
            raise ValueError(f"Unknown intent family ID: {family_id}")
        if label not in {"commercial", "informational", "neutral"}:
            raise ValueError(f"Invalid intent family label: {label}")
        existing[family_id] = label
        applied += 1
    if not applied:
        raise ValueError("At least one intent family decision is required.")
    config_path = job_dir / "job_config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    weak_signals = configured_values(config, "intent", "weak_question_markers")
    forced_neutral = 0
    forced_neutral_weak = 0
    forced_neutral_lexical = 0
    for family_id, label in list(existing.items()):
        record = family_by_id.get(family_id)
        pattern = str(record.get("pattern", "")).strip() if record else ""
        unsafe_lexical_decision = bool(
            record
            and str(record.get("kind", "lexical")) == "lexical"
            and "safe_decisive_lexical" in record
            and not bool(record.get("safe_decisive_lexical"))
        )
        weak_overlap = family_overlaps_signals(pattern, weak_signals)
        if label != "neutral" and (weak_overlap or unsafe_lexical_decision):
            existing[family_id] = "neutral"
            forced_neutral += 1
            forced_neutral_weak += int(weak_overlap)
            forced_neutral_lexical += int(unsafe_lexical_decision)

    temporary = labels_path.with_suffix(".tmp")
    temporary.write_text(
        json.dumps({"labels": existing}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(labels_path)

    rules = {"commercial": [], "informational": [], "neutral": []}
    for family_id, label in existing.items():
        record = family_by_id.get(family_id)
        pattern = str(record.get("pattern", "")).strip() if record else ""
        if pattern and label in rules and pattern not in rules[label]:
            rules[label].append(pattern)
    intent = config.setdefault("intent", {})
    intent["family_rules"] = rules
    config_temporary = config_path.with_suffix(".tmp")
    config_temporary.write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    config_temporary.replace(config_path)
    status = intent_family_status(job_dir)
    write_state(
        job_dir,
        "intent_family_audit",
        intent_family_labeled=status["labeled"],
        intent_family_remaining=status["remaining"],
    )
    return {
        "status": "intent_family_labels_applied",
        "applied": applied,
        "labeled": status["labeled"],
        "remaining": status["remaining"],
        "decisive_commercial": len(rules["commercial"]),
        "decisive_informational": len(rules["informational"]),
        "neutral": len(rules["neutral"]),
        "forced_neutral_families": forced_neutral,
        "forced_neutral_weak_families": forced_neutral_weak,
        "forced_neutral_unsafe_lexical_families": forced_neutral_lexical,
    }


def intent_signal_coverage(
    frame: pd.DataFrame,
    commercial_signals: list[str],
    informational_signals: list[str],
    weak_question_signals: list[str] | None = None,
) -> dict[str, object]:
    labels, phrases = label_summary(frame)
    commercial_rows = labels.eq("commercial") & phrases.ne("")
    informational_rows = labels.eq("informational") & phrases.ne("")
    commercial_hit = signal_mask(phrases, commercial_signals)
    informational_hit = signal_mask(phrases, informational_signals)
    weak_question_hit = signal_mask(phrases, weak_question_signals or [])
    commercial_count = int(commercial_rows.sum())
    informational_count = int(informational_rows.sum())
    eligible = (
        commercial_count >= MIN_LARGE_CORE_EXAMPLES_PER_INTENT
        and informational_count >= MIN_LARGE_CORE_EXAMPLES_PER_INTENT
    )
    commercial_coverage = (
        float(commercial_hit[commercial_rows].mean()) if commercial_count else 0.0
    )
    informational_coverage = (
        float(informational_hit[informational_rows].mean()) if informational_count else 0.0
    )
    informational_false_positive = (
        float((informational_hit & ~commercial_hit)[commercial_rows].mean())
        if commercial_count
        else 0.0
    )
    commercial_false_positive = (
        float((commercial_hit & ~informational_hit)[informational_rows].mean())
        if informational_count
        else 0.0
    )
    weak_informational_rows = weak_question_hit & informational_rows
    weak_commercial_rows = weak_question_hit & commercial_rows
    weak_informational_coverage = (
        float(informational_hit[weak_informational_rows].mean())
        if int(weak_informational_rows.sum())
        else 0.0
    )
    weak_commercial_coverage = (
        float(commercial_hit[weak_commercial_rows].mean())
        if int(weak_commercial_rows.sum())
        else 0.0
    )
    needs_refinement = bool(
        eligible
        and (
            informational_coverage < 0.55
            or informational_false_positive > 0.15
            or commercial_false_positive > 0.15
            or (
                int(weak_informational_rows.sum()) >= 5
                and weak_informational_coverage < 0.50
            )
            or (
                int(weak_commercial_rows.sum()) >= 5
                and weak_commercial_coverage < 0.50
            )
        )
    )
    quality_score = (
        0.5 * (commercial_coverage + informational_coverage)
        - 0.75 * (informational_false_positive + commercial_false_positive)
    )
    return {
        "eligible": eligible,
        "commercial_examples": commercial_count,
        "informational_examples": informational_count,
        "commercial_coverage": round(commercial_coverage, 4),
        "informational_coverage": round(informational_coverage, 4),
        "informational_false_positive": round(informational_false_positive, 4),
        "commercial_false_positive": round(commercial_false_positive, 4),
        "weak_question_informational_examples": int(weak_informational_rows.sum()),
        "weak_question_commercial_examples": int(weak_commercial_rows.sum()),
        "weak_question_informational_context_coverage": round(
            weak_informational_coverage, 4
        ),
        "weak_question_commercial_context_coverage": round(
            weak_commercial_coverage, 4
        ),
        "quality_score": round(quality_score, 4),
        "needs_refinement": needs_refinement,
    }


def final_policy_candidate_is_safe(
    baseline: dict[str, object], candidate: dict[str, object]
) -> tuple[bool, list[str]]:
    """Reject a final signal refresh that regresses reviewed real examples."""
    reasons: list[str] = []
    if not bool(candidate.get("eligible")):
        reasons.append("not_enough_reviewed_intent_examples")
        return False, reasons
    baseline_score = float(baseline.get("quality_score", 0.0) or 0.0)
    candidate_score = float(candidate.get("quality_score", 0.0) or 0.0)
    if candidate_score + 0.02 < baseline_score:
        reasons.append("signal_quality_score_regressed")
    for key in ("informational_false_positive", "commercial_false_positive"):
        old = float(baseline.get(key, 0.0) or 0.0)
        new = float(candidate.get(key, 0.0) or 0.0)
        if new > max(0.20, old + 0.03):
            reasons.append(f"{key}_regressed")
    return not reasons, reasons


def label_quality_status(
    labels: pd.DataFrame,
    config: dict[str, object],
) -> tuple[dict[str, object], list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    lookup = normalized_column_lookup(labels)
    required_columns = {"phrase", "model label", "model confidence"}
    missing_columns = sorted(required_columns - set(lookup))
    if missing_columns:
        errors.append(
            "model_labels.xlsx is missing required columns: "
            + ", ".join(missing_columns)
        )
        return {}, errors, warnings

    phrase_column = lookup["phrase"]
    label_column = lookup["model label"]
    confidence_column = lookup["model confidence"]
    notes_column = lookup.get("model notes")

    phrases = labels[phrase_column].fillna("").astype(str).str.strip()
    raw_labels = labels[label_column].fillna("").astype(str).str.strip().str.lower()
    normalized_labels = raw_labels.map(LABEL_ALIASES)
    labeled_mask = normalized_labels.notna() & phrases.ne("")
    invalid_mask = raw_labels.ne("") & normalized_labels.isna()
    confidence = pd.to_numeric(labels[confidence_column], errors="coerce")
    valid_confidence = confidence.between(0.0, 1.0, inclusive="both")
    confidence_complete = labeled_mask & valid_confidence

    labeled_count = int(labeled_mask.sum())
    sample_rows = int(phrases.ne("").sum())
    required_labeled = min(MIN_REVIEWED_LABELS, sample_rows)
    invalid_count = int(invalid_mask.sum())
    missing_or_invalid_confidence = int((labeled_mask & ~valid_confidence).sum())
    counts = {
        label: int((normalized_labels == label).sum())
        for label in sorted(VALID_LABELS)
    }

    if labeled_count < required_labeled:
        errors.append(
            f"Only {labeled_count} valid labels; at least {required_labeled} are required."
        )
    if invalid_count:
        errors.append(
            f"{invalid_count} labels are invalid; use only commercial, informational, or garbage."
        )
    if missing_or_invalid_confidence:
        errors.append(
            f"{missing_or_invalid_confidence} labeled rows have missing or invalid Model Confidence; "
            "use a numeric value from 0 to 1."
        )
    for label in ("commercial", "informational"):
        if counts[label] < MIN_EXAMPLES_PER_CLASS:
            errors.append(
                f"Class {label} has {counts[label]} examples; at least "
                f"{MIN_EXAMPLES_PER_CLASS} model-reviewed examples are required."
            )
    if counts["garbage"] and counts["garbage"] < MIN_EXAMPLES_PER_CLASS:
        warnings.append(
            f"Class garbage has only {counts['garbage']} real examples. It is optional; "
            "do not invent labels merely to meet a quota."
        )

    # These are quality warnings, not artificial class quotas.  The workflow
    # must allow a genuinely one-sided core, while making a weak training set
    # visible before its labels are propagated to every phrase.
    intent_coverage_target = min(
        30,
        max(MIN_EXAMPLES_PER_CLASS, (labeled_count + 9) // 10),
    )
    weak_intent_classes = [
        label
        for label in ("commercial", "informational")
        if counts[label] < intent_coverage_target
    ]
    if weak_intent_classes:
        warnings.append(
            "Weak intent coverage for "
            + ", ".join(weak_intent_classes)
            + f": fewer than {intent_coverage_target} labels. Review more boundary examples."
        )

    phrase_keys = phrases.str.lower().str.replace("ё", "е", regex=False).str.replace(r"\s+", " ", regex=True)
    duplicate_labels = pd.DataFrame({"phrase": phrase_keys, "label": normalized_labels})
    duplicate_labels = duplicate_labels[duplicate_labels["phrase"].ne("") & duplicate_labels["label"].notna()]
    conflicting_duplicate_labels = int(
        (duplicate_labels.groupby("phrase")["label"].nunique() > 1).sum()
    )
    if conflicting_duplicate_labels:
        warnings.append(
            f"{conflicting_duplicate_labels} duplicate phrases have conflicting labels. "
            "Check them before saving knowledge."
        )

    reviewed_confidence = confidence[confidence_complete]
    low_confidence_share = (
        float(reviewed_confidence.lt(0.70).mean()) if not reviewed_confidence.empty else 0.0
    )
    if len(reviewed_confidence) >= 30 and low_confidence_share > 0.30:
        warnings.append(
            f"{low_confidence_share:.1%} of reviewed labels have confidence below 0.70. "
            "The topic boundary may need more representative examples."
        )

    priority_columns = [
        column
        for column in (lookup.get("search volume"), lookup.get("occurrences"))
        if column is not None
    ]
    high_priority_unreviewed = 0
    if priority_columns:
        priority = pd.Series(0.0, index=labels.index)
        for column in priority_columns:
            priority += pd.to_numeric(labels[column], errors="coerce").fillna(0)
        high_priority_index = priority.nlargest(
            min(HIGH_PRIORITY_REVIEW_ROWS, len(labels))
        ).index
        high_priority_unreviewed = int(
            (~(labeled_mask & valid_confidence).loc[high_priority_index]).sum()
        )
        if high_priority_unreviewed:
            errors.append(
                f"{high_priority_unreviewed} of the top "
                f"{min(HIGH_PRIORITY_REVIEW_ROWS, len(labels))} priority rows "
                "lack a valid label or confidence."
            )

    garbage_share = counts["garbage"] / max(labeled_count, 1)
    if garbage_share > GARBAGE_SHARE_WARNING:
        warnings.append(
            f"Garbage share in reviewed labels is {garbage_share:.1%}. "
            "A dirty semantic core is allowed, but manually audit false garbage before saving knowledge."
        )

    positive_markers = configured_values(config, "topic", "positive_markers")
    positive_markers += configured_values(config, "topic", "relevant_seeds")
    normalized_markers = {
        re.sub(r"\s+", " ", marker.lower().replace("ё", "е")).strip()
        for marker in positive_markers
        if marker.strip()
    }
    garbage_phrases = (
        phrases[normalized_labels.eq("garbage")]
        .str.lower()
        .str.replace("ё", "е", regex=False)
    )
    conflicting_garbage: list[str] = []
    for phrase in garbage_phrases:
        if any(marker in phrase for marker in normalized_markers):
            conflicting_garbage.append(phrase)
        if len(conflicting_garbage) >= 10:
            break
    if conflicting_garbage:
        errors.append(
            f"{len(conflicting_garbage)} sampled garbage labels conflict with configured "
            "positive markers or relevant seeds. Examples: "
            + " | ".join(conflicting_garbage[:5])
        )

    low_confidence_without_notes = 0
    if notes_column:
        notes = labels[notes_column].fillna("").astype(str).str.strip()
        low_confidence_without_notes = int(
            (labeled_mask & valid_confidence & confidence.lt(0.7) & notes.eq("")).sum()
        )
        if low_confidence_without_notes:
            warnings.append(
                f"{low_confidence_without_notes} low-confidence labels have no Model Notes."
            )

    metrics = {
        "sample_rows": sample_rows,
        "valid_labeled_rows": labeled_count,
        "required_labeled_rows": required_labeled,
        "labeled_counts": counts,
        "invalid_labels": invalid_count,
        "confidence_complete_rows": int(confidence_complete.sum()),
        "missing_or_invalid_confidence": missing_or_invalid_confidence,
        "high_priority_unreviewed": high_priority_unreviewed,
        "garbage_share": round(garbage_share, 6),
        "intent_coverage_target": intent_coverage_target,
        "weak_intent_classes": weak_intent_classes,
        "conflicting_duplicate_labels": conflicting_duplicate_labels,
        "low_confidence_share": round(low_confidence_share, 6),
        "positive_marker_garbage_conflicts": len(conflicting_garbage),
        "low_confidence_without_notes": low_confidence_without_notes,
    }
    return metrics, errors, warnings


def job_status(job_dir: Path) -> dict[str, object]:
    errors: list[str] = []
    warnings: list[str] = []
    inspection_path = job_dir / "inspection.json"
    inspection = (
        json.loads(inspection_path.read_text(encoding="utf-8"))
        if inspection_path.is_file()
        else {}
    )
    if not inspection:
        errors.append("inspection.json is missing or empty.")

    config_path = job_dir / "job_config.json"
    config: dict[str, object] = {}
    if config_path.is_file():
        try:
            loaded_config = json.loads(config_path.read_text(encoding="utf-8"))
            if isinstance(loaded_config, dict):
                config = loaded_config
            else:
                errors.append("job_config.json must contain a JSON object.")
        except (OSError, json.JSONDecodeError) as exc:
            errors.append(f"job_config.json cannot be read: {exc}")
    else:
        errors.append("job_config.json is missing.")

    seed_requirements = (
        ("topic", "relevant_seeds"),
        ("intent", "commercial_seeds"),
        ("intent", "informational_seeds"),
    )
    seed_counts: dict[str, int] = {}
    for section, key in seed_requirements:
        count = len(configured_values(config, section, key))
        seed_counts[f"{section}.{key}"] = count
        if count < MIN_SEEDS_PER_CLASS:
            errors.append(
                f"{section}.{key} has {count} values; at least "
                f"{MIN_SEEDS_PER_CLASS} diverse seeds are required."
            )

    garbage_seed_count = len(configured_values(config, "topic", "garbage_seeds"))
    seed_counts["topic.garbage_seeds"] = garbage_seed_count

    labels_path = job_dir / "model_labels.xlsx"
    label_metrics: dict[str, object] = {}
    if labels_path.is_file():
        try:
            labels = pd.read_excel(labels_path, sheet_name="Model labels")
            label_metrics, label_errors, label_warnings = label_quality_status(
                labels, config
            )
            errors.extend(label_errors)
            warnings.extend(label_warnings)
        except (OSError, ValueError, KeyError) as exc:
            errors.append(f"model_labels.xlsx cannot be validated: {exc}")
    else:
        errors.append("model_labels.xlsx is missing.")

    unexpected_python_files = sorted(
        path.name
        for path in PROJECT_DIR.glob("*.py")
        if path.name not in ALLOWED_ROOT_PYTHON_FILES
    )
    if unexpected_python_files:
        errors.append(
            "Unexpected Python files exist in the project root: "
            + ", ".join(unexpected_python_files)
            + ". Use only the maintained workflow scripts; put disposable data under the job tmp directory."
        )

    return {
        "job_directory": str(job_dir),
        "inspection": inspection,
        "labeled_counts": label_metrics.get("labeled_counts", {}),
        "config_exists": config_path.is_file(),
        "seed_counts": seed_counts,
        "label_quality": label_metrics,
        "unexpected_python_files": unexpected_python_files,
        "blocking_errors": errors,
        "warnings": warnings,
        "ready_for_supervised_run": not errors,
        "state": read_state(job_dir),
    }


def load_label_sheet(job_dir: Path) -> pd.DataFrame:
    path = job_dir / "model_labels.xlsx"
    if not path.is_file():
        raise FileNotFoundError(path)
    frame = pd.read_excel(path, sheet_name="Model labels")
    if "Sample ID" not in frame.columns:
        frame.insert(0, "Sample ID", [f"row-{index:06d}" for index in range(1, len(frame) + 1)])
    required = {"Sample ID", "Phrase", "Model Label", "Model Confidence"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError("model_labels.xlsx is missing columns: " + ", ".join(missing))
    for column in ("Sample ID", "Phrase", "Model Label", "Model Confidence", "Model Notes"):
        if column in frame.columns:
            frame[column] = frame[column].astype(object)
    return frame


def label_summary(frame: pd.DataFrame) -> tuple[pd.Series, pd.Series]:
    labels = frame["Model Label"].fillna("").astype(str).str.strip().str.lower().map(LABEL_ALIASES)
    phrases = frame["Phrase"].fillna("").astype(str).str.strip()
    return labels, phrases


def current_sample_intent_counts(frame: pd.DataFrame) -> dict[str, int]:
    """Count job-local reviewed labels, excluding auxiliary saved knowledge."""
    labels, phrases = label_summary(frame)
    if "Knowledge Source" in frame.columns:
        sources = frame["Knowledge Source"].fillna("").astype(str).str.strip().str.lower()
        auxiliary = sources.isin({"model_reviewed", "review_corrected"})
    else:
        auxiliary = pd.Series(False, index=frame.index)
    current = phrases.ne("") & ~auxiliary
    return {
        label: int((current & labels.eq(label)).sum())
        for label in ("commercial", "informational", "garbage")
    }


def sanitize_strong_signals(
    frame: pd.DataFrame,
    signals: list[str],
    expected_label: str,
) -> tuple[list[str], list[dict[str, object]]]:
    """Remove collision-prone signals using structure and reviewed job labels."""
    labels, phrases = label_summary(frame)
    relevant = labels.isin(["commercial", "informational"]) & phrases.ne("")
    retained: list[str] = []
    rejected: list[dict[str, object]] = []
    opposite_label = (
        "informational" if expected_label == "commercial" else "commercial"
    )
    for signal in signals:
        reason = marker_safety_reason(signal)
        matched = signal_mask(phrases, [signal]) & relevant
        expected_count = int((matched & labels.eq(expected_label)).sum())
        opposite_count = int((matched & labels.eq(opposite_label)).sum())
        labeled_matches = expected_count + opposite_count
        if (
            reason is None
            and labeled_matches >= 4
            and opposite_count >= 2
            and opposite_count / labeled_matches >= 0.30
        ):
            reason = "reviewed_label_collision"
        if reason is None:
            retained.append(signal)
        else:
            rejected.append(
                {
                    "signal": signal,
                    "reason": reason,
                    "expected_examples": expected_count,
                    "opposite_examples": opposite_count,
                }
            )
    return retained, rejected


def bootstrap_seeds_from_real_labels(job_dir: Path, frame: pd.DataFrame) -> dict[str, int]:
    """Fill only missing seed slots from model-reviewed, real source phrases.

    Seeds are an aid for the CPU pipeline, not input the user must provide.
    They must never be invented before the model has seen the semantic core.
    Existing manually configured values stay first and are never replaced.
    """
    config_path = job_dir / "job_config.json"
    try:
        config = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(config, dict):
        return {}

    labels, phrases = label_summary(frame)

    def unique_phrases(mask: pd.Series) -> list[str]:
        result: list[str] = []
        seen: set[str] = set()
        for phrase in phrases[mask]:
            normalized = re.sub(r"\s+", " ", str(phrase).lower().replace("ё", "е")).strip()
            if normalized and normalized not in seen:
                result.append(str(phrase).strip())
                seen.add(normalized)
        return result

    candidates = {
        ("topic", "relevant_seeds"): unique_phrases(labels.isin(["commercial", "informational"])),
        ("intent", "commercial_seeds"): unique_phrases(labels.eq("commercial")),
        ("intent", "informational_seeds"): unique_phrases(labels.eq("informational")),
        ("topic", "garbage_seeds"): unique_phrases(labels.eq("garbage")),
    }
    changed = False
    counts: dict[str, int] = {}
    for (section, key), values in candidates.items():
        section_data = config.setdefault(section, {})
        if not isinstance(section_data, dict):
            continue
        current = configured_values(config, section, key)
        seen = {
            re.sub(r"\s+", " ", value.lower().replace("ё", "е")).strip()
            for value in current
        }
        merged = list(current)
        for value in values:
            normalized = re.sub(r"\s+", " ", value.lower().replace("ё", "е")).strip()
            if normalized not in seen:
                merged.append(value)
                seen.add(normalized)
            if len(merged) >= MIN_SEEDS_PER_CLASS:
                break
        if merged != current:
            section_data[key] = merged
            changed = True
        counts[f"{section}.{key}"] = len(merged)

    if changed:
        temporary = config_path.with_suffix(".tmp")
        temporary.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(config_path)
    return counts


def sample_rows(job_dir: Path, offset: int, limit: int, only_unlabeled: bool = True) -> dict[str, object]:
    if offset < 0 or limit < 1 or limit > 50:
        raise ValueError("offset must be non-negative and limit must be from 1 to 50.")
    frame = load_label_sheet(job_dir)
    labels, phrases = label_summary(frame)
    candidates = frame[phrases.ne("") & (labels.isna() if only_unlabeled else pd.Series(True, index=frame.index))].copy()
    if "Search Volume" in candidates:
        candidates["_priority"] = pd.to_numeric(candidates["Search Volume"], errors="coerce").fillna(0)
    else:
        candidates["_priority"] = 0.0
    candidates = candidates.sort_values(["_priority", "Sample ID"], ascending=[False, True])
    view = candidates.iloc[offset : offset + limit]
    rows = [
        {
            "id": str(row["Sample ID"]),
            "phrase": str(row["Phrase"]),
            "priority": round(float(row["_priority"]), 3),
        }
        for _, row in view.iterrows()
    ]
    return {
        "stage": read_state(job_dir).get("stage", "initial_labeling"),
        "topic": read_state(job_dir).get("topic", ""),
        "labeling_contract": INTENT_LABELING_CONTRACT,
        "offset": offset,
        "limit": limit,
        "total_unlabeled": int(len(candidates)),
        "rows": rows,
    }


def family_coverage_status(frame: pd.DataFrame) -> dict[str, int]:
    """Count mandatory real examples reserved for source-family coverage."""
    labels, phrases = label_summary(frame)
    if "Knowledge Source" not in frame.columns:
        return {"total": 0, "labeled": 0, "remaining": 0}
    sources = frame["Knowledge Source"].fillna("").astype(str).str.strip().str.lower()
    coverage = sources.eq("current family coverage sample") & phrases.ne("")
    labeled = coverage & labels.isin(VALID_LABELS)
    return {
        "total": int(coverage.sum()),
        "labeled": int(labeled.sum()),
        "remaining": int((coverage & ~labeled).sum()),
    }


def select_family_coverage_review(job_dir: Path, limit: int) -> dict[str, object]:
    """Return the bounded real-phrase batch that closes family blind spots."""
    if limit < 1 or limit > FAMILY_COVERAGE_REVIEW_BATCH_SIZE:
        raise ValueError(
            f"limit must be from 1 to {FAMILY_COVERAGE_REVIEW_BATCH_SIZE}."
        )
    frame = load_label_sheet(job_dir)
    labels, phrases = label_summary(frame)
    if "Knowledge Source" not in frame.columns:
        return {
            "stage": "intent_family_coverage",
            "selection": "complete",
            "rows": [],
            "remaining_after_batch": 0,
        }
    sources = frame["Knowledge Source"].fillna("").astype(str).str.strip().str.lower()
    candidates = frame[
        sources.eq("current family coverage sample")
        & phrases.ne("")
        & labels.isna()
    ].copy()
    if "Coverage Family" not in candidates.columns:
        candidates["Coverage Family"] = ""
    candidates = candidates.sort_values(
        ["Coverage Family", "Sample ID"], ascending=[True, True], kind="stable"
    )
    view = candidates.head(limit)
    rows = [
        {
            "id": str(row["Sample ID"]),
            "phrase": str(row["Phrase"]),
            "family": str(row.get("Coverage Family", "")),
            "reason": "unrepresented_source_family",
        }
        for _, row in view.iterrows()
    ]
    status = family_coverage_status(frame)
    write_state(
        job_dir,
        "intent_family_coverage",
        family_coverage_labeled=status["labeled"],
        family_coverage_remaining=status["remaining"],
    )
    return {
        "stage": "intent_family_coverage",
        "topic": read_state(job_dir).get("topic", ""),
        "labeling_contract": INTENT_LABELING_CONTRACT,
        "selection": "unrepresented_source_families",
        "rows": rows,
        "remaining_after_batch": max(0, status["remaining"] - len(rows)),
    }


def select_priority_review(job_dir: Path, limit: int) -> dict[str, object]:
    """Return unlabeled rows from the protected high-priority set.

    This must use exactly the same combined priority metric as
    ``label_quality_status``.  Otherwise a job can reach its active-learning
    target and then be blocked by an unreviewed high-volume phrase that was
    never offered to the model.
    """
    if limit < 1 or limit > 50:
        raise ValueError("limit must be from 1 to 50.")
    frame = load_label_sheet(job_dir)
    labels, phrases = label_summary(frame)
    confidence = pd.to_numeric(frame["Model Confidence"], errors="coerce")
    reviewed = labels.notna() & phrases.ne("") & confidence.between(0.0, 1.0, inclusive="both")
    lookup = normalized_column_lookup(frame)
    priority = pd.Series(0.0, index=frame.index)
    for column in (lookup.get("search volume"), lookup.get("occurrences")):
        if column is not None:
            priority += pd.to_numeric(frame[column], errors="coerce").fillna(0)
    protected_index = priority.nlargest(min(HIGH_PRIORITY_REVIEW_ROWS, len(frame))).index
    candidates = frame.loc[protected_index].copy()
    candidates["_priority"] = priority.loc[protected_index]
    candidates = candidates[~reviewed.loc[protected_index]].sort_values(
        ["_priority", "Sample ID"], ascending=[False, True]
    )
    view = candidates.head(limit)
    rows = [
        {
            "id": str(row["Sample ID"]),
            "phrase": str(row["Phrase"]),
            "reason": "required_priority_coverage",
            "priority": round(float(row["_priority"]), 3),
        }
        for _, row in view.iterrows()
    ]
    return {
        "stage": read_state(job_dir).get("stage", "initial_labeling"),
        "selection": "required_priority_coverage",
        "rows": rows,
        "remaining_priority_rows": int(len(candidates) - len(view)),
    }


def policy_context(job_dir: Path, limit: int) -> dict[str, object]:
    """Give the model a small basis for intent and topic-relevance policies."""
    if limit < 3 or limit > 12:
        raise ValueError("limit must be from 3 to 12.")
    frame = load_label_sheet(job_dir)
    labels, phrases = label_summary(frame)
    config = json.loads((job_dir / "job_config.json").read_text(encoding="utf-8"))
    state = read_state(job_dir)
    weak_question_values = configured_values(config, "intent", "weak_question_markers")
    intent_config = config.get("intent", {}) if isinstance(config.get("intent", {}), dict) else {}
    result: dict[str, object] = {
        "topic": state.get("topic", ""),
        "labeling_contract": INTENT_LABELING_CONTRACT,
        "examples": {},
        "current_signals": {
            "commercial": configured_values(config, "intent", "commercial_markers"),
            "informational": configured_values(config, "intent", "informational_markers"),
            "weak_question": weak_question_values,
            "rejected": intent_config.get("rejected_strong_markers", []),
            "reviewed_families": (
                config.get("intent", {}).get("family_rules", {})
                if isinstance(config.get("intent", {}), dict)
                else {}
            ),
        },
        "signal_coverage": state.get("signal_policy_coverage", {}),
        "weak_question_context_examples": {},
        "structural_family_examples": [],
    }
    for label in ("commercial", "informational", "garbage"):
        subset = frame[labels.eq(label) & phrases.ne("")].copy()
        if "Search Volume" in subset:
            subset["_priority"] = pd.to_numeric(subset["Search Volume"], errors="coerce").fillna(0)
            subset = subset.sort_values("_priority", ascending=False)
        result["examples"][label] = subset["Phrase"].astype(str).head(limit).tolist()
    if weak_question_values:
        weak_mask = signal_mask(phrases, weak_question_values)
        for label in ("commercial", "informational"):
            subset = frame[weak_mask & labels.eq(label) & phrases.ne("")].copy()
            if "Search Volume" in subset:
                subset["_priority"] = pd.to_numeric(
                    subset["Search Volume"], errors="coerce"
                ).fillna(0)
                subset = subset.sort_values("_priority", ascending=False)
            result["weak_question_context_examples"][label] = (
                subset["Phrase"].astype(str).head(limit).tolist()
            )
    candidates_path, labels_path = intent_family_paths(job_dir)
    if candidates_path.is_file() and labels_path.is_file():
        candidates_payload = json.loads(candidates_path.read_text(encoding="utf-8"))
        labels_payload = json.loads(labels_path.read_text(encoding="utf-8"))
        family_labels = labels_payload.get("labels", {})
        if isinstance(family_labels, dict):
            for record in candidates_payload.get("families", []):
                if not isinstance(record, dict) or str(record.get("kind", "")) != "structural":
                    continue
                family_id = str(record.get("id", ""))
                label = str(family_labels.get(family_id, "neutral"))
                result["structural_family_examples"].append(
                    {
                        "family": str(record.get("pattern", "")),
                        "decision": label,
                        "examples": [str(value) for value in record.get("examples", [])[:2]],
                    }
                )
                if len(result["structural_family_examples"]) >= 12:
                    break
    result["instruction"] = (
        "Create a universal policy adapted to the stated topic and the reviewed structural-family examples: "
        "3-8 explicit commercial prototypes, "
        "5-12 implicit commercial prototypes without buy/price words, 5-12 informational prototypes, "
        "8-20 strong commercial signals, 12-30 strong informational signals, and 3-10 weak question signals. "
        "Strong commercial signals must express concrete offer/result seeking or a transaction/conversion action, "
        "including implicit marketplace demand without buy/price words. Strong informational signals "
        "must independently express instructions, diagnosis, reference data, specifications, diagrams, comparison, "
        "or explanation. Weak question words (the topic language equivalents of where/how/how much/which) are never "
        "decisive alone: a phrase equivalent to 'where to buy' or 'how much does it cost' remains commercial. "
        "For weak-question examples, derive strong context signals for both transaction/conversion and informational "
        "reference/location meanings; never promote the weak question word itself. Use a trailing * only for productive "
        "word stems of at least four characters and keep one concept per signal. Replace every signal listed under "
        "current_signals.rejected with a safer exact, longer-stem, or multiword signal. Do not copy automotive vocabulary "
        "unless the current topic is automotive. Also create "
        "5-12 broad relevant topic prototypes, "
        "and 8-15 diverse hard negative garbage prototypes. Informational signals must cover questions, "
        "instructions, diagnostics, locations, specifications, diagrams, and comparisons for this topic. "
        "Hard negatives must cover plausible lexical "
        "collisions with the topic, unrelated entities, stories/media, jobs/services, and other meanings "
        "of topic words. They are synthetic boundary examples, never source rows."
        " The informational policy must explicitly cover the topic-language equivalents of reviews, opinions, "
        "user experience, pros and cons, and product/service overviews in addition to instructions, diagnostics, "
        "reference data, specifications, diagrams, comparisons, and explanations. Reuse decisive reviewed_families "
        "as signals where appropriate; never turn neutral brand or product families into intent markers."
    )
    return result


def apply_policy_inline(
    job_dir: Path,
    commercial: str,
    implicit_commercial: str,
    informational: str,
    commercial_signals: str,
    informational_signals: str,
    weak_question_signals: str,
    relevant: str,
    garbage: str,
) -> dict[str, object]:
    def parse(value: str, limit: int = 15) -> list[str]:
        seen: set[str] = set()
        result: list[str] = []
        for item in value.split(";"):
            text = re.sub(r"\s+", " ", item).strip()
            key = text.lower().replace("ё", "е")
            if text and key not in seen:
                seen.add(key)
                result.append(text)
        return result[:limit]
    commercial_values, informational_values = parse(commercial), parse(informational)
    implicit_commercial_values = parse(implicit_commercial)
    commercial_signal_values = parse(commercial_signals, 20)
    informational_signal_values = parse(informational_signals, 30)
    weak_question_signal_values = parse(weak_question_signals, 10)
    relevant_values, garbage_values = parse(relevant), parse(garbage)
    if len(commercial_values) < 3 or len(informational_values) < 5:
        raise ValueError("Provide at least three commercial and five informational prototypes separated by semicolons.")
    if len(implicit_commercial_values) < 5:
        raise ValueError("Provide at least five implicit commercial structures without buy or price words.")
    if len(commercial_signal_values) < 5:
        raise ValueError("Provide at least five strong commercial signals.")
    if len(informational_signal_values) < 12:
        raise ValueError("Provide at least twelve strong informational signals.")
    if len(weak_question_signal_values) < 3:
        raise ValueError("Provide at least three weak question signals.")
    if len(relevant_values) < 5 or len(garbage_values) < 5:
        raise ValueError("Provide at least five relevant and five hard-negative garbage prototypes.")
    path = job_dir / "job_config.json"
    config = json.loads(path.read_text(encoding="utf-8"))
    previous_state = read_state(job_dir)
    final_refresh = bool(previous_state.get("final_policy_refinement_requested"))
    commercial_signal_keys = {normalize_signal(value) for value in commercial_signal_values}
    informational_signal_keys = {normalize_signal(value) for value in informational_signal_values}
    overlap = sorted(commercial_signal_keys & informational_signal_keys)
    if overlap:
        raise ValueError("Strong commercial and informational signals overlap: " + ", ".join(overlap))
    weak_question_signal_values = [
        value
        for value in weak_question_signal_values
        if normalize_signal(value) not in commercial_signal_keys | informational_signal_keys
    ]
    if len(weak_question_signal_values) < 3:
        raise ValueError("At least three weak question signals must remain distinct from strong signals.")
    label_frame = load_label_sheet(job_dir)
    current_commercial_signals = configured_values(config, "intent", "commercial_markers")
    current_informational_signals = configured_values(config, "intent", "informational_markers")
    current_weak_signals = configured_values(config, "intent", "weak_question_markers")
    baseline_coverage = intent_signal_coverage(
        label_frame,
        current_commercial_signals,
        current_informational_signals,
        current_weak_signals,
    )
    commercial_signal_values, rejected_commercial_signals = sanitize_strong_signals(
        label_frame, commercial_signal_values, "commercial"
    )
    informational_signal_values, rejected_informational_signals = sanitize_strong_signals(
        label_frame, informational_signal_values, "informational"
    )
    rejected_strong_markers = [
        {**item, "expected_label": "commercial"}
        for item in rejected_commercial_signals
    ] + [
        {**item, "expected_label": "informational"}
        for item in rejected_informational_signals
    ]
    candidate_coverage = intent_signal_coverage(
        label_frame,
        commercial_signal_values,
        informational_signal_values,
        weak_question_signal_values,
    )
    if final_refresh:
        accepted, rejection_reasons = final_policy_candidate_is_safe(
            baseline_coverage, candidate_coverage
        )
        if not accepted:
            labels_now, _ = label_summary(label_frame)
            reviewed_now = int(labels_now.notna().sum())
            write_state(
                job_dir,
                "active_review",
                final_policy_refinement_requested=False,
                final_policy_checked_at_labels=reviewed_now,
                final_policy_candidate_accepted=False,
                final_policy_candidate_rejection_reasons=rejection_reasons,
                final_policy_candidate_coverage=candidate_coverage,
            )
            return {
                "status": "existing_policy_retained",
                "reason": "Final policy candidate failed reviewed-example regression safeguards.",
                "rejection_reasons": rejection_reasons,
                "baseline_coverage": baseline_coverage,
                "candidate_coverage": candidate_coverage,
            }
    config["intent_policy"] = {
        "commercial_prototypes": commercial_values,
        "implicit_commercial_prototypes": implicit_commercial_values,
        "informational_prototypes": informational_values,
        "synthetic_weight": 0.75,
        "informational_evidence_margin": 0.005,
        "strength": 0.12,
        "minimum_similarity": 0.55,
    }
    intent_config = config.setdefault("intent", {})
    intent_config["commercial_markers"] = commercial_signal_values
    intent_config["informational_markers"] = informational_signal_values
    intent_config["weak_question_markers"] = weak_question_signal_values
    intent_config["rejected_strong_markers"] = rejected_strong_markers
    family_rules = (
        intent_config.get("family_rules", {})
        if isinstance(intent_config.get("family_rules", {}), dict)
        else {}
    )
    family_rules = {
        label: [str(value) for value in family_rules.get(label, [])]
        for label in ("commercial", "informational", "neutral")
    }
    forced_neutral_families: list[str] = []
    for label in ("commercial", "informational"):
        retained = []
        for pattern in family_rules[label]:
            if family_overlaps_signals(pattern, weak_question_signal_values):
                forced_neutral_families.append(pattern)
            else:
                retained.append(pattern)
        family_rules[label] = retained
    family_rules["neutral"] = list(
        dict.fromkeys(family_rules["neutral"] + forced_neutral_families)
    )
    intent_config["family_rules"] = family_rules
    if forced_neutral_families:
        candidates_path, labels_path = intent_family_paths(job_dir)
        if candidates_path.is_file() and labels_path.is_file():
            candidates_payload = json.loads(candidates_path.read_text(encoding="utf-8"))
            labels_payload = json.loads(labels_path.read_text(encoding="utf-8"))
            labels_map = labels_payload.get("labels", {})
            if isinstance(labels_map, dict):
                for record in candidates_payload.get("families", []):
                    if (
                        isinstance(record, dict)
                        and str(record.get("pattern", "")) in forced_neutral_families
                    ):
                        labels_map[str(record.get("id", ""))] = "neutral"
                labels_temporary = labels_path.with_suffix(".tmp")
                labels_temporary.write_text(
                    json.dumps({"labels": labels_map}, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                labels_temporary.replace(labels_path)
    intent_config.setdefault("informational_decision_margin", 0.12)
    intent_config.setdefault("strong_informational_decision_margin", 0.02)
    intent_config.setdefault("weak_question_informational_margin", 0.02)
    intent_config.setdefault("family_override_tolerance", 0.05)
    relevance_config = config.setdefault("relevance", {})
    relevance_config["relevant_prototypes"] = relevant_values
    relevance_config["garbage_prototypes"] = garbage_values
    relevance_config.setdefault("synthetic_weight", 1.0)
    temporary = path.with_suffix(".tmp")
    temporary.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)
    coverage = candidate_coverage
    labels_now, _ = label_summary(label_frame)
    reviewed_now = int(labels_now.notna().sum())
    attempts = int(previous_state.get("signal_policy_attempts", 0) or 0) + 1
    needs_refinement = bool(coverage["needs_refinement"]) and attempts < 2
    write_state(
        job_dir,
        "active_review",
        intent_policy_ready=True,
        signal_policy_attempts=attempts,
        signal_policy_coverage=coverage,
        signal_policy_label_count=reviewed_now,
        signal_policy_refinement_needed=needs_refinement,
        signal_policy_warning=bool(coverage["needs_refinement"]) and attempts >= 2,
        final_policy_refinement_requested=False,
        final_policy_checked_at_labels=(reviewed_now if final_refresh else int(previous_state.get("final_policy_checked_at_labels", 0) or 0)),
        final_policy_candidate_accepted=(True if final_refresh else previous_state.get("final_policy_candidate_accepted")),
    )
    return {
        "status": "classification_policy_applied",
        "commercial_prototypes": len(commercial_values),
        "implicit_commercial_prototypes": len(implicit_commercial_values),
        "informational_prototypes": len(informational_values),
        "commercial_signals": len(commercial_signal_values),
        "informational_signals": len(informational_signal_values),
        "weak_question_signals": len(weak_question_signal_values),
        "rejected_strong_markers": rejected_strong_markers,
        "forced_neutral_weak_families": len(forced_neutral_families),
        "signal_coverage": coverage,
        "refinement_needed": needs_refinement,
        "relevant_prototypes": len(relevant_values),
        "garbage_prototypes": len(garbage_values),
    }


def parse_inline_labels(value: str) -> list[dict[str, object]]:
    """Parse a deliberately small, phrase-free label transport for Bionic."""
    payload: list[dict[str, object]] = []
    for raw_item in value.split(";"):
        item = raw_item.strip()
        if not item:
            continue
        fields = [field.strip() for field in item.split("|")]
        if len(fields) != 3:
            raise ValueError("Each inline label must be SampleID|label|confidence, separated by semicolons.")
        payload.append({"id": fields[0], "label": fields[1], "confidence": fields[2], "notes": ""})
    if not payload:
        raise ValueError("At least one inline label is required.")
    return payload


def apply_label_payload(job_dir: Path, payload: object) -> dict[str, object]:
    if not isinstance(payload, list):
        raise ValueError("Label JSON must be an array of {id, label, confidence, notes?} objects.")
    frame = load_label_sheet(job_dir)
    id_index = {str(value): index for index, value in frame["Sample ID"].items()}
    seen: set[str] = set()
    new_labels = 0
    updated_labels = 0
    unchanged_labels = 0
    for item in payload:
        if not isinstance(item, dict):
            raise ValueError("Every label item must be an object.")
        sample_id = str(item.get("id", "")).strip()
        raw_label = str(item.get("label", "")).strip().lower()
        label = LABEL_ALIASES.get(raw_label)
        if not sample_id or sample_id not in id_index:
            raise ValueError(f"Unknown Sample ID: {sample_id or '<empty>'}")
        if sample_id in seen:
            raise ValueError(f"Duplicate Sample ID in label JSON: {sample_id}")
        if label not in VALID_LABELS:
            raise ValueError(f"Invalid label for {sample_id}: {raw_label}")
        try:
            confidence = float(item.get("confidence"))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid confidence for {sample_id}") from exc
        if not 0.0 <= confidence <= 1.0:
            raise ValueError(f"Confidence for {sample_id} must be from 0 to 1.")
        row = id_index[sample_id]
        raw_previous_label = frame.loc[row, "Model Label"]
        previous_label = LABEL_ALIASES.get("" if pd.isna(raw_previous_label) else str(raw_previous_label).strip().lower())
        previous_confidence = pd.to_numeric(pd.Series([frame.loc[row, "Model Confidence"]]), errors="coerce").iloc[0]
        raw_previous_notes = frame.loc[row, "Model Notes"]
        previous_notes = "" if pd.isna(raw_previous_notes) else str(raw_previous_notes).strip()
        notes = str(item.get("notes", "")).strip()
        if previous_label == label and pd.notna(previous_confidence) and abs(float(previous_confidence) - confidence) < 1e-9 and previous_notes == notes:
            unchanged_labels += 1
            seen.add(sample_id)
            continue
        if previous_label is None:
            new_labels += 1
        else:
            updated_labels += 1
        frame.loc[row, "Model Label"] = label
        frame.loc[row, "Model Confidence"] = confidence
        frame.loc[row, "Model Notes"] = notes
        seen.add(sample_id)
    applied = new_labels + updated_labels
    if applied:
        temporary = job_dir / "tmp" / "model_labels.updated.xlsx"
        temporary.parent.mkdir(parents=True, exist_ok=True)
        frame.to_excel(temporary, sheet_name="Model labels", index=False)
        temporary.replace(job_dir / "model_labels.xlsx")
    labels, _ = label_summary(frame)
    seed_counts = bootstrap_seeds_from_real_labels(job_dir, frame)
    stage = "active_review" if int(labels.notna().sum()) >= 100 else "initial_labeling"
    previous_state = read_state(job_dir)
    pending_relevance_ids = {
        str(value) for value in previous_state.get("pending_relevance_review_ids", [])
    }
    pending_intent_ids = {
        str(value) for value in previous_state.get("pending_intent_review_ids", [])
    }
    pending_policy_conflict_ids = {
        str(value) for value in previous_state.get("pending_policy_conflict_review_ids", [])
    }
    labeled_ids = set(frame.loc[labels.notna(), "Sample ID"].astype(str))
    relevance_audit_completed = bool(pending_relevance_ids) and pending_relevance_ids.issubset(labeled_ids)
    intent_audit_completed = bool(pending_intent_ids) and pending_intent_ids.issubset(labeled_ids)
    policy_conflict_audit_completed = (
        bool(pending_policy_conflict_ids)
        and pending_policy_conflict_ids.issubset(labeled_ids)
    )
    state_updates: dict[str, object] = {
        "labels_reviewed": int(labels.notna().sum()),
        "remaining_review_rows": int(labels.isna().sum()),
    }
    if relevance_audit_completed:
        state_updates["relevance_audit_batches"] = int(previous_state.get("relevance_audit_batches", 0) or 0) + 1
        state_updates["pending_relevance_review_ids"] = []
        garbage_before = int(previous_state.get("relevance_audit_garbage_before", 0) or 0)
        garbage_after = int(labels.eq("garbage").sum())
        yields = [
            int(value)
            for value in (previous_state.get("relevance_audit_garbage_yields", []) or [])
        ]
        yields.append(max(0, garbage_after - garbage_before))
        state_updates["relevance_audit_garbage_yields"] = yields[-6:]
        state_updates["relevance_audit_garbage_before"] = garbage_after
    if intent_audit_completed:
        state_updates["intent_audit_batches"] = int(previous_state.get("intent_audit_batches", 0) or 0) + 1
        state_updates["pending_intent_review_ids"] = []
    if policy_conflict_audit_completed:
        state_updates["policy_conflict_audit_batches"] = int(
            previous_state.get("policy_conflict_audit_batches", 0) or 0
        ) + 1
        state_updates["pending_policy_conflict_review_ids"] = []
    state = write_state(
        job_dir,
        stage,
        **state_updates,
    )
    return {
        "status": "labels_applied" if applied else "no_progress",
        "new": new_labels,
        "updated": updated_labels,
        "unchanged": unchanged_labels,
        "applied": applied,
        "seed_counts": seed_counts,
        "state": state,
    }


def apply_labels(job_dir: Path, source: Path) -> dict[str, object]:
    if not source.is_file():
        raise FileNotFoundError(source)
    try:
        payload = json.loads(source.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Label JSON cannot be read: {exc}") from exc
    return apply_label_payload(job_dir, payload)


def select_active_review(job_dir: Path, limit: int) -> dict[str, object]:
    if limit < 1 or limit > 50:
        raise ValueError("limit must be from 1 to 50.")
    frame = load_label_sheet(job_dir)
    labels, phrases = label_summary(frame)
    labeled = frame[labels.notna() & phrases.ne("")].copy()
    labeled_labels = labels.loc[labeled.index]
    if len(labeled) < 20 or labeled_labels.nunique() < 2:
        return sample_rows(job_dir, 0, limit, only_unlabeled=True) | {
            "selection": "initial_priority",
            "reason": "Need at least 20 labels across two real classes for active review.",
        }
    unlabeled = frame[labels.isna() & phrases.ne("")].copy()
    if unlabeled.empty:
        return {"stage": "ready_to_run", "selection": "complete", "rows": []}
    # This classifier is used only after the initial labels exist.  Import it
    # here so `next`, `sample`, and label import do not pay scikit-learn's
    # startup cost on every agent turn.
    import numpy as np
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.linear_model import LogisticRegression

    vectorizer = TfidfVectorizer(analyzer="char_wb", ngram_range=(3, 5), min_df=1, max_features=12000)
    train = vectorizer.fit_transform(labeled["Phrase"].astype(str))
    candidate = vectorizer.transform(unlabeled["Phrase"].astype(str))
    classifier = LogisticRegression(max_iter=1000, class_weight="balanced", random_state=42).fit(train, labeled_labels)
    probabilities = classifier.predict_proba(candidate)
    ordered = np.sort(probabilities, axis=1)
    margin = ordered[:, -1] - ordered[:, -2] if probabilities.shape[1] > 1 else np.ones(len(unlabeled))
    entropy = -(probabilities * np.log(np.maximum(probabilities, 1e-12))).sum(axis=1)
    priority = pd.to_numeric(unlabeled.get("Search Volume", 0), errors="coerce").fillna(0).to_numpy()
    # Similar text patterns with both commercial and informational labelled
    # neighbours are valuable calibration cases, even when the classifier is
    # superficially confident.
    similarity = (candidate @ train.T).toarray()
    nearest = np.argpartition(similarity, -min(5, similarity.shape[1]), axis=1)[:, -min(5, similarity.shape[1]):]
    known = labeled_labels.to_numpy()
    structural_conflict = np.asarray(
        [len(set(known[indexes]) & {"commercial", "informational"}) == 2 for indexes in nearest],
        dtype=int,
    )
    conflict_by_row = pd.Series(structural_conflict, index=unlabeled.index)
    ranking = np.lexsort((-priority, -entropy, margin, -structural_conflict))[:limit]
    selected = unlabeled.iloc[ranking].copy()
    rows = [
        {
            "id": str(row["Sample ID"]),
            "phrase": str(row["Phrase"]),
            "reason": "structural_label_conflict" if conflict_by_row.loc[row_index] else "low_margin",
            "priority": round(float(pd.to_numeric(row.get("Search Volume", 0), errors="coerce") if pd.notna(pd.to_numeric(row.get("Search Volume", 0), errors="coerce")) else 0), 3),
        }
        for row_index, row in selected.iterrows()
    ]
    write_state(job_dir, "active_review", labels_reviewed=int(labels.notna().sum()), remaining_review_rows=int(labels.isna().sum()))
    return {"stage": "active_review", "selection": "uncertain_tfidf", "rows": rows, "remaining_unlabeled": int(len(unlabeled))}


def select_relevance_review(job_dir: Path, limit: int) -> dict[str, object]:
    """Select likely out-of-topic rows without letting a tiny garbage class dominate."""
    if limit < 1 or limit > 50:
        raise ValueError("limit must be from 1 to 50.")
    frame = load_label_sheet(job_dir)
    labels, phrases = label_summary(frame)
    labeled = frame[labels.notna() & phrases.ne("")].copy()
    unlabeled = frame[labels.isna() & phrases.ne("")].copy()
    if unlabeled.empty:
        return {"stage": "ready_to_run", "selection": "complete", "rows": []}
    relevant = labeled[labels.loc[labeled.index].isin(["commercial", "informational"])]
    if len(relevant) < 10:
        return sample_rows(job_dir, 0, limit, only_unlabeled=True) | {
            "selection": "relevance_bootstrap",
            "reason": "Need relevant examples before out-of-topic candidate selection.",
        }

    import numpy as np
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.linear_model import LogisticRegression

    corpus = pd.concat([labeled["Phrase"], unlabeled["Phrase"]]).astype(str)
    vectorizer = TfidfVectorizer(
        analyzer="char_wb", ngram_range=(3, 5), min_df=1, max_features=20000, sublinear_tf=True
    )
    vectorizer.fit(corpus)
    relevant_features = vectorizer.transform(relevant["Phrase"].astype(str))
    candidate_features = vectorizer.transform(unlabeled["Phrase"].astype(str))
    similarity = candidate_features @ relevant_features.T
    relevant_support = np.asarray(similarity.max(axis=1).toarray()).ravel()
    novelty = 1.0 - relevant_support

    labeled_targets = np.where(labels.loc[labeled.index].eq("garbage"), "garbage", "relevant")
    garbage_probability = np.zeros(len(unlabeled), dtype=float)
    if len(set(labeled_targets)) == 2:
        train_features = vectorizer.transform(labeled["Phrase"].astype(str))
        garbage_count = int(np.sum(labeled_targets == "garbage"))
        relevant_count = int(np.sum(labeled_targets == "relevant"))
        class_weight = {"garbage": min(8.0, relevant_count / max(garbage_count, 1)), "relevant": 1.0}
        classifier = LogisticRegression(max_iter=1000, class_weight=class_weight, random_state=42).fit(
            train_features, labeled_targets
        )
        probabilities = classifier.predict_proba(candidate_features)
        garbage_index = list(classifier.classes_).index("garbage")
        garbage_probability = probabilities[:, garbage_index]

    priority = pd.to_numeric(unlabeled.get("Search Volume", 0), errors="coerce").fillna(0).to_numpy()
    if priority.max() > 0:
        priority = np.log1p(priority) / max(float(np.log1p(priority).max()), 1.0)
    candidate_score = 0.55 * garbage_probability + 0.35 * novelty + 0.10 * priority
    chosen: list[int] = []
    selection_reason: dict[int, str] = {}

    def add_indexes(indexes: object, reason: str, quota: int) -> None:
        for raw_index in list(indexes):
            index = int(raw_index)
            if index not in selection_reason:
                chosen.append(index)
                selection_reason[index] = reason
            if sum(value == reason for value in selection_reason.values()) >= quota:
                break

    model_quota = max(1, limit // 3)
    novelty_quota = max(1, limit // 4)
    if float(garbage_probability.max(initial=0.0)) > 0:
        add_indexes(
            np.argsort(-garbage_probability, kind="stable"),
            "garbage_model_candidate",
            model_quota,
        )
    add_indexes(
        np.argsort(-novelty, kind="stable"),
        "semantic_outlier_candidate",
        novelty_quota,
    )
    remaining = limit - len(chosen)
    if remaining > 0 and len(unlabeled) >= 4:
        # Greedy farthest-first selection is a compact microcluster audit: each
        # new row represents a text neighbourhood not already covered by the
        # higher-risk candidates. It avoids another heavyweight model call.
        available = np.ones(len(unlabeled), dtype=bool)
        available[chosen] = False
        representatives: list[int] = []
        while len(representatives) < remaining and available.any():
            anchors = chosen + representatives
            if anchors:
                similarity_to_anchors = candidate_features @ candidate_features[anchors].T
                maximum_similarity = np.asarray(similarity_to_anchors.max(axis=1).toarray()).ravel()
                diversity = 1.0 - maximum_similarity
            else:
                diversity = novelty.copy()
            diversified_score = 0.75 * diversity + 0.25 * candidate_score
            diversified_score[~available] = -np.inf
            selected_index = int(np.argmax(diversified_score))
            representatives.append(selected_index)
            available[selected_index] = False
        add_indexes(
            representatives,
            "diverse_cluster_representative",
            remaining,
        )
    if len(chosen) < limit:
        add_indexes(
            np.argsort(-candidate_score, kind="stable"),
            "combined_relevance_risk",
            limit - len(chosen),
        )
    ranking = np.asarray(chosen[:limit], dtype=int)
    selected = unlabeled.iloc[ranking].copy()
    selected_scores = candidate_score[ranking]
    selected_support = relevant_support[ranking]
    selected_garbage = garbage_probability[ranking]
    rows = [
        {
            "id": str(row["Sample ID"]),
            "phrase": str(row["Phrase"]),
            "reason": selection_reason.get(int(ranking[position]), "combined_relevance_risk"),
            "candidate_score": round(float(selected_scores[position]), 4),
            "relevant_support": round(float(selected_support[position]), 4),
            "garbage_model_probability": round(float(selected_garbage[position]), 4),
        }
        for position, (_, row) in enumerate(selected.iterrows())
    ]
    pending_ids = [row["id"] for row in rows]
    current_state = read_state(job_dir)
    write_state(
        job_dir,
        "relevance_audit",
        pending_relevance_review_ids=pending_ids,
        relevance_audit_garbage_before=int(labels.eq("garbage").sum()),
        labels_reviewed=int(labels.notna().sum()),
        remaining_review_rows=int(labels.isna().sum()),
        relevance_audit_batches=int(current_state.get("relevance_audit_batches", 0) or 0),
    )
    return {
        "stage": "relevance_audit",
        "labeling_contract": INTENT_LABELING_CONTRACT,
        "selection": "out_of_topic_candidates",
        "rows": rows,
        "remaining_unlabeled": int(len(unlabeled)),
    }


def select_intent_review(job_dir: Path, limit: int) -> dict[str, object]:
    """Select the closest commercial/informational boundary rows for model review."""
    if limit < 1 or limit > 50:
        raise ValueError("limit must be from 1 to 50.")
    frame = load_label_sheet(job_dir)
    labels, phrases = label_summary(frame)
    relevant_mask = labels.isin(["commercial", "informational"]) & phrases.ne("")
    labeled = frame[relevant_mask].copy()
    labeled_targets = labels.loc[labeled.index]
    unlabeled = frame[labels.isna() & phrases.ne("")].copy()
    if unlabeled.empty:
        return {"stage": "ready_to_run", "selection": "complete", "rows": []}
    if len(labeled) < 20 or not {"commercial", "informational"}.issubset(set(labeled_targets)):
        return sample_rows(job_dir, 0, limit, only_unlabeled=True) | {
            "selection": "intent_bootstrap",
            "reason": "Need both relevant intent classes before boundary selection.",
        }

    import numpy as np
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.linear_model import LogisticRegression

    vectorizer = TfidfVectorizer(
        analyzer="char_wb", ngram_range=(3, 5), min_df=1, max_features=20000, sublinear_tf=True
    )
    train = vectorizer.fit_transform(labeled["Phrase"].astype(str))
    candidates = vectorizer.transform(unlabeled["Phrase"].astype(str))
    classifier = LogisticRegression(max_iter=1000, class_weight="balanced", random_state=42).fit(
        train, labeled_targets
    )
    probabilities = classifier.predict_proba(candidates)
    ordered = np.sort(probabilities, axis=1)
    margin = ordered[:, -1] - ordered[:, -2]
    priority = pd.to_numeric(unlabeled.get("Search Volume", 0), errors="coerce").fillna(0).to_numpy()
    word_count = unlabeled["Phrase"].astype(str).str.split().str.len().to_numpy()
    ranking = np.lexsort((word_count, -priority, margin))[:limit]
    selected = unlabeled.iloc[ranking].copy()
    selected_margin = margin[ranking]
    rows = [
        {
            "id": str(row["Sample ID"]),
            "phrase": str(row["Phrase"]),
            "reason": "commercial_informational_boundary",
            "model_margin": round(float(selected_margin[position]), 4),
        }
        for position, (_, row) in enumerate(selected.iterrows())
    ]
    pending_ids = [row["id"] for row in rows]
    current_state = read_state(job_dir)
    write_state(
        job_dir,
        "intent_audit",
        pending_intent_review_ids=pending_ids,
        labels_reviewed=int(labels.notna().sum()),
        remaining_review_rows=int(labels.isna().sum()),
        intent_audit_batches=int(current_state.get("intent_audit_batches", 0) or 0),
    )
    return {
        "stage": "intent_audit",
        "labeling_contract": INTENT_LABELING_CONTRACT,
        "selection": "commercial_informational_boundary",
        "rows": rows,
        "remaining_unlabeled": int(len(unlabeled)),
    }


def select_policy_conflict_review(job_dir: Path, limit: int) -> dict[str, object]:
    """Select signal conflicts plus weak-question contexts needing composition."""
    if limit < 1 or limit > 50:
        raise ValueError("limit must be from 1 to 50.")
    import numpy as np
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.linear_model import LogisticRegression

    frame = load_label_sheet(job_dir)
    labels, phrases = label_summary(frame)
    relevant_mask = labels.isin(["commercial", "informational"]) & phrases.ne("")
    labeled = frame[relevant_mask].copy()
    targets = labels.loc[labeled.index]
    unlabeled = frame[labels.isna() & phrases.ne("")].copy()
    if unlabeled.empty:
        return {"stage": "ready_to_run", "selection": "complete", "rows": []}
    if len(labeled) < 20 or not {"commercial", "informational"}.issubset(set(targets)):
        return sample_rows(job_dir, 0, limit, only_unlabeled=True) | {
            "selection": "policy_conflict_bootstrap",
            "reason": "Need both intent classes before policy-conflict selection.",
        }

    config = json.loads((job_dir / "job_config.json").read_text(encoding="utf-8"))
    commercial_signals = configured_values(config, "intent", "commercial_markers")
    informational_signals = configured_values(config, "intent", "informational_markers")
    intent_config = config.get("intent", {}) if isinstance(config.get("intent", {}), dict) else {}
    family_rules = (
        intent_config.get("family_rules", {})
        if isinstance(intent_config.get("family_rules", {}), dict)
        else {}
    )
    weak_question_signals = configured_values(config, "intent", "weak_question_markers")
    commercial_families = [
        str(value)
        for value in family_rules.get("commercial", [])
        if not family_overlaps_signals(str(value), weak_question_signals)
    ]
    informational_families = [
        str(value)
        for value in family_rules.get("informational", [])
        if not family_overlaps_signals(str(value), weak_question_signals)
    ]
    candidate_phrases = unlabeled["Phrase"].astype(str)
    commercial_hit = signal_mask(candidate_phrases, commercial_signals).to_numpy()
    informational_hit = signal_mask(candidate_phrases, informational_signals).to_numpy()
    weak_question_hit = signal_mask(candidate_phrases, weak_question_signals).to_numpy()
    commercial_family_hit = signal_mask(candidate_phrases, commercial_families).to_numpy()
    informational_family_hit = signal_mask(candidate_phrases, informational_families).to_numpy()

    vectorizer = TfidfVectorizer(
        analyzer="char_wb", ngram_range=(3, 5), min_df=1, max_features=20000, sublinear_tf=True
    )
    train = vectorizer.fit_transform(labeled["Phrase"].astype(str))
    candidates = vectorizer.transform(candidate_phrases)
    classifier = LogisticRegression(
        max_iter=1000, class_weight="balanced", random_state=42
    ).fit(train, targets)
    probabilities = classifier.predict_proba(candidates)
    predicted = classifier.classes_[np.argmax(probabilities, axis=1)]
    confidence = np.max(probabilities, axis=1)
    ordered = np.sort(probabilities, axis=1)
    margin = ordered[:, -1] - ordered[:, -2]
    signal_conflict = (
        (informational_hit & ~commercial_hit & (predicted == "commercial"))
        | (commercial_hit & ~informational_hit & (predicted == "informational"))
        | (commercial_hit & informational_hit)
        | (informational_family_hit & (predicted == "commercial"))
        | (commercial_family_hit & (predicted == "informational"))
        | (informational_family_hit & commercial_family_hit)
        | (informational_family_hit & commercial_hit)
        | (commercial_family_hit & informational_hit)
    )
    weak_context = (
        weak_question_hit
        & ~commercial_hit
        & ~informational_hit
        & ~commercial_family_hit
        & ~informational_family_hit
    )
    weak_indices = sorted(
        (int(index) for index in np.flatnonzero(weak_context)),
        key=lambda index: (predicted[index] != "commercial", -confidence[index]),
    )
    weak_budget = min(len(weak_indices), limit, max(10, limit // 2))
    chosen = weak_indices[:weak_budget]
    conflict_indices = np.flatnonzero(signal_conflict)
    conflict_indices = conflict_indices[
        np.argsort(-confidence[conflict_indices], kind="stable")
    ]
    chosen_set = set(chosen)
    chosen.extend(
        int(index)
        for index in conflict_indices
        if int(index) not in chosen_set and len(chosen) < limit
    )
    if len(chosen) < limit:
        boundary_indices = np.argsort(margin, kind="stable")
        chosen_set = set(chosen)
        chosen.extend(
            int(index)
            for index in boundary_indices
            if int(index) not in chosen_set
            and len(chosen) < limit
        )
    selected = unlabeled.iloc[chosen].copy()
    rows = []
    for position, (_, row) in enumerate(selected.iterrows()):
        candidate_index = chosen[position]
        if weak_context[candidate_index]:
            reason = "weak_question_context_audit"
        elif commercial_family_hit[candidate_index] and informational_family_hit[candidate_index]:
            reason = "opposite_family_conflict"
        elif (
            informational_family_hit[candidate_index] and commercial_hit[candidate_index]
        ) or (
            commercial_family_hit[candidate_index] and informational_hit[candidate_index]
        ):
            reason = "family_strong_signal_conflict"
        elif (
            informational_family_hit[candidate_index] and predicted[candidate_index] == "commercial"
        ) or (
            commercial_family_hit[candidate_index] and predicted[candidate_index] == "informational"
        ):
            reason = "family_classifier_disagreement"
        elif commercial_hit[candidate_index] and informational_hit[candidate_index]:
            reason = "strong_signal_conflict"
        elif signal_conflict[candidate_index]:
            reason = "signal_classifier_disagreement"
        else:
            reason = "intent_boundary_fallback"
        rows.append(
            {
                "id": str(row["Sample ID"]),
                "phrase": str(row["Phrase"]),
                "reason": reason,
                "classifier_prediction": str(predicted[candidate_index]),
                "classifier_confidence": round(float(confidence[candidate_index]), 4),
                "commercial_signal": bool(commercial_hit[candidate_index]),
                "informational_signal": bool(informational_hit[candidate_index]),
                "weak_question_signal": bool(weak_question_hit[candidate_index]),
                "commercial_family": bool(commercial_family_hit[candidate_index]),
                "informational_family": bool(informational_family_hit[candidate_index]),
            }
        )
    pending_ids = [row["id"] for row in rows]
    current_state = read_state(job_dir)
    write_state(
        job_dir,
        "policy_conflict_audit",
        pending_policy_conflict_review_ids=pending_ids,
        labels_reviewed=int(labels.notna().sum()),
        remaining_review_rows=int(labels.isna().sum()),
        policy_conflict_audit_batches=int(
            current_state.get("policy_conflict_audit_batches", 0) or 0
        ),
    )
    return {
        "stage": "policy_conflict_audit",
        "labeling_contract": INTENT_LABELING_CONTRACT,
        "selection": "signal_family_and_weak_question_conflicts",
        "rows": rows,
        "remaining_unlabeled": int(len(unlabeled)),
    }


def cluster_relevance_paths(job_dir: Path) -> tuple[Path, Path, Path]:
    state = read_state(job_dir)
    output_value = str(state.get("cluster_review_output_dir", "")).strip()
    if not output_value:
        raise ValueError("Cluster relevance output directory is not recorded in state.json.")
    output_dir = Path(output_value)
    return (
        output_dir / "cluster_relevance_audit.json",
        output_dir / "cluster_relevance_labels.json",
        output_dir,
    )


def cluster_relevance_status(job_dir: Path) -> dict[str, object]:
    audit_path, labels_path, _ = cluster_relevance_paths(job_dir)
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    clusters = audit.get("clusters", []) if isinstance(audit, dict) else []
    decisions: dict[str, str] = {}
    if labels_path.is_file():
        loaded = json.loads(labels_path.read_text(encoding="utf-8"))
        if isinstance(loaded, dict) and isinstance(loaded.get("labels"), dict):
            decisions = {str(key): str(value) for key, value in loaded["labels"].items()}
    pending = []
    total_decisions = 0
    labeled_decisions = 0
    for source in clusters:
        required_ids = [str(source.get("id"))]
        required_ids.extend(
            str(item["review_id"])
            for item in source.get("representative_evidence", [])
            if isinstance(item, dict) and item.get("review_id")
        )
        total_decisions += len(required_ids)
        labeled_decisions += sum(decision_id in decisions for decision_id in required_ids)
        missing_ids = [decision_id for decision_id in required_ids if decision_id not in decisions]
        if missing_ids:
            record = dict(source)
            record["missing_decision_ids"] = missing_ids
            pending.append(record)
    return {
        "total_clusters": len(clusters),
        "labeled_clusters": len(clusters) - len(pending),
        "remaining_clusters": len(pending),
        "total_decisions": total_decisions,
        "labeled_decisions": labeled_decisions,
        "remaining_decisions": total_decisions - labeled_decisions,
        "pending": pending,
    }


def cluster_relevance_batch(job_dir: Path, limit: int) -> dict[str, object]:
    if limit < 1 or limit > 30:
        raise ValueError("limit must be from 1 to 30.")
    status = cluster_relevance_status(job_dir)
    rows = []
    for record in list(status["pending"])[:limit]:
        evidence = list(
            record.get("representative_evidence")
            or [
                {"role": "unspecified", "phrase": phrase}
                for phrase in record.get("representatives", [])
            ]
        )[:5]
        rows.append(
            {
                "id": str(record["id"]),
                "current_intent": str(record.get("current_intent", "")),
                "cluster": str(record.get("cluster", "")),
                "rows": int(record.get("rows", 0)),
                "average_garbage_probability": record.get("average_garbage_probability", 0),
                "representatives": evidence,
                "required_decisions": list(record.get("missing_decision_ids", [])),
            }
        )
    return {
        "stage": "cluster_relevance_review",
        "topic": read_state(job_dir).get("topic", ""),
        "instruction": (
            "Judge topic membership, not commercial intent. Use relevant only when the cluster belongs to the "
            "user's stated business scope; garbage for an adjacent or unrelated domain even if brand words collide; "
            "mixed only when the representatives genuinely contain both. Evidence roles distinguish central, "
            "high-volume, relevance-risk, and boundary phrases; do not let one boundary outlier override a clearly "
            "relevant cluster, but use mixed when the central evidence itself spans both scopes. Return a cluster "
            "decision for the RC ID and an individual relevant/garbage decision for every representative review_id."
        ),
        "rows": rows,
        "remaining_after_batch": int(status["remaining_clusters"]) - len(rows),
    }


def apply_cluster_labels_inline(job_dir: Path, value: str) -> dict[str, object]:
    audit_path, labels_path, _ = cluster_relevance_paths(job_dir)
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    cluster_ids = {str(record.get("id")) for record in audit.get("clusters", [])}
    review_ids = {
        str(item["review_id"])
        for record in audit.get("clusters", [])
        for item in record.get("representative_evidence", [])
        if isinstance(item, dict) and item.get("review_id")
    }
    valid_ids = cluster_ids | review_ids
    existing: dict[str, str] = {}
    if labels_path.is_file():
        loaded = json.loads(labels_path.read_text(encoding="utf-8"))
        if isinstance(loaded, dict) and isinstance(loaded.get("labels"), dict):
            existing = {str(key): str(label) for key, label in loaded["labels"].items()}
    applied = 0
    for item in value.split(";"):
        fields = [field.strip() for field in item.split("|")]
        if not item.strip():
            continue
        if len(fields) != 2:
            raise ValueError(
                "Each decision must be ID|label. Cluster IDs allow relevant, garbage, or mixed; "
                "representative review IDs allow relevant or garbage."
            )
        cluster_id, label = fields[0], fields[1].lower()
        if cluster_id not in valid_ids:
            raise ValueError(f"Unknown cluster relevance ID: {cluster_id}")
        allowed = {"relevant", "garbage", "mixed"} if cluster_id in cluster_ids else {"relevant", "garbage"}
        if label not in allowed:
            raise ValueError(f"Invalid cluster relevance label: {label}")
        existing[cluster_id] = label
        applied += 1
    if not applied:
        raise ValueError("At least one cluster relevance label is required.")
    temporary = labels_path.with_suffix(".tmp")
    temporary.write_text(
        json.dumps({"labels": existing}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    temporary.replace(labels_path)
    status = cluster_relevance_status(job_dir)
    write_state(
        job_dir,
        "cluster_relevance_review",
        cluster_relevance_labeled=status["labeled_clusters"],
        cluster_relevance_remaining=status["remaining_clusters"],
    )
    return {
        "status": "cluster_labels_applied",
        "applied": applied,
        "labeled_clusters": status["labeled_clusters"],
        "remaining_clusters": status["remaining_clusters"],
        "labeled_decisions": status["labeled_decisions"],
        "remaining_decisions": status["remaining_decisions"],
    }


def apply_cluster_relevance_decisions(job_dir: Path) -> dict[str, object]:
    lock, existing_lock = acquire_run_lock(job_dir)
    if lock is None:
        return {
            "status": "already_running",
            "pid": existing_lock.get("pid") if existing_lock else None,
        }
    _, _, output_dir = cluster_relevance_paths(job_dir)
    from seo_pipeline import apply_large_cluster_relevance_decisions

    write_state(
        job_dir,
        "cluster_relevance_finalize",
        autopilot_last_status="running",
        run_pid=lock["pid"],
        run_started_at=lock["started_at"],
    )
    try:
        quiet_library_logs()
        manifest = apply_large_cluster_relevance_decisions(
            output_dir, job_dir / "job_config.json"
        )
    except Exception as error:
        record_run_failure(job_dir, error)
        release_run_lock(job_dir, str(lock["token"]))
        raise
    state = write_state(
        job_dir,
        "quality_review",
        last_large_run=manifest["excel_summary"],
        cluster_relevance_audit=manifest.get("cluster_relevance_audit", {}),
    )
    result = {
        "status": "completed",
        "output": manifest["excel_summary"],
        "rows": int(manifest.get("classified_rows_per_chunk_deduplicated", 0)),
        "intent_counts": manifest.get("intent_counts", {}),
        "cluster_relevance_audit": manifest.get("cluster_relevance_audit", {}),
        "intent_family_audit": manifest.get("intent_family_audit", {}),
        "stage": state["stage"],
    }
    release_run_lock(job_dir, str(lock["token"]))
    return result


def import_feedback(job_dir: Path, reviewed_workbook: Path) -> int:
    if not reviewed_workbook.is_file():
        raise FileNotFoundError(reviewed_workbook)
    review = pd.read_excel(reviewed_workbook, sheet_name="Human review")
    lookup = normalized_column_lookup(review)
    phrase_column = lookup.get("phrase")
    intent_column = lookup.get("correct intent")
    notes_column = lookup.get("reviewer notes")
    confidence_column = lookup.get("classification confidence")
    if not phrase_column or not intent_column:
        raise ValueError("Human review must contain Phrase and Correct Intent columns.")

    corrections = pd.DataFrame(
        {
            "Phrase": review[phrase_column].astype(str).str.strip(),
            "Model Label": review[intent_column].fillna("").astype(str).str.strip(),
            "Model Confidence": (
                review[confidence_column] if confidence_column else ""
            ),
            "Model Notes": review[notes_column].fillna("").astype(str) if notes_column else "",
        }
    )
    corrections["_normalized_label"] = (
        corrections["Model Label"].str.lower().map(LABEL_ALIASES)
    )
    corrections = corrections[
        corrections["_normalized_label"].notna() & corrections["Phrase"].ne("")
    ].copy()
    corrections["Model Label"] = corrections["_normalized_label"]
    corrections = corrections.drop(columns="_normalized_label")
    if corrections.empty:
        return 0

    labels_path = job_dir / "model_labels.xlsx"
    existing = load_label_sheet(job_dir)
    for column in ("Model Label", "Model Confidence", "Model Notes"):
        if column not in existing.columns:
            existing[column] = ""
    combined = pd.concat([existing, corrections], ignore_index=True)
    combined["Sample ID"] = [f"row-{index:06d}" for index in range(1, len(combined) + 1)]
    combined["_phrase_key"] = combined["Phrase"].astype(str).str.strip().str.lower()
    combined = combined.drop_duplicates("_phrase_key", keep="last").drop(columns="_phrase_key")
    combined.to_excel(labels_path, sheet_name="Model labels", index=False)
    return len(corrections)


def resolve_reviewed_workbook(value: Path) -> Path:
    """Accept an explicit review path or a filename in files/ or outputs/."""
    if value.is_file():
        return value.resolve()
    for directory in (PROJECT_DIR / "files", OUTPUT_DIR):
        candidate = directory / value.name
        if candidate.is_file():
            return candidate.resolve()
    raise FileNotFoundError(
        f"Reviewed workbook not found: {value}. Put it in {PROJECT_DIR / 'files'} or provide its full path."
    )


def workflow_job_id_from_workbook(reviewed_workbook: Path) -> str:
    with pd.ExcelFile(reviewed_workbook) as workbook:
        sheet_names = set(workbook.sheet_names)
    metadata_sheet = next(
        (name for name in ("_Workflow", "Run configuration") if name in sheet_names),
        None,
    )
    if metadata_sheet is None:
        raise ValueError(
            "Reviewed workbook has no _Workflow or Run configuration sheet with Workflow Job ID."
        )
    configuration = pd.read_excel(reviewed_workbook, sheet_name=metadata_sheet)
    lookup = normalized_column_lookup(configuration)
    parameter_column = lookup.get("parameter")
    value_column = lookup.get("value")
    if not parameter_column or not value_column:
        raise ValueError(f"{metadata_sheet} must contain Parameter and Value columns.")
    # The generated configuration sheet uses the Python key spelling
    # ``workflow_job_id``. Treat underscores and spaces alike so this stays
    # compatible with both the machine-readable key and the human label.
    parameters = configuration[parameter_column].fillna("").astype(str).map(
        lambda value: re.sub(r"[_\s]+", " ", value).strip().lower()
    )
    rows = configuration[parameters.eq("workflow job id")]
    if len(rows) != 1:
        raise ValueError("Reviewed workbook must contain exactly one Workflow Job ID.")
    workflow_job_id = str(rows.iloc[0][value_column]).strip()
    if not re.fullmatch(r"seo-[a-f0-9]{16}", workflow_job_id):
        raise ValueError("Workflow Job ID is missing or invalid; no import was made.")
    return workflow_job_id


def reviewed_corrections(reviewed_workbook: Path) -> pd.DataFrame:
    if not reviewed_workbook.is_file():
        return pd.DataFrame(columns=["Phrase", "Label", "Confidence"])
    review = pd.read_excel(reviewed_workbook, sheet_name="Human review")
    lookup = normalized_column_lookup(review)
    phrase_column = lookup.get("phrase")
    intent_column = lookup.get("correct intent")
    if not phrase_column or not intent_column:
        return pd.DataFrame(columns=["Phrase", "Label", "Confidence"])
    result = pd.DataFrame(
        {
            "Phrase": review[phrase_column].astype(str).str.strip(),
            "Label": review[intent_column].fillna("").astype(str).str.strip().str.lower().map(LABEL_ALIASES),
            "Confidence": 1.0,
        }
    )
    return result[result["Label"].notna() & result["Phrase"].ne("")]


def learn_job(job_dir: Path, reviewed_workbook: Path | None) -> dict[str, object]:
    config_path = job_dir / "job_config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    description = str(config["topic"]["description"])
    reused = config.get("knowledge_reuse", {}).get("topic_key")
    key = reused or topic_key(description)
    save_topic_profile(key, description, config)

    labels_path = job_dir / "model_labels.xlsx"
    labels = pd.read_excel(labels_path, sheet_name="Model labels")
    lookup = normalized_column_lookup(labels)
    phrase_column = lookup.get("phrase")
    label_column = lookup.get("model label")
    confidence_column = lookup.get("model confidence")
    if not phrase_column or not label_column:
        raise ValueError("model_labels.xlsx must contain Phrase and Model Label.")
    examples = pd.DataFrame(
        {
            "Phrase": labels[phrase_column].astype(str).str.strip(),
            "Label": labels[label_column].fillna("").astype(str).str.strip(),
            "Confidence": (
                labels[confidence_column] if confidence_column else 0.8
            ),
        }
    )
    learned = store_examples(
        key, examples, job_dir.name, "model_reviewed", default_confidence=0.8
    )
    corrected = 0
    if reviewed_workbook:
        corrections = reviewed_corrections(reviewed_workbook)
        corrected = store_examples(
            key,
            corrections,
            job_dir.name,
            "review_corrected",
            default_confidence=1.0,
        )
    return {
        "topic_key": key,
        "model_reviewed_examples_processed": learned,
        "review_corrected_examples_processed": corrected,
        "knowledge": knowledge_status(key),
    }


def apply_review_and_learn(reviewed_value: Path) -> dict[str, object]:
    """Import a human review and persist knowledge without recalculating SEO output."""
    reviewed_workbook = resolve_reviewed_workbook(reviewed_value)
    workflow_job_id = workflow_job_id_from_workbook(reviewed_workbook)
    job_dir = find_job_by_workflow_id(workflow_job_id)
    imported = import_feedback(job_dir, reviewed_workbook)
    learned = learn_job(job_dir, reviewed_workbook)
    state = write_state(
        job_dir,
        "knowledge_saved",
        feedback_imported=imported,
        knowledge_topic_key=learned["topic_key"],
        reviewed_workbook=str(reviewed_workbook),
    )
    return {
        "status": "knowledge_saved",
        "recalculated": False,
        "workflow_job_id": workflow_job_id,
        "feedback_imported": imported,
        "reviewed_workbook": str(reviewed_workbook),
        "model_reviewed_examples_processed": learned["model_reviewed_examples_processed"],
        "review_corrected_examples_processed": learned["review_corrected_examples_processed"],
        # Keep the terminal response ASCII and compact. The complete topic
        # metadata remains in the local knowledge database and state.json.
        "stage": state["stage"],
    }


def cleanup_job(job_dir: Path) -> list[str]:
    removed: list[str] = []
    safe_targets = [job_dir / "tmp", job_dir / "cache"]
    for target in safe_targets:
        if target.is_dir():
            shutil.rmtree(target)
            removed.append(str(target))
    cache_targets = [PROJECT_DIR / "__pycache__", *job_dir.rglob("__pycache__")]
    for cache in cache_targets:
        if cache.is_dir():
            shutil.rmtree(cache)
            removed.append(str(cache))
    for temporary in job_dir.rglob("*"):
        if temporary.is_file() and (
            temporary.suffix.lower() in {".tmp", ".part"}
            or temporary.name.startswith("~$")
        ):
            temporary.unlink()
            removed.append(str(temporary))
    lock_path = job_dir / RUN_LOCK_NAME
    if lock_path.is_file():
        lock_data = read_run_lock(lock_path)
        if not lock_process_is_active(lock_data):
            lock_path.unlink()
            removed.append(str(lock_path))
    return removed


def read_run_lock(lock_path: Path) -> dict[str, object]:
    try:
        data = json.loads(lock_path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def lock_process_is_active(lock_data: dict[str, object]) -> bool:
    """Return True only when a lock owner still appears to be running."""
    try:
        process_id = int(lock_data.get("pid", 0))
    except (TypeError, ValueError):
        return False
    if process_id <= 0:
        return False
    if os.name == "nt":
        # Python 3.14 on Windows may turn ``os.kill(pid, 0)`` with access
        # denied into a SystemError.  Query the process handle directly so a
        # lock check never breaks `next` while a worker is running.
        import ctypes

        process_query_limited_information = 0x1000
        still_active = 259
        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        handle = kernel32.OpenProcess(process_query_limited_information, False, process_id)
        if not handle:
            error_code = ctypes.get_last_error()
            # Access denied means that Windows knows the process exists but
            # this interpreter may not inspect it (for example, an elevated
            # worker launched by Hermes).
            return error_code == 5
        try:
            exit_code = ctypes.c_ulong()
            if not kernel32.GetExitCodeProcess(handle, ctypes.byref(exit_code)):
                return False
            return exit_code.value == still_active
        finally:
            kernel32.CloseHandle(handle)
    try:
        os.kill(process_id, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    except OSError:
        return False
    return True


def acquire_run_lock(job_dir: Path) -> tuple[dict[str, object] | None, dict[str, object] | None]:
    """Atomically claim the one permitted final run for a job."""
    lock_path = job_dir / RUN_LOCK_NAME
    token = uuid.uuid4().hex
    mine = {
        "token": token,
        "pid": os.getpid(),
        "started_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    }
    for _ in range(2):
        try:
            descriptor = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            existing = read_run_lock(lock_path)
            if lock_process_is_active(existing):
                return None, existing
            try:
                lock_path.unlink()
            except FileNotFoundError:
                continue
            continue
        try:
            os.write(descriptor, json.dumps(mine, ensure_ascii=False).encode("utf-8"))
        finally:
            os.close(descriptor)
        return mine, None
    return None, read_run_lock(lock_path)


def release_run_lock(job_dir: Path, token: str) -> None:
    """Release only the lock created by this exact workflow process."""
    lock_path = job_dir / RUN_LOCK_NAME
    if read_run_lock(lock_path).get("token") == token:
        try:
            lock_path.unlink()
        except FileNotFoundError:
            pass


def next_output_path(job_dir: Path, input_value: str) -> Path:
    input_path = resolve_input(input_value)
    job_output_dir = OUTPUT_DIR / job_dir.name
    job_output_dir.mkdir(parents=True, exist_ok=True)
    existing = list(job_output_dir.glob(f"{input_path.stem}_run_*.xlsx"))
    numbers = []
    for path in existing:
        match = re.search(r"_run_(\d+)$", path.stem)
        if match:
            numbers.append(int(match.group(1)))
    return job_output_dir / f"{input_path.stem}_run_{max(numbers, default=0) + 1:03d}.xlsx"


def latest_run_path(job_dir: Path, input_value: str) -> Path:
    input_path = resolve_input(input_value)
    job_output_dir = OUTPUT_DIR / job_dir.name
    candidates: list[tuple[int, Path]] = []
    for path in job_output_dir.glob(f"{input_path.stem}_run_*.xlsx"):
        match = re.search(r"_run_(\d+)$", path.stem)
        if match:
            candidates.append((int(match.group(1)), path))
    if not candidates:
        raise FileNotFoundError(
            f"No versioned run found for {input_path.name} in {job_output_dir}"
        )
    return max(candidates, key=lambda item: item[0])[1]


def finalize_workbook(
    job_dir: Path,
    input_value: str,
    source: Path | None = None,
) -> dict[str, object]:
    input_path = resolve_input(input_value)
    job_output_dir = OUTPUT_DIR / job_dir.name
    job_output_dir.mkdir(parents=True, exist_ok=True)

    source_path = source.resolve() if source else latest_run_path(job_dir, input_value)
    if not source_path.is_file():
        raise FileNotFoundError(source_path)
    if source_path.suffix.lower() != ".xlsx":
        raise ValueError("The final source must be an .xlsx workbook.")

    with pd.ExcelFile(source_path) as workbook:
        available_sheets = set(workbook.sheet_names)
    missing_sheets = sorted(REQUIRED_RESULT_SHEETS - available_sheets)
    if missing_sheets:
        raise ValueError(
            "Cannot finalize an incomplete result workbook. Missing sheets: "
            + ", ".join(missing_sheets)
        )
    if not ({"Cluster summary", "Cluster review"} & available_sheets):
        raise ValueError(
            "Cannot finalize an incomplete result workbook. Missing Cluster summary."
        )
    if not ({"_Workflow", "Run configuration"} & available_sheets):
        raise ValueError(
            "Cannot finalize an incomplete result workbook. Missing workflow metadata."
        )

    final_path = job_output_dir / f"{input_path.stem}_FINAL.xlsx"
    temporary_path = job_output_dir / f".{input_path.stem}_FINAL.tmp"
    replaced_previous_final = final_path.exists()
    shutil.copy2(source_path, temporary_path)
    temporary_path.replace(final_path)
    return {
        "input": str(input_path.resolve()),
        "source": str(source_path.resolve()),
        "final": str(final_path.resolve()),
        "source_preserved": True,
        "replaced_previous_final": replaced_previous_final,
    }


def record_run_failure(job_dir: Path, error: Exception) -> None:
    """Make deterministic pipeline failures terminal for the Hermes autopilot."""
    write_state(
        job_dir,
        "blocked",
        autopilot_last_status="blocked",
        run_error=str(error),
    )


def run_with_log(job_dir: Path, input_value: str, output_path: Path | None) -> dict[str, object]:
    lock, existing_lock = acquire_run_lock(job_dir)
    if lock is None:
        return {
            "status": "already_running",
            "job": str(job_dir.resolve()),
            "pid": existing_lock.get("pid") if existing_lock else None,
            "started_at": existing_lock.get("started_at") if existing_lock else None,
        }
    logs = job_dir / "logs"
    logs.mkdir(parents=True, exist_ok=True)
    log_path = logs / "run.log"
    started = time.perf_counter()
    write_state(
        job_dir,
        "run",
        autopilot_last_status="running",
        run_pid=lock["pid"],
        run_started_at=lock["started_at"],
        run_mode="standard",
    )
    try:
        try:
            ensure_workflow_job_id(job_dir)
            # Importing the pipeline initializes torch/transformers/UMAP.  Keep it
            # out of all short control commands used during agent labelling.
            from seo_pipeline import run_pipeline

            quiet_library_logs()
            with log_path.open("a", encoding="utf-8") as log, contextlib.redirect_stdout(log), contextlib.redirect_stderr(log):
                print(f"[{time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())}] run started")
                output = run_pipeline(
                    input_value,
                    job_dir / "job_config.json",
                    job_dir / "model_labels.xlsx",
                    output_path,
                )
                print(f"[{time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())}] run completed")
        except Exception:
            with log_path.open("a", encoding="utf-8") as log:
                traceback.print_exc(file=log)
            raise
    except Exception as error:
        record_run_failure(job_dir, error)
        release_run_lock(job_dir, str(lock["token"]))
        raise
    duration = round(time.perf_counter() - started, 3)
    inspection = json.loads((job_dir / "inspection.json").read_text(encoding="utf-8"))
    state = write_state(
        job_dir,
        "quality_review",
        last_run=output.name,
        last_run_path=str(output.resolve()),
        duration_seconds=duration,
        labels_reviewed=job_status(job_dir).get("label_quality", {}).get("valid_labeled_rows", 0),
    )
    result = {
        "status": "completed",
        "output": str(output.resolve()),
        "rows": int(inspection.get("unique_normalized_phrases", 0)),
        "duration_seconds": duration,
        "log": str(log_path.resolve()),
        "stage": state["stage"],
    }
    release_run_lock(job_dir, str(lock["token"]))
    return result


def stage_large_input(job_dir: Path, input_value: str) -> tuple[str, Path | None]:
    """Create a workflow-owned TSV only when a large XLSX needs streaming.

    The language model never converts user files.  This short-lived staging
    file contains only the configured phrase and optional frequency columns
    and is removed by ``run_large_with_log`` in every exit path.
    """
    input_path = resolve_input(input_value)
    if input_path.suffix.lower() in {".csv", ".tsv"}:
        return str(input_path), None
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("Large mode supports CSV, TSV, and XLSX inputs.")
    config = json.loads((job_dir / "job_config.json").read_text(encoding="utf-8"))
    requested_phrase = str(config.get("phrase_column") or "").strip()
    requested_frequency = str(config.get("frequency_column") or "").strip()
    sheet_ref = config.get("source_sheet", 0)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_ref] if isinstance(sheet_ref, str) else workbook.worksheets[int(sheet_ref)]
        rows = sheet.iter_rows(values_only=True)
        header_values = next(rows, None)
        if not header_values:
            raise ValueError("The XLSX source sheet has no header row.")
        headers = ["" if value is None else str(value).strip() for value in header_values]
        normalized = {re.sub(r"\s+", " ", value.lower()): index for index, value in enumerate(headers)}
        phrase_index = normalized.get(re.sub(r"\s+", " ", requested_phrase.lower()))
        if phrase_index is None:
            raise ValueError(f"Phrase column not found in XLSX staging: {requested_phrase}")
        frequency_index = normalized.get(re.sub(r"\s+", " ", requested_frequency.lower())) if requested_frequency else None
        staging = job_dir / "tmp" / f"large-input-{uuid.uuid4().hex}.tsv"
        staging.parent.mkdir(parents=True, exist_ok=True)
        with staging.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, delimiter="\t")
            output_header = [headers[phrase_index]]
            if frequency_index is not None:
                output_header.append(headers[frequency_index])
            writer.writerow(output_header)
            for values in rows:
                phrase = values[phrase_index] if phrase_index < len(values) else None
                if phrase is None or not str(phrase).strip():
                    continue
                output_row = [phrase]
                if frequency_index is not None:
                    output_row.append(values[frequency_index] if frequency_index < len(values) else "")
                writer.writerow(output_row)
        return str(staging), staging
    finally:
        workbook.close()


def run_large_with_log(job_dir: Path, input_value: str, chunk_size: int) -> dict[str, object]:
    lock, existing_lock = acquire_run_lock(job_dir)
    if lock is None:
        return {
            "status": "already_running",
            "job": str(job_dir.resolve()),
            "pid": existing_lock.get("pid") if existing_lock else None,
            "started_at": existing_lock.get("started_at") if existing_lock else None,
        }
    logs = job_dir / "logs"
    logs.mkdir(parents=True, exist_ok=True)
    log_path = logs / "run_large.log"
    output_dir = OUTPUT_DIR / job_dir.name / "large" / Path(input_value).stem
    started = time.perf_counter()
    staged_path: Path | None = None
    write_state(
        job_dir,
        "run",
        autopilot_last_status="running",
        run_pid=lock["pid"],
        run_started_at=lock["started_at"],
        run_mode="large",
    )
    try:
        try:
            ensure_workflow_job_id(job_dir)
            from seo_pipeline import run_large_pipeline

            staged_input, staged_path = stage_large_input(job_dir, input_value)
            quiet_library_logs()
            with log_path.open("a", encoding="utf-8") as log, contextlib.redirect_stdout(log), contextlib.redirect_stderr(log):
                print(f"[{time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())}] large run started")
                manifest = run_large_pipeline(staged_input, job_dir / "job_config.json", job_dir / "model_labels.xlsx", output_dir, chunk_size, source_name=Path(input_value).name)
                print(f"[{time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())}] large run completed")
        except Exception:
            with log_path.open("a", encoding="utf-8") as log:
                traceback.print_exc(file=log)
            raise
    except Exception as error:
        record_run_failure(job_dir, error)
        release_run_lock(job_dir, str(lock["token"]))
        raise
    finally:
        if staged_path and staged_path.is_file():
            staged_path.unlink()
    duration = round(time.perf_counter() - started, 3)
    if manifest.get("status") == "awaiting_cluster_relevance_review":
        state = write_state(
            job_dir,
            "cluster_relevance_review",
            cluster_review_output_dir=str(output_dir.resolve()),
            cluster_relevance_labeled=0,
            cluster_relevance_remaining=int(
                manifest.get("cluster_relevance_audit", {}).get("clusters", 0)
            ),
            duration_seconds=duration,
        )
        result = {
            "status": "awaiting_cluster_relevance_review",
            "clusters": state["cluster_relevance_remaining"],
            "rows": manifest["classified_rows_per_chunk_deduplicated"],
            "duration_seconds": duration,
            "log": str(log_path.resolve()),
            "stage": state["stage"],
        }
    else:
        state = write_state(
            job_dir,
            "quality_review",
            last_large_run=manifest["excel_summary"],
            duration_seconds=duration,
        )
        result = {
            "status": "completed",
            "output": manifest["excel_summary"],
            "rows": manifest["classified_rows_per_chunk_deduplicated"],
            "duration_seconds": duration,
            "log": str(log_path.resolve()),
            "stage": state["stage"],
        }
    release_run_lock(job_dir, str(lock["token"]))
    return result


def run_auto(job_dir: Path, input_value: str, threshold: int, chunk_size: int) -> dict[str, object]:
    if threshold < 1:
        raise ValueError("large-threshold must be at least 1.")
    inspection_path = job_dir / "inspection.json"
    if not inspection_path.is_file():
        raise FileNotFoundError("inspection.json is required for run-auto.")
    inspection = json.loads(inspection_path.read_text(encoding="utf-8"))
    rows = int(inspection.get("unique_normalized_phrases", 0))
    input_path = resolve_input(input_value)
    if rows < threshold:
        result = run_with_log(job_dir, input_value, None)
        result["mode"] = "standard"
        result["large_threshold"] = threshold
        return result
    if input_path.suffix.lower() not in {".csv", ".tsv", ".xlsx"}:
        raise ValueError(
            f"run-auto selected large mode for {rows} phrases, but the input format "
            f"{input_path.suffix or '<none>'} is unsupported. Use CSV, TSV, or XLSX."
        )
    result = run_large_with_log(job_dir, input_value, chunk_size)
    result["mode"] = "large"
    result["large_threshold"] = threshold
    return result


def export_completed_large_result(job_dir: Path, input_value: str) -> dict[str, object]:
    """Build the human review workbook from an already completed large run.

    This deliberately reads only the classified part files.  It never starts
    embeddings, classification, or clustering again.
    """
    input_path = resolve_input(input_value)
    output_dir = OUTPUT_DIR / job_dir.name / "large" / input_path.stem
    manifest_path = output_dir / "manifest.json"
    if not manifest_path.is_file():
        raise FileNotFoundError(f"Completed large-run manifest not found: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("status") != "completed":
        raise ValueError("Large run is not completed; export was not started.")
    parts_dir = output_dir / "parts"
    if not list(parts_dir.glob("part-*.csv.gz")):
        raise FileNotFoundError("Completed large-run parts are missing.")
    from seo_pipeline import build_intent_family_audit, export_large_result_workbook

    summary_path = output_dir / "cluster_summary.csv"
    uncertain_path = output_dir / "uncertain_review.csv"
    summary = pd.read_csv(summary_path, encoding="utf-8-sig") if summary_path.is_file() else pd.DataFrame()
    uncertain = pd.read_csv(uncertain_path, encoding="utf-8-sig") if uncertain_path.is_file() else pd.DataFrame()
    config = json.loads((job_dir / "job_config.json").read_text(encoding="utf-8"))
    family_audit_path = output_dir / "intent_family_audit.csv"
    family_audit = (
        pd.read_csv(family_audit_path, encoding="utf-8-sig")
        if family_audit_path.is_file()
        else build_intent_family_audit(
            (
                pd.read_csv(part, compression="gzip", usecols=["Phrase", "Intent"])
                for part in sorted(parts_dir.glob("part-*.csv.gz"))
            ),
            config,
        )
    )
    result_path = output_dir / f"{input_path.stem}_clustered.xlsx"
    counts = export_large_result_workbook(
        parts_dir, result_path, summary, uncertain, family_audit, config, manifest
    )
    manifest["excel_result"] = str(result_path.resolve())
    manifest["excel_summary"] = str(result_path.resolve())
    manifest["intent_counts"] = counts
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    write_state(job_dir, "quality_review", last_large_run=str(result_path.resolve()))
    return {
        "status": "completed",
        "recalculated": False,
        "output": str(result_path.resolve()),
        "rows": int(manifest.get("classified_rows_per_chunk_deduplicated", 0)),
        "intent_counts": counts,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Orchestrate the model-directed SEO workflow.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    prepare = subparsers.add_parser("prepare", help="Inspect input and create a representative labeling job.")
    prepare.add_argument("input")
    prepare.add_argument("--topic", required=True)
    prepare.add_argument("--sample-size", type=int, default=800)
    prepare.add_argument("--phrase-column")
    prepare.add_argument("--frequency-column")
    prepare.add_argument("--job-name")
    prepare.add_argument(
        "--reuse-topic",
        default="auto",
        help="Prior knowledge topic, auto for current topic, or none",
    )
    prepare.add_argument("--prior-limit", type=int, default=150)
    prepare.add_argument("--quiet", action="store_true")

    status = subparsers.add_parser("status", help="Show labeling and configuration readiness.")
    status.add_argument("--job", required=True)
    status.add_argument("--compact", action="store_true")
    status.add_argument("--quiet", action="store_true")

    next_step = subparsers.add_parser("next", help="Return the next safe action from the persistent job state.")
    next_step.add_argument("--job")
    next_step.add_argument("--input", help="Input filename used to safely locate an existing job or create the first prepare action.")
    next_step.add_argument("--topic", help="Required with a new input when no prepared job exists.")
    next_step.add_argument("--quiet", action="store_true")

    sample = subparsers.add_parser("sample", help="Return a small UTF-8 JSON labeling batch; never print XLSX rows.")
    sample.add_argument("--job", required=True)
    sample.add_argument("--offset", type=int, default=0)
    sample.add_argument("--limit", type=int, default=25)
    sample.add_argument("--include-labeled", action="store_true")
    sample.add_argument("--quiet", action="store_true")

    labels = subparsers.add_parser("apply-labels", help="Atomically import model labels from a job-local JSON batch.")
    labels.add_argument("--job", required=True)
    labels.add_argument("--input", type=Path, required=True)
    labels.add_argument("--quiet", action="store_true")

    inline_labels = subparsers.add_parser("apply-labels-inline", help="Import a compact phrase-free batch: SampleID|label|confidence;...")
    inline_labels.add_argument("--job", required=True)
    inline_labels.add_argument("--labels", required=True)
    inline_labels.add_argument("--quiet", action="store_true")

    family_review = subparsers.add_parser(
        "family-review",
        help="Return a compact batch of frequent lexical families for intent decisions.",
    )
    family_review.add_argument("--job", required=True)
    family_review.add_argument("--limit", type=int, default=INTENT_FAMILY_REVIEW_BATCH_SIZE)
    family_review.add_argument("--quiet", action="store_true")

    family_coverage_review = subparsers.add_parser(
        "family-coverage-review",
        help="Return mandatory real phrases that cover otherwise unrepresented source families.",
    )
    family_coverage_review.add_argument("--job", required=True)
    family_coverage_review.add_argument(
        "--limit", type=int, default=FAMILY_COVERAGE_REVIEW_BATCH_SIZE
    )
    family_coverage_review.add_argument("--quiet", action="store_true")

    family_labels = subparsers.add_parser(
        "apply-family-labels-inline",
        help="Apply compact intent-family decisions: ID|commercial/informational/neutral.",
    )
    family_labels.add_argument("--job", required=True)
    family_labels.add_argument("--labels", required=True)
    family_labels.add_argument("--quiet", action="store_true")

    review = subparsers.add_parser("review", help="Select an uncertainty-focused active-learning labeling batch.")
    review.add_argument("--job", required=True)
    review.add_argument("--limit", type=int, default=20)
    review.add_argument("--quiet", action="store_true")

    relevance_review = subparsers.add_parser(
        "relevance-review",
        help="Select likely out-of-topic candidates for a bounded relevance audit.",
    )
    relevance_review.add_argument("--job", required=True)
    relevance_review.add_argument("--limit", type=int, default=50)
    relevance_review.add_argument("--quiet", action="store_true")

    intent_review = subparsers.add_parser(
        "intent-review",
        help="Select commercial/informational boundary rows for a mandatory intent audit.",
    )
    intent_review.add_argument("--job", required=True)
    intent_review.add_argument("--limit", type=int, default=50)
    intent_review.add_argument("--quiet", action="store_true")

    policy_conflict_review = subparsers.add_parser(
        "policy-conflict-review",
        help="Select rows where strong intent signals and the learned classifier disagree.",
    )
    policy_conflict_review.add_argument("--job", required=True)
    policy_conflict_review.add_argument("--limit", type=int, default=50)
    policy_conflict_review.add_argument("--quiet", action="store_true")

    cluster_review = subparsers.add_parser(
        "cluster-review",
        help="Return a compact post-classification cluster relevance batch.",
    )
    cluster_review.add_argument("--job", required=True)
    cluster_review.add_argument("--limit", type=int, default=30)
    cluster_review.add_argument("--quiet", action="store_true")

    cluster_labels = subparsers.add_parser(
        "apply-cluster-labels-inline",
        help="Apply compact cluster and representative ID|label decisions.",
    )
    cluster_labels.add_argument("--job", required=True)
    cluster_labels.add_argument("--labels", required=True)
    cluster_labels.add_argument("--quiet", action="store_true")

    cluster_decisions = subparsers.add_parser(
        "apply-cluster-decisions",
        help="Apply completed cluster relevance decisions and export the final workbook.",
    )
    cluster_decisions.add_argument("--job", required=True)
    cluster_decisions.add_argument("--quiet", action="store_true")

    priority_review = subparsers.add_parser(
        "priority-review",
        help="Select unlabeled phrases from the required high-priority coverage set.",
    )
    priority_review.add_argument("--job", required=True)
    priority_review.add_argument("--limit", type=int, default=20)
    priority_review.add_argument("--quiet", action="store_true")

    policy_context_parser = subparsers.add_parser("policy-context", help="Return compact labeled examples for a topic-specific intent policy.")
    policy_context_parser.add_argument("--job", required=True)
    policy_context_parser.add_argument("--limit", type=int, default=8)
    policy_context_parser.add_argument("--quiet", action="store_true")

    apply_policy = subparsers.add_parser("apply-policy-inline", help="Save compact topic-specific commercial and informational prototypes.")
    apply_policy.add_argument("--job", required=True)
    apply_policy.add_argument("--commercial", required=True)
    apply_policy.add_argument("--implicit-commercial", required=True)
    apply_policy.add_argument("--informational", required=True)
    apply_policy.add_argument("--commercial-signals", required=True)
    apply_policy.add_argument("--informational-signals", required=True)
    apply_policy.add_argument("--weak-question-signals", required=True)
    apply_policy.add_argument("--relevant", required=True)
    apply_policy.add_argument("--garbage", required=True)
    apply_policy.add_argument("--quiet", action="store_true")

    run = subparsers.add_parser("run", help="Run classification and clustering for a prepared job.")
    run.add_argument("input")
    run.add_argument("--job", required=True)
    run.add_argument("--output", type=Path)
    run.add_argument("--quiet", action="store_true")

    run_large = subparsers.add_parser("run-large", help="Chunked local GPU/CPU mode for CSV/TSV semantic cores; full data is partitioned CSV, not Excel.")
    run_large.add_argument("input")
    run_large.add_argument("--job", required=True)
    run_large.add_argument("--chunk-size", type=int, default=50000)
    run_large.add_argument("--quiet", action="store_true")

    run_auto_parser = subparsers.add_parser("run-auto", help="Choose standard or large local GPU/CPU mode from the prepared unique-phrase count.")
    run_auto_parser.add_argument("input")
    run_auto_parser.add_argument("--job", required=True)
    run_auto_parser.add_argument("--large-threshold", type=int, default=DEFAULT_LARGE_THRESHOLD)
    run_auto_parser.add_argument("--chunk-size", type=int, default=50000)
    run_auto_parser.add_argument("--quiet", action="store_true")

    export_large = subparsers.add_parser(
        "export-large", help="Create the full review workbook from completed large-run parts without recalculation."
    )
    export_large.add_argument("input")
    export_large.add_argument("--job", required=True)
    export_large.add_argument("--quiet", action="store_true")

    finalize = subparsers.add_parser(
        "finalize",
        help="Create a stable *_FINAL.xlsx copy from the latest or specified reviewed run.",
    )
    finalize.add_argument("input")
    finalize.add_argument("--job", required=True)
    finalize.add_argument(
        "--source",
        type=Path,
        help="Reviewed result workbook; defaults to the highest numbered run.",
    )
    finalize.add_argument("--quiet", action="store_true")

    feedback = subparsers.add_parser(
        "feedback", help="Import Correct Intent values from a reviewed output workbook."
    )
    feedback.add_argument("reviewed_workbook", type=Path)
    feedback.add_argument("--job", required=True)
    feedback.add_argument("--quiet", action="store_true")

    apply_review = subparsers.add_parser(
        "apply-review",
        help="Import Human review corrections and save topic knowledge; never reruns cleaning.",
    )
    apply_review.add_argument("reviewed_workbook", type=Path)
    apply_review.add_argument("--quiet", action="store_true")

    provenance = subparsers.add_parser(
        "provenance",
        help="Assign or show the permanent ASCII Workflow Job ID for one job.",
    )
    provenance.add_argument("--job", required=True)
    provenance.add_argument("--quiet", action="store_true")

    learn = subparsers.add_parser(
        "learn", help="Persist model-reviewed labels and optional corrected feedback."
    )
    learn.add_argument("--job", required=True)
    learn.add_argument("--reviewed-workbook", type=Path)
    learn.add_argument("--quiet", action="store_true")

    knowledge = subparsers.add_parser(
        "knowledge", help="Show persistent topic knowledge."
    )
    knowledge.add_argument("--topic")
    knowledge.add_argument("--quiet", action="store_true")

    cleanup = subparsers.add_parser(
        "cleanup", help="Remove only known temporary files and Python caches."
    )
    cleanup.add_argument("--job", required=True)
    cleanup.add_argument("--quiet", action="store_true")

    forget = subparsers.add_parser(
        "forget", help="Delete one topic's persistent knowledge after explicit confirmation."
    )
    forget.add_argument("--topic", required=True)
    forget.add_argument(
        "--confirm",
        required=True,
        help="Must exactly match the normalized topic key shown by knowledge status",
    )
    forget.add_argument("--quiet", action="store_true")

    args = parser.parse_args()
    if args.command == "prepare":
        # Preparation uses scikit-learn only once; do not import it while the
        # agent merely asks for `next`, samples a batch, or imports labels.
        from seo_prepare import prepare_job

        job_dir = prepare_job(
            args.input,
            args.topic,
            args.sample_size,
            args.phrase_column,
            args.frequency_column,
            args.job_name,
            args.reuse_topic,
            args.prior_limit,
        )
        inspection = json.loads((job_dir / "inspection.json").read_text(encoding="utf-8"))
        workflow_job_id = ensure_workflow_job_id(job_dir)
        write_state(job_dir, "initial_labeling", input=inspection["input_file"], topic=inspection["topic"], workflow_job_id=workflow_job_id, labels_reviewed=0, remaining_review_rows=inspection["sample_rows"])
        # The empty seed lists are expected before the first model labels.  Do
        # not expose readiness checks here: Hermes must receive its first
        # compact labeling command, not a misleading configuration block.
        print_result(recorded_next_action(job_dir), args.quiet)
    elif args.command == "status":
        result = job_status(resolve_job(args.job))
        print_result(compact_status(result) if args.compact else result, args.quiet)
    elif args.command == "next":
        if args.job:
            job_dir = resolve_job(args.job)
            print_result(recorded_next_action(job_dir), args.quiet)
        elif args.input:
            # Validate the requested source before looking up an existing job.
            # This keeps slash-command entry points from silently accepting a
            # misspelled or missing filename.
            input_path = resolve_input(args.input)
            job_dir = find_job_for_input(input_path.name)
            if job_dir:
                existing_state = read_state(job_dir)
                if args.topic and str(existing_state.get("stage", "")) in {
                    "quality_review",
                    "knowledge_saved",
                    "learned",
                    "finalized",
                }:
                    rerun_name = f"{job_dir.name}-rerun-{time.strftime('%Y%m%d-%H%M%S')}"
                    print_result(
                        {
                            "status": "continue",
                            "stage": "prepare",
                            "action": "prepare_new_run",
                            "command": (
                                f'{HERMES_COMMAND} prepare "{input_path.name}" '
                                f'--topic "{args.topic}" --job-name "{rerun_name}" --quiet'
                            ),
                        },
                        args.quiet,
                    )
                else:
                    print_result(recorded_next_action(job_dir), args.quiet)
            elif args.topic:
                print_result(
                    {
                        "status": "continue",
                        "stage": "prepare",
                        "action": "prepare_job",
                        "command": f'{HERMES_COMMAND} prepare "{input_path.name}" --topic "{args.topic}" --quiet',
                    },
                    args.quiet,
                )
            else:
                raise ValueError("next needs --job, or both --input and --topic for a new job.")
        else:
            raise ValueError("next needs --job, or both --input and --topic for a new job.")
    elif args.command == "sample":
        print_result(sample_rows(resolve_job(args.job), args.offset, args.limit, not args.include_labeled), args.quiet)
    elif args.command == "apply-labels":
        print_result(apply_labels(resolve_job(args.job), args.input), args.quiet)
    elif args.command == "apply-labels-inline":
        print_result(
            apply_label_payload(resolve_job(args.job), parse_inline_labels(args.labels)),
            args.quiet,
        )
    elif args.command == "family-review":
        print_result(intent_family_batch(resolve_job(args.job), args.limit), args.quiet)
    elif args.command == "family-coverage-review":
        print_result(
            select_family_coverage_review(resolve_job(args.job), args.limit),
            args.quiet,
        )
    elif args.command == "apply-family-labels-inline":
        print_result(
            apply_intent_family_labels_inline(resolve_job(args.job), args.labels),
            args.quiet,
        )
    elif args.command == "review":
        print_result(select_active_review(resolve_job(args.job), args.limit), args.quiet)
    elif args.command == "relevance-review":
        print_result(select_relevance_review(resolve_job(args.job), args.limit), args.quiet)
    elif args.command == "intent-review":
        print_result(select_intent_review(resolve_job(args.job), args.limit), args.quiet)
    elif args.command == "policy-conflict-review":
        print_result(select_policy_conflict_review(resolve_job(args.job), args.limit), args.quiet)
    elif args.command == "cluster-review":
        print_result(cluster_relevance_batch(resolve_job(args.job), args.limit), args.quiet)
    elif args.command == "apply-cluster-labels-inline":
        print_result(
            apply_cluster_labels_inline(resolve_job(args.job), args.labels), args.quiet
        )
    elif args.command == "apply-cluster-decisions":
        print_result(apply_cluster_relevance_decisions(resolve_job(args.job)), args.quiet)
    elif args.command == "priority-review":
        print_result(select_priority_review(resolve_job(args.job), args.limit), args.quiet)
    elif args.command == "policy-context":
        print_result(policy_context(resolve_job(args.job), args.limit), args.quiet)
    elif args.command == "apply-policy-inline":
        print_result(
            apply_policy_inline(
                resolve_job(args.job),
                args.commercial,
                args.implicit_commercial,
                args.informational,
                args.commercial_signals,
                args.informational_signals,
                args.weak_question_signals,
                args.relevant,
                args.garbage,
            ),
            args.quiet,
        )
    elif args.command == "run":
        job_dir = resolve_job(args.job)
        readiness = job_status(job_dir)
        if not readiness["ready_for_supervised_run"]:
            details = "\n- ".join(readiness["blocking_errors"])
            raise ValueError(
                "Job is not ready for a supervised run. Fix these blocking errors:\n- "
                + details
            )
        output_path = args.output or next_output_path(job_dir, args.input)
        print_result(run_with_log(job_dir, args.input, output_path), args.quiet)
    elif args.command == "run-large":
        job_dir = resolve_job(args.job)
        readiness = job_status(job_dir)
        if not readiness["ready_for_supervised_run"]:
            details = "\n- ".join(readiness["blocking_errors"])
            raise ValueError(
                "Job is not ready for a supervised large run. Fix these blocking errors:\n- " + details
            )
        print_result(run_large_with_log(job_dir, args.input, args.chunk_size), args.quiet)
    elif args.command == "run-auto":
        job_dir = resolve_job(args.job)
        readiness = job_status(job_dir)
        if not readiness["ready_for_supervised_run"]:
            details = "\n- ".join(readiness["blocking_errors"])
            raise ValueError(
                "Job is not ready for a supervised run. Fix these blocking errors:\n- " + details
            )
        print_result(run_auto(job_dir, args.input, args.large_threshold, args.chunk_size), args.quiet)
    elif args.command == "export-large":
        print_result(export_completed_large_result(resolve_job(args.job), args.input), args.quiet)
    elif args.command == "finalize":
        job_dir = resolve_job(args.job)
        result = finalize_workbook(job_dir, args.input, args.source)
        write_state(job_dir, "finalized", final=result["final"])
        print_result(result, args.quiet)
    elif args.command == "feedback":
        job_dir = resolve_job(args.job)
        reviewed_workbook = resolve_reviewed_workbook(args.reviewed_workbook)
        imported = import_feedback(job_dir, reviewed_workbook)
        write_state(job_dir, "quality_review", feedback_imported=imported)
        print_result({"status": "feedback_imported", "imported": imported, "job": str(job_dir)}, args.quiet)
    elif args.command == "apply-review":
        print_result(
            apply_review_and_learn(args.reviewed_workbook),
            args.quiet,
        )
    elif args.command == "provenance":
        job_dir = resolve_job(args.job)
        print_result(
            {
                "status": "provenance_ready",
                "workflow_job_id": ensure_workflow_job_id(job_dir),
            },
            args.quiet,
        )
    elif args.command == "learn":
        job_dir = resolve_job(args.job)
        result = learn_job(job_dir, args.reviewed_workbook)
        write_state(job_dir, "learned", knowledge_topic_key=result["topic_key"])
        print_result(result, args.quiet)
    elif args.command == "knowledge":
        key = topic_key(args.topic) if args.topic else None
        print_result(knowledge_status(key), args.quiet)
    elif args.command == "cleanup":
        job_dir = resolve_job(args.job)
        removed = cleanup_job(job_dir)
        print_result({"removed": removed, "count": len(removed)}, args.quiet)
    elif args.command == "forget":
        key = topic_key(args.topic)
        if args.confirm != key:
            raise ValueError(f"Confirmation mismatch. Required: --confirm {key}")
        print_result({"topic_key": key, **forget_topic(key)}, args.quiet)


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(json.dumps({"status": "error", "error": str(error)}, ensure_ascii=False), file=sys.stderr)
        raise SystemExit(1)
