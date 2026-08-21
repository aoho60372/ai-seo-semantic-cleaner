"""Run model-directed SEO classification, validation, and semantic clustering."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import MiniBatchKMeans
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import classification_report
from sklearn.model_selection import StratifiedKFold, train_test_split
from scipy.sparse import csr_matrix, hstack, vstack

from seo_embeddings import encode_queries, load_local_embedding_model
from seo_io import iter_delimited_chunks, read_delimited_table

PROJECT_DIR = Path(__file__).resolve().parent
INPUT_DIR = PROJECT_DIR / "files"
OUTPUT_DIR = PROJECT_DIR / "outputs"
VALID_LABELS = {"commercial", "informational", "garbage"}
EXCEL_MAX_DATA_ROWS = 1_048_575
MIN_PRODUCTIVE_MARKER_STEM_LENGTH = 4

PROBABILITY_RESULT_COLUMNS = ["P(commercial)", "P(informational)", "P(garbage)"]
CORE_RESULT_COLUMNS = [
    "Phrase",
    "Search Volume",
    *PROBABILITY_RESULT_COLUMNS,
    "Cluster",
    "Source Row",
]
GARBAGE_RESULT_COLUMNS = [
    "Phrase",
    "Search Volume",
    *PROBABILITY_RESULT_COLUMNS,
    "Source Row",
]
HUMAN_REVIEW_RESULT_COLUMNS = [
    "Phrase",
    "Search Volume",
    "Predicted Intent",
    *PROBABILITY_RESULT_COLUMNS,
    "Cluster",
    "Review reason",
    "Correct Intent",
    "Reviewer Notes",
    "Source Row",
]
CLUSTER_SUMMARY_COLUMNS = ["Intent", "Cluster", "Phrases", "Total Search Volume"]

LABEL_ALIASES = {
    "коммерческий": "commercial",
    "коммерческие": "commercial",
    "commercial": "commercial",
    "transactional": "commercial",
    "информационный": "informational",
    "информационные": "informational",
    "informational": "informational",
    "information": "informational",
    "мусор": "garbage",
    "garbage": "garbage",
    "irrelevant": "garbage",
}


def normalize(value: object) -> str:
    text = re.sub(r"[^a-zа-я0-9]+", " ", str(value).lower().replace("ё", "е"))
    return re.sub(r"\s+", " ", text).strip()


def normalize_marker(value: object) -> str:
    """Normalize a model-generated signal while preserving its wildcard."""
    text = re.sub(r"[^a-zа-я0-9*]+", " ", str(value).lower().replace("ё", "е"))
    return re.sub(r"\s+", " ", text).strip()


def read_table(path: Path, sheet_name: str | int = 0) -> pd.DataFrame:
    if path.suffix.lower() in {".csv", ".tsv"}:
        return read_delimited_table(path)
    return pd.read_excel(path, sheet_name=sheet_name)


def resolve_input(value: str) -> Path:
    candidate = Path(value)
    if candidate.is_file():
        return candidate.resolve()
    candidate = INPUT_DIR / Path(value).name
    if candidate.is_file():
        return candidate.resolve()
    raise FileNotFoundError(f"Input not found: {value}. Expected it in {INPUT_DIR}.")


def choose_phrase_column(frame: pd.DataFrame, requested: str | None) -> str:
    normalized = {normalize(column): column for column in frame.columns}
    if requested:
        matched = normalized.get(normalize(requested))
        if matched is None:
            raise ValueError(f"Column not found: {requested}. Available: {list(frame.columns)}")
        return matched
    for candidate in ("поисковый запрос", "фраза", "keyword", "query", "запрос"):
        if candidate in normalized:
            return normalized[candidate]
    text = [column for column in frame.columns if frame[column].dtype == "object"]
    if not text:
        raise ValueError("No text phrase column found.")
    return text[0]


def build_phrase_table(
    source: pd.DataFrame,
    phrase_column: str,
    frequency_column: str | None,
) -> pd.DataFrame:
    raw = source[phrase_column].dropna().astype(str).str.strip()
    raw = raw[raw.ne("")]
    table = pd.DataFrame({"Phrase": raw, "Source Row": raw.index + 2})
    table["Normalized"] = table["Phrase"].map(normalize)
    if frequency_column and frequency_column in source.columns:
        table["Search Volume"] = pd.to_numeric(
            source.loc[raw.index, frequency_column], errors="coerce"
        ).fillna(0).to_numpy()
    else:
        table["Search Volume"] = 0.0
    return (
        table.groupby("Normalized", as_index=False)
        .agg(
            Phrase=("Phrase", "first"),
            Occurrences=("Phrase", "size"),
            **{"Search Volume": ("Search Volume", "max"), "Source Row": ("Source Row", "first")},
        )
    )


def load_config(path: Path) -> dict[str, Any]:
    config = json.loads(path.read_text(encoding="utf-8"))
    for section in ("topic", "intent", "clustering"):
        if not isinstance(config.get(section), dict):
            raise ValueError(f"Missing config section: {section}")
    if config["intent"].get("default") not in {"commercial", "informational"}:
        raise ValueError("intent.default must be commercial or informational")
    return config


def load_model_labels(path: Path | None) -> pd.DataFrame:
    if not path or not path.is_file():
        return pd.DataFrame(columns=["Phrase", "Label", "Model Notes", "Knowledge Source"])
    frame = read_table(path, "Model labels" if path.suffix.lower() == ".xlsx" else 0)
    column_lookup = {normalize(column): column for column in frame.columns}
    phrase_column = column_lookup.get("phrase") or column_lookup.get("фраза")
    label_column = column_lookup.get("model label") or column_lookup.get("label") or column_lookup.get("метка")
    notes_column = column_lookup.get("model notes") or column_lookup.get("notes")
    source_column = column_lookup.get("knowledge source")
    if not phrase_column or not label_column:
        raise ValueError("Label file must contain Phrase and Model Label columns.")
    labels = pd.DataFrame(
        {
            "Phrase": frame[phrase_column].astype(str).str.strip(),
            "Raw Label": frame[label_column].astype(str).str.strip().str.lower(),
            "Model Notes": frame[notes_column].fillna("").astype(str) if notes_column else "",
            "Knowledge Source": (
                frame[source_column].fillna("").astype(str)
                if source_column
                else "current representative sample"
            ),
        }
    )
    labels["Label"] = labels["Raw Label"].map(LABEL_ALIASES)
    labels = labels[labels["Label"].isin(VALID_LABELS) & labels["Phrase"].ne("")]
    labels["Normalized"] = labels["Phrase"].map(normalize)
    return labels.drop_duplicates("Normalized", keep="last")


def marker_pattern(marker: str) -> re.Pattern[str]:
    escaped = re.escape(normalize_marker(marker)).replace(r"\*", r"[a-zа-я0-9]*")
    return re.compile(rf"(?:^|\s){escaped}(?:$|\s)", re.IGNORECASE)


def marker_hits(texts: object, markers: list[str] | None) -> np.ndarray:
    patterns = [marker_pattern(marker) for marker in (markers or []) if normalize_marker(marker)]
    normalized_texts = [normalize(value) for value in list(texts)]
    return np.asarray(
        [any(pattern.search(text) for pattern in patterns) for text in normalized_texts],
        dtype=bool,
    )


def marker_safety_reason(marker: str) -> str | None:
    """Reject structurally collision-prone wildcards without knowing the topic."""
    normalized = normalize_marker(marker)
    if not normalized:
        return "empty"
    for token in normalized.split():
        if "*" not in token:
            continue
        stem = token.replace("*", "")
        if len(stem) < MIN_PRODUCTIVE_MARKER_STEM_LENGTH:
            return "wildcard_stem_too_short"
    return None


def safe_strong_markers(markers: list[str] | None) -> list[str]:
    return [
        str(marker)
        for marker in (markers or [])
        if marker_safety_reason(str(marker)) is None
    ]


def marker_overlaps_any(marker: str, signals: list[str]) -> bool:
    marker_stem = normalize_marker(marker).rstrip("*")
    if not marker_stem:
        return False
    return any(
        signal_stem
        and (
            marker_stem.startswith(signal_stem)
            or signal_stem.startswith(marker_stem)
        )
        for signal in signals
        for token in normalize_marker(signal).split()
        for signal_stem in [token.rstrip("*")]
    )


def intent_family_rules(config: dict[str, Any]) -> list[tuple[str, str]]:
    intent = config.get("intent", {}) if isinstance(config.get("intent", {}), dict) else {}
    configured = (
        intent.get("family_rules", {})
        if isinstance(intent.get("family_rules", {}), dict)
        else {}
    )
    weak_markers = [str(value) for value in intent.get("weak_question_markers", [])]
    rules: dict[str, str] = {}
    for label in ("commercial", "informational", "neutral"):
        for pattern in configured.get(label, []):
            normalized_pattern = str(pattern).strip()
            if not normalized_pattern:
                continue
            rules[normalized_pattern] = (
                "neutral"
                if label != "neutral" and marker_overlaps_any(normalized_pattern, weak_markers)
                else label
            )
    return list(rules.items())


def build_intent_family_audit(
    frames: object,
    config: dict[str, Any],
) -> pd.DataFrame:
    """Audit reviewed families against predictions made before family rules."""
    rules = intent_family_rules(config)
    columns = [
        "Family",
        "Reviewed decision",
        "Matched rows",
        "Commercial",
        "Informational",
        "Garbage",
        "Pre-family commercial",
        "Pre-family informational",
        "Pre-family expected share",
        "Family conflicts",
        "Family overrides",
        "Audit status",
        "Examples",
    ]
    if not rules:
        return pd.DataFrame(columns=columns)
    accumulators = {
        pattern: {
            "label": label,
            "counts": {"commercial": 0, "informational": 0, "garbage": 0},
            "pre_counts": {"commercial": 0, "informational": 0},
            "pre_matched": 0,
            "conflicts": 0,
            "overrides": 0,
            "matched": 0,
            "examples": [],
        }
        for pattern, label in rules
    }
    for frame in frames:
        if frame.empty or not {"Phrase", "Intent"}.issubset(frame.columns):
            continue
        phrases = frame["Phrase"].fillna("").astype(str)
        intents = frame["Intent"].fillna("").astype(str)
        for pattern, _ in rules:
            mask = marker_hits(phrases, [pattern])
            if not mask.any():
                continue
            selected = frame.loc[mask]
            item = accumulators[pattern]
            item["matched"] += int(len(selected))
            values = selected["Intent"].value_counts()
            for label in item["counts"]:
                item["counts"][label] += int(values.get(label, 0))
            if "Pre-family intent" in selected.columns:
                pre_values = selected["Pre-family intent"].value_counts()
                item["pre_matched"] += int(len(selected))
                for intent_label in item["pre_counts"]:
                    item["pre_counts"][intent_label] += int(pre_values.get(intent_label, 0))
                if "Intent family override" in selected.columns:
                    item["overrides"] += int(
                        selected["Intent family override"].fillna(False).astype(bool).sum()
                    )
            if "Intent family conflict" in selected.columns:
                item["conflicts"] += int(selected["Intent family conflict"].fillna(False).astype(bool).sum())
            for phrase in selected["Phrase"].astype(str):
                if phrase not in item["examples"] and len(item["examples"]) < 3:
                    item["examples"].append(phrase)
    rows = []
    for pattern, label in rules:
        item = accumulators[pattern]
        relevant = item["pre_counts"]["commercial"] + item["pre_counts"]["informational"]
        expected_share: float | None = None
        status = "neutral_observation"
        if label in {"commercial", "informational"}:
            if item["pre_matched"] == 0:
                status = "unverifiable_legacy"
            else:
                expected_share = item["pre_counts"][label] / max(relevant, 1)
                conflict_share = item["conflicts"] / max(item["matched"], 1)
                status = (
                    "pass"
                    if relevant > 0 and expected_share >= 0.70 and conflict_share <= 0.10
                    else "review"
                )
        rows.append(
            {
                "Family": pattern,
                "Reviewed decision": label,
                "Matched rows": item["matched"],
                "Commercial": item["counts"]["commercial"],
                "Informational": item["counts"]["informational"],
                "Garbage": item["counts"]["garbage"],
                "Pre-family commercial": item["pre_counts"]["commercial"],
                "Pre-family informational": item["pre_counts"]["informational"],
                "Pre-family expected share": expected_share,
                "Family conflicts": item["conflicts"],
                "Family overrides": item["overrides"],
                "Audit status": status,
                "Examples": " | ".join(item["examples"]),
            }
        )
    return pd.DataFrame(rows, columns=columns).sort_values(
        ["Matched rows", "Family"], ascending=[False, True]
    )


def seed_scores(
    vectors: np.ndarray,
    seeds: list[str],
    model: SentenceTransformer,
    batch_size: int,
) -> np.ndarray:
    if not seeds:
        return np.zeros(len(vectors), dtype=float)
    seed_vectors = encode_queries(model, [normalize(seed) for seed in seeds], batch_size)
    return np.max(vectors @ np.asarray(seed_vectors).T, axis=1)


def seed_classification(
    phrase_table: pd.DataFrame,
    vectors: np.ndarray,
    model: SentenceTransformer,
    config: dict[str, Any],
    batch_size: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    topic, intent = config["topic"], config["intent"]
    required = {
        "topic.relevant_seeds": topic.get("relevant_seeds"),
        "intent.commercial_seeds": intent.get("commercial_seeds"),
        "intent.informational_seeds": intent.get("informational_seeds"),
    }
    missing = [name for name, values in required.items() if not values]
    if missing:
        raise ValueError(
            "Not enough labeled examples and missing seed lists: " + ", ".join(missing)
        )

    texts = phrase_table["Normalized"].tolist()
    relevant = seed_scores(vectors, topic["relevant_seeds"], model, batch_size)
    garbage = seed_scores(vectors, topic.get("garbage_seeds", []), model, batch_size)
    commercial = seed_scores(vectors, intent["commercial_seeds"], model, batch_size)
    informational = seed_scores(vectors, intent["informational_seeds"], model, batch_size)
    positive_hits = marker_hits(texts, topic.get("positive_markers"))
    garbage_hits = marker_hits(texts, topic.get("garbage_markers"))
    commercial_hits = marker_hits(
        texts, safe_strong_markers(intent.get("commercial_markers"))
    )
    informational_hits = marker_hits(
        texts, safe_strong_markers(intent.get("informational_markers"))
    )
    minimum_relevance = float(topic.get("minimum_relevance", 0.35))
    garbage_margin = float(topic.get("garbage_margin", 0.05))
    intent_margin = float(intent.get("semantic_margin", 0.03))
    labels, bases, confidences = [], [], []

    for index in range(len(phrase_table)):
        if garbage_hits[index] and not positive_hits[index]:
            label, basis, confidence = "garbage", "Task garbage marker", 1.0
        elif (
            relevant[index] < minimum_relevance
            and garbage[index] >= relevant[index] + garbage_margin
            and not positive_hits[index]
        ):
            label, basis = "garbage", "Seed-based topic relevance"
            confidence = min(1.0, garbage[index] - relevant[index] + 0.5)
        elif commercial_hits[index] and not informational_hits[index]:
            label, basis, confidence = "commercial", "Task commercial marker", 1.0
        elif informational_hits[index] and not commercial_hits[index]:
            label, basis, confidence = "informational", "Task informational marker", 1.0
        elif commercial[index] >= informational[index] + intent_margin:
            label, basis = "commercial", "Commercial seed similarity"
            confidence = min(1.0, commercial[index] - informational[index] + 0.5)
        elif informational[index] >= commercial[index] + intent_margin:
            label, basis = "informational", "Informational seed similarity"
            confidence = min(1.0, informational[index] - commercial[index] + 0.5)
        else:
            label, basis, confidence = intent["default"], "Ambiguous seed result", 0.5
        labels.append(label)
        bases.append(basis)
        confidences.append(confidence)

    result = phrase_table.copy()
    result["Intent"] = labels
    result["Classification confidence"] = confidences
    result["Decision basis"] = bases
    result["Topic relevance"] = relevant
    return result, pd.DataFrame([("Mode", "Seed fallback")], columns=["Metric", "Value"])


def relevance_settings(config: dict[str, Any]) -> dict[str, float]:
    configured = config.get("relevance", {})
    configured = configured if isinstance(configured, dict) else {}
    return {
        "garbage_threshold": float(configured.get("garbage_threshold", 0.45)),
        "configured_garbage_threshold": float(configured.get("garbage_threshold", 0.45)),
        "quarantine_margin": float(configured.get("quarantine_margin", 0.10)),
        "minimum_garbage_precision": float(configured.get("minimum_garbage_precision", 0.80)),
        "maximum_relevant_false_positive_rate": float(
            configured.get("maximum_relevant_false_positive_rate", 0.05)
        ),
        "tfidf_weight": float(configured.get("tfidf_weight", 0.35)),
        "max_garbage_class_weight": float(configured.get("max_garbage_class_weight", 8.0)),
        "intent_tfidf_weight": float(configured.get("intent_tfidf_weight", 1.25)),
    }


def calibrate_garbage_threshold(
    features: Any,
    targets: np.ndarray,
    class_weight: Any,
    settings: dict[str, float],
) -> tuple[float, dict[str, object]]:
    """Choose a conservative topic threshold from out-of-fold real labels."""
    configured = float(settings["configured_garbage_threshold"])
    counts = pd.Series(targets).value_counts()
    garbage_count = int(counts.get("garbage", 0))
    relevant_count = int(counts.get("relevant", 0))
    if garbage_count < 8 or relevant_count < 20:
        return configured, {
            "mode": "configured_fallback",
            "reason": "Need at least 8 real garbage and 20 real relevant labels for calibration.",
        }
    folds = min(5, garbage_count, relevant_count)
    probabilities = np.zeros(len(targets), dtype=float)
    splitter = StratifiedKFold(n_splits=folds, shuffle=True, random_state=42)
    for train_index, test_index in splitter.split(features, targets):
        fitted = LogisticRegression(
            max_iter=2000, class_weight=class_weight, random_state=42
        ).fit(features[train_index], targets[train_index])
        class_index = list(fitted.classes_).index("garbage")
        probabilities[test_index] = fitted.predict_proba(features[test_index])[:, class_index]

    candidates: list[tuple[float, float, float, float]] = []
    actual_garbage = targets == "garbage"
    actual_relevant = ~actual_garbage
    for threshold in np.linspace(0.20, 0.60, 81):
        predicted = probabilities >= threshold
        true_positive = int(np.sum(predicted & actual_garbage))
        false_positive = int(np.sum(predicted & actual_relevant))
        precision = true_positive / max(int(np.sum(predicted)), 1)
        recall = true_positive / max(garbage_count, 1)
        false_positive_rate = false_positive / max(relevant_count, 1)
        if (
            precision >= settings["minimum_garbage_precision"]
            and false_positive_rate <= settings["maximum_relevant_false_positive_rate"]
        ):
            candidates.append((recall, precision, -false_positive_rate, float(threshold)))
    if not candidates:
        return configured, {
            "mode": "configured_fallback",
            "reason": "No cross-validation threshold met the precision and false-positive safeguards.",
        }
    recall, precision, negative_fpr, threshold = max(candidates)
    return threshold, {
        "mode": "out_of_fold",
        "folds": folds,
        "garbage_examples": garbage_count,
        "relevant_examples": relevant_count,
        "garbage_precision": round(precision, 4),
        "garbage_recall": round(recall, 4),
        "relevant_false_positive_rate": round(-negative_fpr, 4),
    }


def _binary_validation(features: Any, targets: np.ndarray, name: str, class_weight: Any) -> list[tuple[str, Any]]:
    counts = pd.Series(targets).value_counts()
    rows: list[tuple[str, Any]] = [(f"{name} labeled phrases", len(targets))]
    if len(counts) < 2 or len(targets) < 40 or int(counts.min()) < 4:
        rows.append((f"{name} validation", "Insufficient per-class labels for holdout validation"))
        return rows
    x_train, x_test, y_train, y_test = train_test_split(
        features, targets, test_size=0.25, random_state=42, stratify=targets
    )
    fitted = LogisticRegression(max_iter=2000, class_weight=class_weight, random_state=42).fit(x_train, y_train)
    report = classification_report(y_test, fitted.predict(x_test), output_dict=True, zero_division=0)
    rows.extend([(f"{name} validation samples", len(y_test)), (f"{name} validation accuracy", report["accuracy"])])
    for label in sorted(set(y_test)):
        rows.extend([
            (f"{name} {label} precision", report[label]["precision"]),
            (f"{name} {label} recall", report[label]["recall"]),
            (f"{name} {label} f1", report[label]["f1-score"]),
        ])
    return rows


def fit_two_stage_classifier(
    labels: pd.DataFrame, model: SentenceTransformer, batch_size: int, config: dict[str, Any]
) -> dict[str, Any]:
    """Fit topic relevance separately from commercial/informational intent."""
    counts = labels["Label"].value_counts()
    if len(labels) < 30 or not {"commercial", "informational"}.issubset(counts.index):
        raise ValueError(
            "At least 30 valid model labels covering commercial and informational are required. "
            "Garbage is optional and must never be invented."
        )
    settings = relevance_settings(config)
    knowledge_source = labels.get(
        "Knowledge Source", pd.Series("current representative sample", index=labels.index)
    ).fillna("").astype(str).str.lower()
    actual_weights = np.ones(len(labels), dtype=float)
    actual_weights[knowledge_source.eq("model_reviewed").to_numpy()] = 0.35
    actual_weights[knowledge_source.eq("review_corrected").to_numpy()] = 0.80
    train_vectors = encode_queries(model, labels["Normalized"], batch_size)
    relevance_policy = config.get("relevance", {}) if isinstance(config.get("relevance", {}), dict) else {}
    intent_policy = config.get("intent_policy", {}) if isinstance(config.get("intent_policy", {}), dict) else {}
    relevant_prototypes = [normalize(value) for value in relevance_policy.get("relevant_prototypes", []) if normalize(value)]
    garbage_prototypes = [normalize(value) for value in relevance_policy.get("garbage_prototypes", []) if normalize(value)]
    commercial_prototypes = [
        normalize(value)
        for key in ("commercial_prototypes", "implicit_commercial_prototypes")
        for value in intent_policy.get(key, [])
        if normalize(value)
    ]
    informational_prototypes = [
        normalize(value) for value in intent_policy.get("informational_prototypes", []) if normalize(value)
    ]
    relevance_synthetic_texts = relevant_prototypes + garbage_prototypes
    intent_synthetic_texts = commercial_prototypes + informational_prototypes
    structural_corpus = pd.concat(
        [
            labels["Normalized"],
            pd.Series(relevance_synthetic_texts, dtype=object),
            pd.Series(intent_synthetic_texts, dtype=object),
        ],
        ignore_index=True,
    )
    structural = TfidfVectorizer(
        analyzer="char_wb", ngram_range=(3, 5), min_df=1, max_features=30000, sublinear_tf=True
    )
    structural.fit(structural_corpus)
    train_structural = structural.transform(labels["Normalized"])

    actual_relevance_targets = np.where(labels["Label"].eq("garbage"), "garbage", "relevant")
    relevance_targets = np.concatenate([
        actual_relevance_targets,
        np.asarray(["relevant"] * len(relevant_prototypes) + ["garbage"] * len(garbage_prototypes)),
    ])
    relevance_vectors = train_vectors
    if relevance_synthetic_texts:
        relevance_vectors = np.vstack([
            train_vectors, encode_queries(model, relevance_synthetic_texts, batch_size)
        ])
    relevance_structural = structural.transform(pd.concat([
        labels["Normalized"], pd.Series(relevance_synthetic_texts, dtype=object)
    ], ignore_index=True))
    relevance_features = hstack(
        [csr_matrix(relevance_vectors), relevance_structural * settings["tfidf_weight"]], format="csr"
    )
    relevance_classifier = None
    relevance_class_weight = None
    if len(set(relevance_targets)) == 2:
        garbage_count = int(np.sum(relevance_targets == "garbage"))
        relevant_count = int(np.sum(relevance_targets == "relevant"))
        relevance_class_weight = {
            "garbage": min(settings["max_garbage_class_weight"], relevant_count / max(garbage_count, 1)),
            "relevant": 1.0,
        }
        relevance_classifier = LogisticRegression(
            max_iter=2000, class_weight=relevance_class_weight, random_state=42
        ).fit(
            relevance_features,
            relevance_targets,
            sample_weight=np.concatenate([
                actual_weights,
                np.full(
                    len(relevance_synthetic_texts),
                    float(relevance_policy.get("synthetic_weight", 1.0)),
                ),
            ]),
        )

    intent_mask = labels["Label"].isin(["commercial", "informational"]).to_numpy()
    actual_intent_features = hstack([
        csr_matrix(train_vectors[intent_mask]),
        train_structural[intent_mask] * settings["intent_tfidf_weight"],
    ], format="csr")
    actual_intent_targets = labels.loc[intent_mask, "Label"].to_numpy()
    intent_features = actual_intent_features
    intent_targets = actual_intent_targets
    intent_weights = actual_weights[intent_mask]
    if intent_synthetic_texts:
        intent_synthetic_vectors = encode_queries(model, intent_synthetic_texts, batch_size)
        intent_synthetic_structural = structural.transform(intent_synthetic_texts)
        intent_synthetic_features = hstack([
            csr_matrix(intent_synthetic_vectors),
            intent_synthetic_structural * settings["intent_tfidf_weight"],
        ], format="csr")
        intent_features = vstack([actual_intent_features, intent_synthetic_features], format="csr")
        intent_targets = np.concatenate([
            actual_intent_targets,
            np.asarray(
                ["commercial"] * len(commercial_prototypes)
                + ["informational"] * len(informational_prototypes)
            ),
        ])
        intent_weights = np.concatenate([
            intent_weights,
            np.full(
                len(intent_synthetic_texts),
                float(intent_policy.get("synthetic_weight", 0.75)),
            ),
        ])
    intent_classifier = LogisticRegression(max_iter=2000, class_weight="balanced", random_state=42).fit(
        intent_features, intent_targets, sample_weight=intent_weights
    )
    validation_rows: list[tuple[str, Any]] = [
        ("Mode", "Two-stage supervised relevance then intent"),
        ("Labeled phrases", len(labels)),
        ("Garbage labels", int(counts.get("garbage", 0))),
        ("Synthetic relevant boundary examples", len(relevant_prototypes)),
        ("Synthetic garbage boundary examples", len(garbage_prototypes)),
        ("Synthetic commercial intent examples", len(commercial_prototypes)),
        ("Synthetic informational intent examples", len(informational_prototypes)),
        ("Auxiliary prior model labels", int(knowledge_source.eq("model_reviewed").sum())),
        ("Auxiliary prior corrected labels", int(knowledge_source.eq("review_corrected").sum())),
        ("Garbage class weight cap", settings["max_garbage_class_weight"]),
        ("Configured garbage threshold", settings["configured_garbage_threshold"]),
    ]
    actual_relevance_features = hstack(
        [csr_matrix(train_vectors), train_structural * settings["tfidf_weight"]], format="csr"
    )
    calibrated_threshold, calibration = calibrate_garbage_threshold(
        actual_relevance_features,
        actual_relevance_targets,
        relevance_class_weight,
        settings,
    )
    settings["garbage_threshold"] = calibrated_threshold
    settings["quarantine_threshold"] = max(
        0.05, calibrated_threshold - settings["quarantine_margin"]
    )
    validation_rows.extend(
        [
            ("Garbage decision threshold", calibrated_threshold),
            ("Relevance quarantine threshold", settings["quarantine_threshold"]),
            ("Garbage threshold calibration", calibration.get("mode", "unknown")),
            ("Calibration garbage precision", calibration.get("garbage_precision", "n/a")),
            ("Calibration garbage recall", calibration.get("garbage_recall", "n/a")),
            (
                "Calibration relevant false-positive rate",
                calibration.get("relevant_false_positive_rate", "n/a"),
            ),
        ]
    )
    validation_rows.extend(_binary_validation(
        actual_relevance_features, actual_relevance_targets, "Relevance", relevance_class_weight
    ))
    validation_rows.extend(_binary_validation(
        actual_intent_features, actual_intent_targets, "Intent", "balanced"
    ))
    return {
        "structural": structural,
        "relevance_classifier": relevance_classifier,
        "intent_classifier": intent_classifier,
        "settings": settings,
        "validation": pd.DataFrame(validation_rows, columns=["Metric", "Value"]),
    }


def predict_two_stage(
    bundle: dict[str, Any],
    texts: pd.Series,
    vectors: np.ndarray,
    config: dict[str, Any],
    policy_commercial_vectors: np.ndarray | None = None,
    policy_informational_vectors: np.ndarray | None = None,
) -> dict[str, np.ndarray]:
    structural = bundle["structural"].transform(texts)
    settings = bundle["settings"]
    relevance_classifier = bundle["relevance_classifier"]
    if relevance_classifier is None:
        garbage_probability = np.zeros(len(texts), dtype=float)
    else:
        relevance_features = hstack(
            [csr_matrix(vectors), structural * settings["tfidf_weight"]], format="csr"
        )
        relevance_probabilities = relevance_classifier.predict_proba(relevance_features)
        relevance_index = {label: index for index, label in enumerate(relevance_classifier.classes_)}
        garbage_probability = relevance_probabilities[:, relevance_index["garbage"]]

    intent_features = hstack(
        [csr_matrix(vectors), structural * settings["intent_tfidf_weight"]], format="csr"
    )
    intent_classifier = bundle["intent_classifier"]
    conditional = intent_classifier.predict_proba(intent_features)
    intent_index = {label: index for index, label in enumerate(intent_classifier.classes_)}
    commercial_conditional = conditional[:, intent_index["commercial"]]
    informational_conditional = conditional[:, intent_index["informational"]]

    policy = config.get("intent_policy", {}) if isinstance(config.get("intent_policy", {}), dict) else {}
    policy_adjustment = np.zeros(len(texts), dtype=float)
    commercial_similarity = np.zeros(len(texts), dtype=float)
    informational_similarity = np.zeros(len(texts), dtype=float)
    if policy_commercial_vectors is not None and policy_informational_vectors is not None:
        commercial_similarity = np.max(vectors @ policy_commercial_vectors.T, axis=1)
        informational_similarity = np.max(vectors @ policy_informational_vectors.T, axis=1)
        applicable = np.maximum(commercial_similarity, informational_similarity) >= float(policy.get("minimum_similarity", 0.55))
        policy_adjustment = np.where(
            applicable,
            np.clip(commercial_similarity - informational_similarity, -0.20, 0.20) * float(policy.get("strength", 0.12)),
            0.0,
        )
        commercial_conditional = np.clip(commercial_conditional + policy_adjustment, 1e-8, None)
        informational_conditional = np.clip(informational_conditional - policy_adjustment, 1e-8, None)
        total = commercial_conditional + informational_conditional
        commercial_conditional /= total
        informational_conditional /= total

    relevant_probability = 1.0 - garbage_probability
    probabilities = np.column_stack([
        relevant_probability * commercial_conditional,
        relevant_probability * informational_conditional,
        garbage_probability,
    ])
    intent_config = config.get("intent", {}) if isinstance(config.get("intent", {}), dict) else {}
    default_intent = str(intent_config.get("default", "commercial"))
    decision_margin = float(intent_config.get("informational_decision_margin", 0.12))
    strong_information_margin = float(
        intent_config.get("strong_informational_decision_margin", 0.02)
    )
    policy_evidence_margin = float(policy.get("informational_evidence_margin", 0.005))
    informational_marker_hit = marker_hits(
        texts,
        safe_strong_markers(list(intent_config.get("informational_markers", []))),
    )
    commercial_marker_hit = marker_hits(
        texts,
        safe_strong_markers(list(intent_config.get("commercial_markers", []))),
    )
    weak_question_marker_hit = marker_hits(
        texts, list(intent_config.get("weak_question_markers", []))
    )
    reviewed_family_rules = intent_family_rules(config)
    informational_family_hit = marker_hits(
        texts,
        [pattern for pattern, label in reviewed_family_rules if label == "informational"],
    )
    commercial_family_hit = marker_hits(
        texts,
        [pattern for pattern, label in reviewed_family_rules if label == "commercial"],
    )
    required_information_margin = np.where(
        informational_marker_hit
        | (informational_similarity >= commercial_similarity + policy_evidence_margin),
        strong_information_margin,
        decision_margin,
    )
    weak_question_margin = float(
        intent_config.get("weak_question_informational_margin", strong_information_margin)
    )
    weak_question_without_transaction = (
        weak_question_marker_hit
        & ~commercial_marker_hit
        & ~commercial_family_hit
    )
    required_information_margin = np.where(
        weak_question_without_transaction,
        np.minimum(required_information_margin, weak_question_margin),
        required_information_margin,
    )
    if default_intent == "informational":
        relevant_intent = np.where(
            commercial_conditional >= informational_conditional + decision_margin,
            "commercial",
            "informational",
        )
    else:
        relevant_intent = np.where(
            informational_conditional >= commercial_conditional + required_information_margin,
            "informational",
            "commercial",
        )
    relevant_intent = np.where(
        informational_marker_hit & ~commercial_marker_hit,
        "informational",
        relevant_intent,
    )
    relevant_intent = np.where(
        commercial_marker_hit & ~informational_marker_hit,
        "commercial",
        relevant_intent,
    )
    pre_family_intent = relevant_intent.copy()
    family_tolerance = float(intent_config.get("family_override_tolerance", 0.05))
    informational_family_supported = (
        informational_conditional + family_tolerance >= commercial_conditional
    )
    commercial_family_supported = (
        commercial_conditional + family_tolerance >= informational_conditional
    )
    informational_family_applies = (
        informational_family_hit
        & ~commercial_family_hit
        & ~commercial_marker_hit
        & informational_family_supported
    )
    commercial_family_applies = (
        commercial_family_hit
        & ~informational_family_hit
        & ~informational_marker_hit
        & commercial_family_supported
    )
    family_conflict = (
        (informational_family_hit & commercial_family_hit)
        | (informational_family_hit & commercial_marker_hit)
        | (commercial_family_hit & informational_marker_hit)
        | (informational_family_hit & ~informational_family_supported)
        | (commercial_family_hit & ~commercial_family_supported)
    )
    # A reviewed lexical family is bounded evidence, not a universal rule. It
    # may resolve a nearby decision, but never defeats an opposite strong
    # marker, another reviewed family, or a confident classifier prediction.
    relevant_intent = np.where(
        commercial_family_applies,
        "commercial",
        relevant_intent,
    )
    relevant_intent = np.where(
        informational_family_applies,
        "informational",
        relevant_intent,
    )
    family_override = (
        (relevant_intent != pre_family_intent)
        & (informational_family_applies | commercial_family_applies)
    )
    quarantine_threshold = float(
        settings.get(
            "quarantine_threshold",
            max(0.05, settings["garbage_threshold"] - settings.get("quarantine_margin", 0.10)),
        )
    )
    relevance_quarantine = (
        (garbage_probability >= quarantine_threshold)
        & (garbage_probability < settings["garbage_threshold"])
    )
    prediction = np.where(
        garbage_probability >= settings["garbage_threshold"],
        "garbage",
        relevant_intent,
    ).astype(object)
    confidence = np.max(probabilities, axis=1)
    predicted_index = np.asarray(
        [
            {"commercial": 0, "informational": 1, "garbage": 2}.get(str(value), 0)
            for value in prediction
        ]
    )
    confidence = probabilities[np.arange(len(probabilities)), predicted_index]
    family_override &= prediction != "garbage"
    return {
        "prediction": prediction,
        "confidence": confidence,
        "probabilities": probabilities,
        "policy_adjustment": policy_adjustment,
        "policy_commercial_similarity": commercial_similarity,
        "policy_informational_similarity": informational_similarity,
        "informational_marker_hit": informational_marker_hit,
        "commercial_marker_hit": commercial_marker_hit,
        "weak_question_marker_hit": weak_question_marker_hit,
        "weak_question_without_transaction": weak_question_without_transaction,
        "informational_family_hit": informational_family_hit,
        "commercial_family_hit": commercial_family_hit,
        "family_override": family_override,
        "family_conflict": family_conflict,
        "pre_family_intent": pre_family_intent,
        "relevance_quarantine": relevance_quarantine,
        "garbage_threshold": np.full(len(texts), settings["garbage_threshold"]),
        "quarantine_threshold": np.full(len(texts), quarantine_threshold),
    }


def supervised_classification(
    phrase_table: pd.DataFrame,
    vectors: np.ndarray,
    labels: pd.DataFrame,
    config: dict[str, Any],
    model: SentenceTransformer,
    batch_size: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    counts = labels["Label"].value_counts()
    required_intents = {"commercial", "informational"}
    if not required_intents.issubset(set(counts.index)) or len(labels) < 30:
        raise ValueError(
            "At least 30 valid model labels covering commercial and informational are required. "
            "Garbage is optional and must never be invented."
        )

    bundle = fit_two_stage_classifier(labels, model, batch_size, config)

    policy = config.get("intent_policy", {}) if isinstance(config.get("intent_policy", {}), dict) else {}
    commercial_prototypes = [
        str(value)
        for key in ("commercial_prototypes", "implicit_commercial_prototypes")
        for value in policy.get(key, [])
        if str(value).strip()
    ]
    informational_prototypes = [str(value) for value in policy.get("informational_prototypes", []) if str(value).strip()]
    policy_adjustment = np.zeros(len(phrase_table), dtype=float)
    policy_commercial_similarity = np.zeros(len(phrase_table), dtype=float)
    policy_informational_similarity = np.zeros(len(phrase_table), dtype=float)
    commercial_vectors = encode_queries(model, commercial_prototypes, batch_size) if commercial_prototypes else None
    informational_vectors = encode_queries(model, informational_prototypes, batch_size) if informational_prototypes else None
    prediction_data = predict_two_stage(
        bundle,
        phrase_table["Normalized"],
        vectors,
        config,
        commercial_vectors,
        informational_vectors,
    )
    prediction = prediction_data["prediction"]
    confidence = prediction_data["confidence"]
    probabilities = prediction_data["probabilities"]
    policy_adjustment = prediction_data["policy_adjustment"]
    policy_commercial_similarity = prediction_data["policy_commercial_similarity"]
    policy_informational_similarity = prediction_data["policy_informational_similarity"]
    relevance_quarantine = prediction_data["relevance_quarantine"].copy()

    exact_mask = ~labels["Knowledge Source"].fillna("").astype(str).str.lower().isin(
        ["model_reviewed", "review_corrected"]
    )
    exact_labels = labels.loc[exact_mask].set_index("Normalized")["Label"].to_dict()
    bases = ["Classifier trained on model labels"] * len(phrase_table)
    for index, adjustment in enumerate(policy_adjustment):
        if abs(adjustment) >= 0.01:
            bases[index] = "Classifier plus topic intent policy"
    for index in np.flatnonzero(prediction_data["family_override"]):
        bases[index] = "Main-model intent family rule"
    for index, normalized in enumerate(phrase_table["Normalized"]):
        if normalized in exact_labels:
            prediction[index] = exact_labels[normalized]
            confidence[index] = 1.0
            probabilities[index] = 0.0
            probabilities[index, {"commercial": 0, "informational": 1, "garbage": 2}[prediction[index]]] = 1.0
            relevance_quarantine[index] = False
            bases[index] = "Exact model-reviewed label"

    result = phrase_table.copy()
    result["Intent"] = prediction
    result["Classification confidence"] = confidence
    result["Decision basis"] = bases
    result["Policy commercial similarity"] = policy_commercial_similarity
    result["Policy informational similarity"] = policy_informational_similarity
    result["Policy adjustment"] = policy_adjustment
    result["Informational marker hit"] = prediction_data["informational_marker_hit"]
    result["Commercial marker hit"] = prediction_data["commercial_marker_hit"]
    result["Weak question marker hit"] = prediction_data["weak_question_marker_hit"]
    result["Weak question without transaction"] = prediction_data[
        "weak_question_without_transaction"
    ]
    result["Informational family hit"] = prediction_data["informational_family_hit"]
    result["Commercial family hit"] = prediction_data["commercial_family_hit"]
    result["Pre-family intent"] = prediction_data["pre_family_intent"]
    result["Intent family conflict"] = prediction_data["family_conflict"]
    result["Intent family override"] = prediction_data["family_override"]
    result["Relevance quarantine"] = relevance_quarantine
    result["Garbage threshold"] = prediction_data["garbage_threshold"]
    result["Quarantine threshold"] = prediction_data["quarantine_threshold"]
    for column_index, label in enumerate(("commercial", "informational", "garbage")):
        result[f"P({label})"] = probabilities[:, column_index]
    result["Topic relevance"] = 1.0 - result["P(garbage)"]
    result["Relevance confidence"] = np.maximum(result["P(garbage)"], result["Topic relevance"])
    return result, bundle["validation"]


def cluster_labels(phrases: list[str], labels: np.ndarray, stop_words: list[str]) -> list[str]:
    result = ["Unclustered" if label == -1 else "" for label in labels]
    for label in sorted(set(labels)):
        if label == -1:
            continue
        positions = np.flatnonzero(labels == label)
        corpus = [normalize(phrases[position]) for position in positions]
        try:
            vectorizer = TfidfVectorizer(
                ngram_range=(1, 2),
                stop_words=[normalize(word) for word in stop_words],
                max_features=300,
            )
            matrix = vectorizer.fit_transform(corpus)
            terms = vectorizer.get_feature_names_out()
            scores = np.asarray(matrix.mean(axis=0)).ravel()
            best = [terms[position] for position in scores.argsort()[::-1] if len(terms[position]) > 2][:3]
            name = " / ".join(best) or f"Cluster {label}"
        except ValueError:
            name = f"Cluster {label}"
        for position in positions:
            result[position] = name
    return result


def cluster_subset(
    subset: pd.DataFrame,
    vectors: np.ndarray,
    config: dict[str, Any],
) -> pd.DataFrame:
    result = subset.copy().reset_index(drop=True)
    if len(result) < 8:
        result["Cluster"] = "Small cluster"
        result["Cluster probability"] = 1.0
        return result
    # UMAP and HDBSCAN are needed only for the final clustering operation.
    # Delaying their imports keeps workflow-control commands fast.
    import hdbscan
    import umap

    clustering = config["clustering"]
    neighbors = min(int(clustering.get("n_neighbors", 30)), len(result) - 1)
    dimensions = min(int(clustering.get("umap_dimensions", 20)), len(result) - 2)
    minimum_size = min(
        int(clustering.get("min_cluster_size", 20)), max(2, len(result) // 2)
    )
    reduced = umap.UMAP(
        n_neighbors=max(2, neighbors),
        n_components=max(2, dimensions),
        metric="cosine",
        random_state=42,
        low_memory=True,
    ).fit_transform(vectors)
    fitted = hdbscan.HDBSCAN(
        min_cluster_size=max(2, minimum_size),
        min_samples=max(2, minimum_size // 2),
        cluster_selection_method="eom",
    ).fit(reduced)
    result["Cluster"] = cluster_labels(
        result["Phrase"].tolist(), fitted.labels_, config.get("cluster_stop_words", [])
    )
    result["Cluster probability"] = fitted.probabilities_
    return result


def build_cluster_summary(frame: pd.DataFrame, intent: str) -> pd.DataFrame:
    rows = []
    for cluster, group in frame.groupby("Cluster", dropna=False):
        ranked = group.sort_values(
            ["Search Volume", "Cluster probability", "Occurrences"],
            ascending=False,
        )
        examples = "\n".join(ranked["Phrase"].head(8).tolist())
        rows.append(
            {
                "Intent": intent,
                "Cluster": cluster,
                "Phrases": len(group),
                "Total Search Volume": float(group["Search Volume"].sum()),
                "Average Classification Confidence": float(
                    group["Classification confidence"].mean()
                ),
                "Average Cluster Probability": float(group["Cluster probability"].mean()),
                "Representative Phrases": examples,
                "Model Cluster Name": "",
                "Model Action": "",
                "Model Notes": "",
            }
        )
    return pd.DataFrame(rows).sort_values(
        ["Total Search Volume", "Phrases"], ascending=False
    )


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cells in worksheet.columns:
            values = [len(str(cell.value or "")) for cell in cells[:200]]
            width = min(max(values, default=10) + 2, 60)
            worksheet.column_dimensions[cells[0].column_letter].width = width
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def flatten_config(value: Any, prefix: str = "") -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    if isinstance(value, dict):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(flatten_config(child, child_prefix))
    elif isinstance(value, list):
        rows.append((prefix, " | ".join(str(item) for item in value)))
    else:
        rows.append((prefix, str(value)))
    return rows


def compact_result_frame(frame: pd.DataFrame, result_type: str) -> pd.DataFrame:
    """Return only the columns needed in the user-facing result workbook."""
    result = frame.copy()
    if result_type == "human_review":
        if "Predicted Intent" not in result.columns:
            result["Predicted Intent"] = result.get("Intent", "")
        columns = HUMAN_REVIEW_RESULT_COLUMNS
    elif result_type == "garbage":
        columns = GARBAGE_RESULT_COLUMNS
    else:
        columns = CORE_RESULT_COLUMNS
    for column in columns:
        if column not in result.columns:
            result[column] = ""
    for column in PROBABILITY_RESULT_COLUMNS:
        if column in result.columns:
            result[column] = pd.to_numeric(result[column], errors="coerce").round(4)
    return result[columns]


def compact_cluster_summary(frame: pd.DataFrame) -> pd.DataFrame:
    """Normalize and trim the cluster-level table for practical SEO work."""
    result = frame.copy()
    if "Total Search Volume" not in result.columns and "Total_Search_Volume" in result.columns:
        result = result.rename(columns={"Total_Search_Volume": "Total Search Volume"})
    for column in CLUSTER_SUMMARY_COLUMNS:
        if column not in result.columns:
            result[column] = ""
    return result[CLUSTER_SUMMARY_COLUMNS]


def workflow_metadata(
    config: dict[str, Any], input_name: str, processing_mode: str
) -> pd.DataFrame:
    """Small visible provenance sheet required by the review-import workflow."""
    return pd.DataFrame(
        [
            ("Workflow Job ID", config.get("workflow_job_id", "")),
            ("Job", config.get("job_name", "")),
            ("Topic", config.get("topic", {}).get("description", "")),
            ("Input file", input_name),
            ("Processing mode", processing_mode),
        ],
        columns=["Parameter", "Value"],
    )


def build_human_review_queue(
    combined: pd.DataFrame,
    garbage: pd.DataFrame,
    quarantine: pd.DataFrame,
    classification_threshold: float,
    cluster_threshold: float,
    garbage_threshold: float,
) -> pd.DataFrame:
    """Build a bounded review queue with coverage, not only uncertainty.

    A classifier can be confidently wrong.  The queue therefore reserves space
    for high-value phrases, cluster representatives and class-boundary rows in
    addition to low-confidence predictions.
    """
    pieces: list[pd.DataFrame] = []

    def with_reason(frame: pd.DataFrame, reason: str, order: int, limit: int) -> None:
        if frame.empty:
            return
        selected = frame.head(limit).copy()
        selected["Review reason"] = reason
        selected["_review_order"] = order
        pieces.append(selected)

    def ranked(frame: pd.DataFrame, columns: list[str], ascending: list[bool]) -> pd.DataFrame:
        result = frame.copy()
        if result.empty:
            return result
        result["_review_volume"] = pd.to_numeric(
            result.get("Search Volume", pd.Series(0, index=result.index)), errors="coerce"
        ).fillna(0)
        return result.sort_values(columns, ascending=ascending, kind="stable")

    with_reason(
        quarantine,
        "Relevance quarantine: probability is between review and garbage thresholds",
        0,
        100,
    )

    confidence = pd.to_numeric(combined.get("Classification confidence", 0), errors="coerce").fillna(0)
    cluster_probability = pd.to_numeric(combined.get("Cluster probability", 0), errors="coerce").fillna(0)
    uncertain = combined[
        confidence.lt(classification_threshold)
        | combined.get("Cluster", pd.Series("Unclustered", index=combined.index)).eq("Unclustered")
        | cluster_probability.lt(cluster_threshold)
    ].copy()
    uncertain["_confidence"] = confidence.loc[uncertain.index]
    uncertain["_cluster_probability"] = cluster_probability.loc[uncertain.index]
    uncertain["_review_volume"] = pd.to_numeric(
        uncertain.get("Search Volume", pd.Series(0, index=uncertain.index)), errors="coerce"
    ).fillna(0)
    uncertain = uncertain.sort_values(
        ["_confidence", "_cluster_probability", "_review_volume"],
        ascending=[True, True, False],
        kind="stable",
    )
    with_reason(uncertain, "Low confidence, unclustered, or weak cluster", 1, 150)

    high_risk_garbage = ranked(garbage, ["Topic relevance", "_review_volume"], [False, False])
    with_reason(high_risk_garbage, "Garbage near the topic-relevance boundary", 2, 75)

    if "P(garbage)" in combined.columns:
        relevance_boundary = combined.copy()
        relevance_boundary["_relevance_margin"] = (
            pd.to_numeric(relevance_boundary["P(garbage)"], errors="coerce").fillna(0)
            - garbage_threshold
        ).abs()
        relevance_boundary = ranked(relevance_boundary, ["_relevance_margin", "_review_volume"], [True, False])
        with_reason(relevance_boundary, "Topic relevance probability boundary", 2, 100)

    if {"P(commercial)", "P(informational)"}.issubset(combined.columns):
        boundary = combined.copy()
        boundary["_intent_margin"] = (
            pd.to_numeric(boundary["P(commercial)"], errors="coerce").fillna(0)
            - pd.to_numeric(boundary["P(informational)"], errors="coerce").fillna(0)
        ).abs()
        boundary = ranked(boundary, ["_intent_margin", "_review_volume"], [True, False])
        with_reason(boundary, "Commercial/informational probability boundary", 3, 75)

        if "Weak question without transaction" in combined.columns:
            weak_questions = combined[
                combined["Weak question without transaction"].fillna(False).astype(bool)
                & combined["Intent"].eq("commercial")
            ].copy()
            weak_questions["_weak_question_gap"] = (
                pd.to_numeric(weak_questions["P(informational)"], errors="coerce").fillna(0)
                - pd.to_numeric(weak_questions["P(commercial)"], errors="coerce").fillna(0)
            )
            weak_questions = ranked(
                weak_questions, ["_weak_question_gap", "_review_volume"], [False, False]
            )
            with_reason(
                weak_questions,
                "Weak question without a transaction signal remained commercial",
                3,
                50,
            )

    if "Intent family conflict" in combined.columns:
        family_conflicts = combined[
            combined["Intent family conflict"].fillna(False).astype(bool)
        ].copy()
        family_conflicts = ranked(family_conflicts, ["_review_volume"], [False])
        with_reason(
            family_conflicts,
            "Reviewed lexical family conflicts with stronger phrase-level evidence",
            3,
            75,
        )

    if "Policy adjustment" in combined.columns:
        policy_conflicts = combined[pd.to_numeric(combined["Policy adjustment"], errors="coerce").abs().ge(0.01)].copy()
        policy_conflicts = ranked(policy_conflicts, ["_review_volume"], [False])
        with_reason(policy_conflicts, "Topic-policy and classifier conflict", 4, 100)

    high_value = ranked(combined, ["_review_volume"], [False])
    with_reason(high_value, "High-priority phrase", 5, 50)

    if not combined.empty and {"Intent", "Cluster"}.issubset(combined.columns):
        cluster_sizes = (
            combined[combined["Cluster"].ne("Unclustered")]
            .groupby(["Intent", "Cluster"], dropna=False)
            .size()
            .sort_values(ascending=False)
        )
        representatives: list[pd.DataFrame] = []
        for (intent, cluster), _ in cluster_sizes.items():
            candidates = combined[combined["Intent"].eq(intent) & combined["Cluster"].eq(cluster)].copy()
            candidates["_review_volume"] = pd.to_numeric(
                candidates.get("Search Volume", pd.Series(0, index=candidates.index)),
                errors="coerce",
            ).fillna(0)
            representatives.append(candidates.nlargest(2, "_review_volume"))
            if sum(len(part) for part in representatives) >= 75:
                break
        if representatives:
            with_reason(pd.concat(representatives, ignore_index=True), "Representative high-priority phrase in cluster", 6, 50)

    if not pieces:
        return pd.DataFrame()
    review = pd.concat(pieces, ignore_index=True, sort=False)
    key = review.get("Normalized", review.get("Phrase", pd.Series("", index=review.index))).astype(str)
    reason_map = (
        pd.DataFrame({"key": key, "reason": review["Review reason"]})
        .groupby("key", sort=False)["reason"]
        .agg(lambda values: " | ".join(dict.fromkeys(values)))
    )
    review["_review_key"] = key
    review = review.sort_values(["_review_order", "_review_volume"], ascending=[True, False], kind="stable")
    review = review.drop_duplicates("_review_key", keep="first").head(300).copy()
    review["Review reason"] = review["_review_key"].map(reason_map)
    review = review.drop(columns=[column for column in review.columns if column.startswith("_review_") or column in {"_confidence", "_cluster_probability", "_intent_margin", "_weak_question_gap"}])
    for column in ("Correct Intent", "Correct Cluster", "Reviewer Notes"):
        review[column] = ""
    return review


def run_pipeline(
    input_value: str,
    config_path: Path,
    labels_path: Path | None,
    output_path: Path | None,
) -> Path:
    input_path = resolve_input(input_value)
    config = load_config(config_path)
    source = read_table(input_path, config.get("source_sheet", 0))
    phrase_column = choose_phrase_column(source, config.get("phrase_column"))
    phrase_table = build_phrase_table(source, phrase_column, config.get("frequency_column"))
    batch_size = int(config.get("embedding_batch_size", config["clustering"].get("batch_size", 256)))
    model, execution_device, model_path = load_local_embedding_model(config)
    vectors = encode_queries(model, phrase_table["Normalized"], batch_size, str(config.get("embedding_prefix", "query: ")))

    model_labels = load_model_labels(labels_path)
    try:
        classified, validation = supervised_classification(
            phrase_table, vectors, model_labels, config, model, batch_size
        )
    except ValueError as error:
        classified, validation = seed_classification(
            phrase_table, vectors, model, config, batch_size
        )
        validation = pd.concat(
            [
                validation,
                pd.DataFrame([("Supervised fallback reason", str(error))], columns=["Metric", "Value"]),
            ],
            ignore_index=True,
        )

    intent_frames: dict[str, pd.DataFrame] = {}
    summaries = []
    for intent, title in (
        ("commercial", "Commercial clusters"),
        ("informational", "Informational clusters"),
    ):
        positions = np.flatnonzero(classified["Intent"].to_numpy() == intent)
        clustered = cluster_subset(classified.iloc[positions], vectors[positions], config)
        intent_frames[title] = clustered
        if len(clustered):
            summaries.append(build_cluster_summary(clustered, intent))
    garbage = classified[classified["Intent"].eq("garbage")].copy()
    quarantine = classified[
        classified.get(
            "Relevance quarantine", pd.Series(False, index=classified.index)
        ).fillna(False).astype(bool)
    ].copy()
    combined = pd.concat(intent_frames.values(), ignore_index=True)
    review_threshold = float(config["clustering"].get("review_threshold", 0.35))
    classification_threshold = float(config.get("classification_review_threshold", 0.65))
    effective_garbage_threshold = (
        float(classified["Garbage threshold"].iloc[0])
        if "Garbage threshold" in classified.columns and not classified.empty
        else relevance_settings(config)["garbage_threshold"]
    )
    review = build_human_review_queue(
        combined,
        garbage,
        quarantine,
        classification_threshold,
        review_threshold,
        effective_garbage_threshold,
    )

    cluster_summary = (
        pd.concat(summaries, ignore_index=True)
        if summaries
        else pd.DataFrame(columns=["Intent", "Cluster", "Phrases"])
    )
    family_audit = build_intent_family_audit([classified], config)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = output_path or OUTPUT_DIR / f"{input_path.stem}_clustered.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        compact_result_frame(intent_frames["Commercial clusters"], "core").to_excel(
            writer, sheet_name="Commercial clusters", index=False
        )
        compact_result_frame(intent_frames["Informational clusters"], "core").to_excel(
            writer, sheet_name="Informational clusters", index=False
        )
        compact_result_frame(garbage, "garbage").to_excel(
            writer, sheet_name="Garbage", index=False
        )
        compact_result_frame(review, "human_review").to_excel(
            writer, sheet_name="Human review", index=False
        )
        compact_cluster_summary(cluster_summary).to_excel(
            writer, sheet_name="Cluster summary", index=False
        )
        workflow_metadata(config, input_path.name, "Standard").to_excel(
            writer, sheet_name="_Workflow", index=False
        )
        style_workbook(writer)
    return output_path


def fit_large_classifier(
    labels: pd.DataFrame,
    model: SentenceTransformer,
    batch_size: int,
    config: dict[str, Any],
) -> dict[str, Any]:
    return fit_two_stage_classifier(labels, model, batch_size, config)


def large_part_path(parts_dir: Path, number: int) -> Path:
    return parts_dir / f"part-{number:06d}.csv.gz"


def stable_sample_mask(values: pd.Series, divisor: int = 40) -> np.ndarray:
    return np.asarray(
        [int(hashlib.blake2b(str(value).encode("utf-8"), digest_size=4).hexdigest(), 16) % divisor == 0 for value in values]
    )


def build_large_cluster_relevance_audit(
    parts_dir: Path,
    cluster_summary: pd.DataFrame,
    audit_path: Path,
    workflow_job_id: str,
) -> dict[str, Any]:
    """Persist compact representatives for the main model's topic audit."""
    if cluster_summary.empty:
        payload = {
            "workflow_job_id": workflow_job_id,
            "allowed_labels": ["relevant", "garbage", "mixed"],
            "clusters": [],
        }
        audit_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return {"clusters": 0, "audit_path": str(audit_path.resolve())}
    eligible = cluster_summary[
        cluster_summary["Intent"].isin(["commercial", "informational"])
        & cluster_summary["Cluster"].astype(str).ne("Unclustered")
    ].copy()
    eligible = eligible.sort_values("Phrases", ascending=False, kind="stable")
    records: list[dict[str, Any]] = []
    lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for number, row in enumerate(eligible.itertuples(index=False), start=1):
        intent = str(getattr(row, "Intent"))
        cluster = str(getattr(row, "Cluster"))
        record = {
            "id": f"RC{number:04d}",
            "current_intent": intent,
            "cluster": cluster,
            "rows": int(getattr(row, "Phrases")),
            "average_confidence": round(float(getattr(row, "Average_Confidence", 0.0)), 4),
            "average_garbage_probability": round(
                float(getattr(row, "Average_Garbage_Probability", 0.0)), 4
            ),
            "representatives": [],
        }
        records.append(record)
        lookup[(intent, cluster)] = record
    candidate_rows: dict[tuple[str, str], list[dict[str, Any]]] = {
        key: [] for key in lookup
    }
    metric_samples: dict[tuple[str, str], list[tuple[float, float]]] = {
        key: [] for key in lookup
    }
    for part in sorted(parts_dir.glob("part-*.csv.gz")):
        frame = pd.read_csv(
            part,
            compression="gzip",
            usecols=lambda column: column
            in {
                "Phrase",
                "Intent",
                "Cluster",
                "P(garbage)",
                "Classification confidence",
                "Cluster probability",
                "Search Volume",
            },
        )
        frame = frame[frame["Intent"].isin(["commercial", "informational"])]
        if frame.empty:
            continue
        for column in (
            "P(garbage)",
            "Classification confidence",
            "Cluster probability",
            "Search Volume",
        ):
            if column not in frame.columns:
                frame[column] = 0.0
            frame[column] = pd.to_numeric(frame[column], errors="coerce").fillna(0.0)
        for (intent, cluster), group in frame.groupby(["Intent", "Cluster"], sort=False):
            key = (str(intent), str(cluster))
            if key not in lookup:
                continue
            local = pd.concat(
                [
                    group.nlargest(2, "Cluster probability"),
                    group.nlargest(1, "Search Volume"),
                    group.nlargest(1, "P(garbage)"),
                    group.nsmallest(1, "Cluster probability"),
                ],
                ignore_index=True,
            ).drop_duplicates("Phrase")
            candidate_rows[key].extend(local.to_dict("records"))
            sampled = group.loc[
                stable_sample_mask(group["Phrase"], divisor=20),
                ["Cluster probability", "P(garbage)"],
            ]
            if sampled.empty:
                sampled = local[["Cluster probability", "P(garbage)"]]
            metric_samples[key].extend(
                (float(cluster_probability), float(garbage_probability))
                for cluster_probability, garbage_probability in sampled.itertuples(
                    index=False, name=None
                )
            )
    for key, record in lookup.items():
        candidates = pd.DataFrame(candidate_rows[key])
        if candidates.empty:
            continue
        candidates = candidates.drop_duplicates("Phrase")
        metrics = np.asarray(metric_samples[key], dtype=float)
        if metrics.size:
            record["outlier_cluster_probability_max"] = round(
                float(np.quantile(metrics[:, 0], 0.05)), 6
            )
            record["outlier_garbage_probability_min"] = round(
                float(np.quantile(metrics[:, 1], 0.90)), 6
            )
        evidence: list[dict[str, str]] = []
        seen_phrases: set[str] = set()
        selections = [
            ("cluster_center", candidates.nlargest(2, "Cluster probability")),
            ("high_volume", candidates.nlargest(1, "Search Volume")),
            ("relevance_risk", candidates.nlargest(1, "P(garbage)")),
            ("cluster_boundary", candidates.nsmallest(1, "Cluster probability")),
        ]
        for role, selected in selections:
            for phrase in selected["Phrase"].astype(str):
                if phrase in seen_phrases:
                    continue
                item = {"role": role, "phrase": phrase}
                suffix = {"relevance_risk": "R", "cluster_boundary": "B"}.get(role)
                if suffix:
                    item["review_id"] = f"{record['id']}-{suffix}"
                evidence.append(item)
                seen_phrases.add(phrase)
                if len(evidence) >= 5:
                    break
            if len(evidence) >= 5:
                break
        record["representatives"] = [item["phrase"] for item in evidence]
        record["representative_evidence"] = evidence
    payload = {
        "workflow_job_id": workflow_job_id,
        "allowed_labels": ["relevant", "garbage", "mixed"],
        "clusters": records,
    }
    audit_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"clusters": len(records), "audit_path": str(audit_path.resolve())}


def build_cluster_outlier_model(
    records: list[dict[str, Any]],
    decisions: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    """Learn a conservative row-level relevance model from the existing audit."""
    examples: list[dict[str, str]] = []
    exact_labels: dict[str, str] = {}
    for record in records:
        cluster_decision = str(decisions.get(str(record.get("id")), "")).lower()
        for evidence in record.get("representative_evidence", []):
            if not isinstance(evidence, dict):
                continue
            phrase = str(evidence.get("phrase", "")).strip()
            normalized = normalize(phrase)
            if not normalized:
                continue
            review_id = str(evidence.get("review_id", "")).strip()
            explicit_label = str(decisions.get(review_id, "")).lower() if review_id else ""
            if explicit_label in {"relevant", "garbage"}:
                examples.append({"Phrase": phrase, "Normalized": normalized, "Label": explicit_label})
                exact_labels[normalized] = explicit_label
            elif (
                cluster_decision == "relevant"
                and str(evidence.get("role", "")) in {"cluster_center", "high_volume"}
            ):
                examples.append({"Phrase": phrase, "Normalized": normalized, "Label": "relevant"})

    training = pd.DataFrame(examples).drop_duplicates(["Normalized", "Label"])
    if not training.empty:
        conflicts = training.groupby("Normalized")["Label"].nunique()
        conflicting_keys = set(conflicts[conflicts.gt(1)].index)
        training = training[~training["Normalized"].isin(conflicting_keys)].copy()
        for key in conflicting_keys:
            exact_labels.pop(str(key), None)
    counts = training["Label"].value_counts().to_dict() if not training.empty else {}
    result: dict[str, Any] = {
        "enabled": False,
        "exact_labels": exact_labels,
        "training_examples": int(len(training)),
        "relevant_examples": int(counts.get("relevant", 0)),
        "garbage_examples": int(counts.get("garbage", 0)),
        "threshold": float(
            config.get("relevance", {}).get("outlier_garbage_threshold", 0.80)
            if isinstance(config.get("relevance", {}), dict)
            else 0.80
        ),
    }
    if counts.get("relevant", 0) < 20 or counts.get("garbage", 0) < 8:
        return result

    batch_size = int(config.get("embedding_batch_size", config["clustering"].get("batch_size", 256)))
    embedding_model, device, model_path = load_local_embedding_model(config)
    vectors = encode_queries(
        embedding_model,
        training["Normalized"],
        batch_size,
        str(config.get("embedding_prefix", "query: ")),
    )
    structural = TfidfVectorizer(
        analyzer="char_wb",
        ngram_range=(3, 5),
        min_df=1,
        max_features=20000,
        sublinear_tf=True,
    )
    structural_features = structural.fit_transform(training["Normalized"])
    features = hstack([csr_matrix(vectors), structural_features * 0.35], format="csr")
    targets = training["Label"].to_numpy()
    folds = min(5, int(counts["relevant"]), int(counts["garbage"]))
    out_of_fold = np.zeros(len(targets), dtype=float)
    splitter = StratifiedKFold(n_splits=folds, shuffle=True, random_state=42)
    for train_index, test_index in splitter.split(features, targets):
        validation_model = LogisticRegression(
            max_iter=2000,
            class_weight="balanced",
            random_state=42,
        ).fit(features[train_index], targets[train_index])
        garbage_index = list(validation_model.classes_).index("garbage")
        out_of_fold[test_index] = validation_model.predict_proba(features[test_index])[
            :, garbage_index
        ]
    actual_garbage = targets == "garbage"
    actual_relevant = targets == "relevant"
    safe_thresholds: list[tuple[float, float, float, float]] = []
    configured_threshold = float(result["threshold"])
    for threshold in np.linspace(configured_threshold, 0.98, 37):
        predicted = out_of_fold >= threshold
        predicted_count = int(predicted.sum())
        if predicted_count < 3:
            continue
        true_positive = int(np.sum(predicted & actual_garbage))
        false_positive = int(np.sum(predicted & actual_relevant))
        precision = true_positive / predicted_count
        recall = true_positive / int(actual_garbage.sum())
        false_positive_rate = false_positive / int(actual_relevant.sum())
        if precision >= 0.90 and false_positive_rate <= 0.02:
            safe_thresholds.append((recall, precision, -false_positive_rate, float(threshold)))
    if not safe_thresholds:
        result["validation"] = "disabled_no_safe_out_of_fold_threshold"
        return result
    recall, precision, negative_fpr, selected_threshold = max(safe_thresholds)
    result.update(
        {
            "threshold": selected_threshold,
            "validation": "out_of_fold",
            "validation_folds": folds,
            "validation_precision": round(precision, 4),
            "validation_recall": round(recall, 4),
            "validation_false_positive_rate": round(-negative_fpr, 4),
        }
    )
    classifier = LogisticRegression(
        max_iter=2000,
        class_weight="balanced",
        random_state=42,
    ).fit(features, targets)
    result.update(
        {
            "enabled": True,
            "embedding_model": embedding_model,
            "embedding_device": device,
            "embedding_model_path": str(model_path),
            "embedding_batch_size": batch_size,
            "structural": structural,
            "classifier": classifier,
        }
    )
    return result


def predict_cluster_outlier_garbage(
    bundle: dict[str, Any],
    phrases: pd.Series,
    config: dict[str, Any],
) -> np.ndarray:
    if not bundle.get("enabled") or phrases.empty:
        return np.zeros(len(phrases), dtype=float)
    normalized = phrases.map(normalize)
    vectors = encode_queries(
        bundle["embedding_model"],
        normalized,
        int(bundle["embedding_batch_size"]),
        str(config.get("embedding_prefix", "query: ")),
    )
    features = hstack(
        [
            csr_matrix(vectors),
            bundle["structural"].transform(normalized) * 0.35,
        ],
        format="csr",
    )
    probabilities = bundle["classifier"].predict_proba(features)
    garbage_index = list(bundle["classifier"].classes_).index("garbage")
    return probabilities[:, garbage_index]


def apply_large_cluster_relevance_decisions(
    output_dir: Path,
    config_path: Path,
) -> dict[str, Any]:
    """Apply model-reviewed cluster relevance without recomputing embeddings."""
    audit_path = output_dir / "cluster_relevance_audit.json"
    labels_path = output_dir / "cluster_relevance_labels.json"
    manifest_path = output_dir / "manifest.json"
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    decisions = (
        json.loads(labels_path.read_text(encoding="utf-8"))
        if labels_path.is_file()
        else {"labels": {}}
    )
    decision_by_id = decisions.get("labels", {}) if isinstance(decisions, dict) else {}
    records = audit.get("clusters", []) if isinstance(audit, dict) else []
    required_ids = [str(record["id"]) for record in records]
    required_ids.extend(
        str(evidence["review_id"])
        for record in records
        for evidence in record.get("representative_evidence", [])
        if isinstance(evidence, dict) and evidence.get("review_id")
    )
    missing = [decision_id for decision_id in required_ids if decision_id not in decision_by_id]
    if missing:
        raise ValueError(f"Cluster relevance decisions are incomplete: {len(missing)} missing.")
    allowed = {"relevant", "garbage", "mixed"}
    cluster_decisions: dict[tuple[str, str], str] = {}
    for record in records:
        decision = str(decision_by_id[record["id"]]).strip().lower()
        if decision not in allowed:
            raise ValueError(f"Invalid cluster decision for {record['id']}: {decision}")
        cluster_decisions[(str(record["current_intent"]), str(record["cluster"]))] = decision

    config = load_config(config_path)
    outlier_bundle = build_cluster_outlier_model(records, decision_by_id, config)
    outlier_thresholds = {
        (str(record["current_intent"]), str(record["cluster"])): (
            float(record.get("outlier_cluster_probability_max", float("-inf"))),
            float(record.get("outlier_garbage_probability_min", float("inf"))),
        )
        for record in records
    }
    exact_outlier_labels = outlier_bundle.get("exact_labels", {})

    parts_dir = output_dir / "parts"
    prior_review_path = output_dir / "uncertain_review.csv"
    try:
        prior_review = (
            pd.read_csv(prior_review_path, encoding="utf-8-sig")
            if prior_review_path.is_file() and prior_review_path.stat().st_size
            else pd.DataFrame()
        )
    except pd.errors.EmptyDataError:
        prior_review = pd.DataFrame()
    mixed_candidates: list[pd.DataFrame] = []
    summaries: list[dict[str, Any]] = []
    moved_to_garbage = 0
    outlier_candidates_evaluated = 0
    outlier_rows_moved = 0
    outlier_garbage_keys: set[str] = set()
    for part in sorted(parts_dir.glob("part-*.csv.gz")):
        frame = pd.read_csv(part, compression="gzip")
        audit_source_column = "Intent before cluster audit"
        if audit_source_column not in frame.columns:
            frame[audit_source_column] = frame["Intent"].astype(str)
        original_intent = frame[audit_source_column].astype(str).copy()
        # Restore the preliminary prediction first so an interrupted finalize
        # can be retried safely without accumulating partial decisions.
        frame["Intent"] = original_intent
        if "Decision basis" in frame.columns:
            stale_audit_basis = frame["Decision basis"].fillna("").astype(str).str.startswith(
                "Main-model cluster"
            )
            frame.loc[stale_audit_basis, "Decision basis"] = np.nan
        frame["Outlier garbage probability"] = np.nan
        decisions_for_rows = np.asarray(
            [
                cluster_decisions.get((intent, str(cluster)), "relevant")
                for intent, cluster in zip(original_intent, frame["Cluster"])
            ],
            dtype=object,
        )
        garbage_mask = decisions_for_rows == "garbage"
        mixed_mask = decisions_for_rows == "mixed"
        moved_to_garbage += int(garbage_mask.sum())
        frame.loc[garbage_mask, "Intent"] = "garbage"
        frame.loc[garbage_mask, "Decision basis"] = "Main-model cluster relevance audit"

        normalized = frame.get("Normalized", frame["Phrase"].map(normalize)).astype(str)
        exact_labels = normalized.map(exact_outlier_labels).fillna("")
        exact_garbage_mask = exact_labels.eq("garbage").to_numpy() & ~garbage_mask
        exact_relevant_mask = exact_labels.eq("relevant").to_numpy()
        frame.loc[exact_garbage_mask, "Intent"] = "garbage"
        frame.loc[exact_garbage_mask, "Outlier garbage probability"] = 1.0
        frame.loc[exact_garbage_mask, "Decision basis"] = "Main-model cluster outlier label"

        threshold_pairs = [
            outlier_thresholds.get((intent, str(cluster)), (float("-inf"), float("inf")))
            for intent, cluster in zip(original_intent, frame["Cluster"])
        ]
        cluster_probability_max = np.asarray([pair[0] for pair in threshold_pairs], dtype=float)
        garbage_probability_min = np.asarray([pair[1] for pair in threshold_pairs], dtype=float)
        quarantine = frame.get(
            "Relevance quarantine", pd.Series(False, index=frame.index)
        ).fillna(False).astype(str).str.lower().isin({"true", "1", "yes"}).to_numpy()
        cluster_probability = pd.to_numeric(
            frame.get("Cluster probability", 0.0), errors="coerce"
        ).fillna(0.0).to_numpy()
        garbage_probability = pd.to_numeric(
            frame.get("P(garbage)", 0.0), errors="coerce"
        ).fillna(0.0).to_numpy()
        outlier_candidate_mask = (
            (decisions_for_rows == "relevant")
            & quarantine
            & (cluster_probability <= cluster_probability_max)
            & (garbage_probability >= garbage_probability_min)
            & ~exact_garbage_mask
            & ~exact_relevant_mask
        )
        candidate_positions = np.flatnonzero(outlier_candidate_mask)
        outlier_candidates_evaluated += int(len(candidate_positions))
        learned_garbage_mask = np.zeros(len(frame), dtype=bool)
        if len(candidate_positions) and outlier_bundle.get("enabled"):
            learned_probabilities = predict_cluster_outlier_garbage(
                outlier_bundle,
                frame.iloc[candidate_positions]["Phrase"],
                config,
            )
            frame.iloc[
                candidate_positions,
                frame.columns.get_loc("Outlier garbage probability"),
            ] = learned_probabilities
            selected_positions = candidate_positions[
                learned_probabilities >= float(outlier_bundle["threshold"])
            ]
            learned_garbage_mask[selected_positions] = True
            frame.loc[learned_garbage_mask, "Intent"] = "garbage"
            frame.loc[learned_garbage_mask, "Decision basis"] = (
                "Main-model cluster outlier classifier"
            )
        outlier_garbage_mask = exact_garbage_mask | learned_garbage_mask
        outlier_rows_moved += int(outlier_garbage_mask.sum())
        outlier_garbage_keys.update(normalized.loc[outlier_garbage_mask].astype(str))

        effective_mixed_mask = mixed_mask & ~outlier_garbage_mask
        if effective_mixed_mask.any():
            mixed = frame.loc[effective_mixed_mask].copy()
            mixed["_audit_original_intent"] = original_intent.loc[mixed.index]
            mixed["Review reason"] = "Mixed relevance cluster after main-model audit"
            mixed_review = (
                mixed.sort_values(
                    ["P(garbage)", "Classification confidence"],
                    ascending=[False, True],
                    kind="stable",
                )
                .groupby(["_audit_original_intent", "Cluster"], group_keys=False)
                .head(3)
                .drop(columns=["_audit_original_intent"])
            )
            mixed_candidates.append(mixed_review)
        frame.to_csv(part, index=False, compression="gzip", encoding="utf-8")
        grouped = (
            frame.groupby(["Intent", "Cluster"], dropna=False)
            .agg(
                Phrases=("Phrase", "size"),
                Total_Search_Volume=("Search Volume", "sum"),
                Average_Confidence=("Classification confidence", "mean"),
                Average_Garbage_Probability=("P(garbage)", "mean"),
            )
            .reset_index()
        )
        summaries.extend(grouped.to_dict("records"))

    summary = pd.DataFrame(summaries)
    if not summary.empty:
        summary = (
            summary.groupby(["Intent", "Cluster"], dropna=False)
            .agg(
                Phrases=("Phrases", "sum"),
                Total_Search_Volume=("Total_Search_Volume", "sum"),
                Average_Confidence=("Average_Confidence", "mean"),
                Average_Garbage_Probability=("Average_Garbage_Probability", "mean"),
            )
            .reset_index()
            .sort_values("Total_Search_Volume", ascending=False)
        )
    mixed_review = (
        pd.concat(mixed_candidates, ignore_index=True, sort=False)
        if mixed_candidates
        else pd.DataFrame()
    )
    if not prior_review.empty and {"Intent", "Cluster"}.issubset(prior_review.columns):
        prior_decisions = [
            cluster_decisions.get((str(intent), str(cluster)), "relevant")
            for intent, cluster in zip(prior_review["Intent"], prior_review["Cluster"])
        ]
        # Garbage clusters need no manual confirmation, while mixed clusters
        # are represented by the smaller, cluster-balanced sample above.
        prior_review = prior_review[np.asarray(prior_decisions, dtype=object) == "relevant"]
    if not prior_review.empty and outlier_garbage_keys:
        prior_keys = prior_review.get(
            "Normalized", prior_review["Phrase"].map(normalize)
        ).astype(str)
        prior_review = prior_review[~prior_keys.isin(outlier_garbage_keys)]
    review = pd.concat([mixed_review, prior_review], ignore_index=True, sort=False)
    if not review.empty:
        key = review.get("Normalized", review["Phrase"]).astype(str)
        review = review.assign(_review_key=key).drop_duplicates("_review_key").head(300)
        review = review.drop(columns="_review_key")
    review.to_csv(output_dir / "uncertain_review.csv", index=False, encoding="utf-8-sig")
    summary.to_csv(output_dir / "cluster_summary.csv", index=False, encoding="utf-8-sig")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    report_path = Path(str(manifest["excel_result"]))
    manifest["status"] = "completed"
    cluster_values = [str(decision_by_id[record["id"]]).lower() for record in records]
    manifest["cluster_relevance_audit"] = {
        "clusters": len(records),
        "relevant": sum(value == "relevant" for value in cluster_values),
        "garbage": sum(value == "garbage" for value in cluster_values),
        "mixed": sum(value == "mixed" for value in cluster_values),
        "rows_moved_to_garbage": moved_to_garbage,
        "outlier_training_examples": int(outlier_bundle.get("training_examples", 0)),
        "outlier_garbage_examples": int(outlier_bundle.get("garbage_examples", 0)),
        "outlier_classifier_enabled": bool(outlier_bundle.get("enabled")),
        "outlier_classifier_threshold": float(outlier_bundle.get("threshold", 0.80)),
        "outlier_validation": outlier_bundle.get("validation", "insufficient_labels"),
        "outlier_validation_precision": outlier_bundle.get("validation_precision"),
        "outlier_validation_recall": outlier_bundle.get("validation_recall"),
        "outlier_candidates_evaluated": outlier_candidates_evaluated,
        "outlier_rows_moved_to_garbage": outlier_rows_moved,
    }
    family_audit = build_intent_family_audit(
        (
            pd.read_csv(part, compression="gzip", usecols=["Phrase", "Intent"])
            for part in sorted(parts_dir.glob("part-*.csv.gz"))
        ),
        config,
    )
    family_audit_path = output_dir / "intent_family_audit.csv"
    family_audit.to_csv(family_audit_path, index=False, encoding="utf-8-sig")
    manifest["intent_family_audit"] = {
        "rules": int(len(family_audit)),
        "warnings": int(family_audit["Audit status"].eq("review").sum())
        if not family_audit.empty
        else 0,
        "path": str(family_audit_path.resolve()),
    }
    manifest["intent_counts"] = export_large_result_workbook(
        parts_dir, report_path, summary, review, family_audit, config, manifest
    )
    manifest["excel_summary"] = str(report_path.resolve())
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def _excel_value(value: object) -> object:
    """Convert pandas missing values to cells that Excel can store."""
    return None if pd.isna(value) else value


def export_large_result_workbook(
    parts_dir: Path,
    report_path: Path,
    cluster_summary: pd.DataFrame,
    uncertain_review: pd.DataFrame,
    intent_family_audit: pd.DataFrame,
    config: dict[str, Any],
    metadata: dict[str, Any],
) -> dict[str, int]:
    """Create the same review workbook contract as the standard pipeline.

    Classified rows are streamed from the compressed parts, so exporting a
    large core does not require loading the whole result into RAM.  Excel has a
    hard per-sheet row limit; exceptionally large intent groups continue on
    numbered sheets while preserving the familiar first-sheet names.
    """
    parts = sorted(parts_dir.glob("part-*.csv.gz"))
    if not parts:
        raise FileNotFoundError(f"No classified parts found in {parts_dir}")
    source_columns = list(pd.read_csv(parts[0], compression="gzip", nrows=0).columns)
    output_columns = {
        "commercial": CORE_RESULT_COLUMNS,
        "informational": CORE_RESULT_COLUMNS,
        "garbage": GARBAGE_RESULT_COLUMNS,
        "human_review": HUMAN_REVIEW_RESULT_COLUMNS,
    }
    workbook = Workbook(write_only=True)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sheet_states: dict[str, tuple[object, int, int]] = {}
    total_by_intent = {
        "commercial": 0,
        "informational": 0,
        "garbage": 0,
        "human_review": 0,
    }
    sheet_names = {
        "commercial": "Commercial clusters",
        "informational": "Informational clusters",
        "garbage": "Garbage",
        "human_review": "Human review",
    }
    human_review_seen: set[str] = set()

    def create_sheet(base_name: str, part_number: int, headers: list[str]) -> object:
        title = base_name if part_number == 1 else f"{base_name} {part_number}"
        sheet = workbook.create_sheet(title)
        header_cells = []
        for header in headers:
            cell = WriteOnlyCell(sheet, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_cells.append(cell)
        sheet.append(header_cells)
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 2, 14), 42)
        return sheet

    def target_sheet(intent: str) -> tuple[object, int, int]:
        state = sheet_states.get(intent)
        if state is None or state[1] >= EXCEL_MAX_DATA_ROWS:
            part_number = 1 if state is None else state[2] + 1
            state = (
                create_sheet(sheet_names[intent], part_number, output_columns[intent]),
                0,
                part_number,
            )
            sheet_states[intent] = state
        return state

    for part in parts:
        frame = pd.read_csv(part, compression="gzip")
        for column in source_columns:
            if column not in frame.columns:
                frame[column] = ""
        for intent in total_by_intent:
            subset = frame[frame["Intent"].eq(intent)]
            if subset.empty:
                continue
            exported = compact_result_frame(subset, intent)
            for row in exported.itertuples(index=False, name=None):
                sheet, written, sheet_number = target_sheet(intent)
                sheet.append([_excel_value(value) for value in row])
                sheet_states[intent] = (sheet, written + 1, sheet_number)
                total_by_intent[intent] += 1
            if intent == "human_review":
                human_review_seen.update(
                    subset.get("Normalized", subset["Phrase"]).astype(str)
                )

    for intent in total_by_intent:
        if intent not in sheet_states:
            sheet_states[intent] = (
                create_sheet(sheet_names[intent], 1, output_columns[intent]),
                0,
                1,
            )

    review = uncertain_review.copy()
    for column in ("Correct Intent", "Reviewer Notes"):
        if column not in review.columns:
            review[column] = ""
    if not review.empty:
        for column in source_columns:
            if column not in review.columns:
                review[column] = ""
        review_keys = review.get("Normalized", review["Phrase"]).astype(str)
        review = review[~review_keys.isin(human_review_seen)]
        exported_review = compact_result_frame(review, "human_review")
        for row in exported_review.itertuples(index=False, name=None):
            sheet, written, sheet_number = target_sheet("human_review")
            sheet.append([_excel_value(value) for value in row])
            sheet_states["human_review"] = (sheet, written + 1, sheet_number)
        total_by_intent["human_review"] += len(review)

    def append_small_sheet(name: str, frame: pd.DataFrame) -> None:
        sheet = create_sheet(name, 1, list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            sheet.append([_excel_value(value) for value in row])
        sheet.auto_filter.ref = f"A1:{get_column_letter(max(len(frame.columns), 1))}{len(frame) + 1}"

    append_small_sheet("Cluster summary", compact_cluster_summary(cluster_summary))
    append_small_sheet(
        "_Workflow",
        workflow_metadata(
            config,
            str(metadata.get("input", "")),
            "Large / streamed",
        ),
    )

    for intent, (sheet, written, _) in sheet_states.items():
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(output_columns[intent]))}{written + 1}"
        )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)
    return total_by_intent


def run_large_pipeline(
    input_value: str,
    config_path: Path,
    labels_path: Path,
    output_dir: Path,
    chunk_size: int = 50000,
    source_name: str | None = None,
) -> dict[str, Any]:
    """Chunked classification plus sample-based centroid clustering for large cores."""

    input_path = resolve_input(input_value)
    if input_path.suffix.lower() not in {".csv", ".tsv"}:
        raise ValueError("run-large accepts CSV or TSV only. Export Excel first; do not create a converted copy beside the input.")
    if chunk_size < 1000:
        raise ValueError("chunk-size must be at least 1000.")
    config = load_config(config_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    parts_dir = output_dir / "parts"
    parts_dir.mkdir(parents=True, exist_ok=True)
    header = read_delimited_table(input_path, nrows=0)
    phrase_column = choose_phrase_column(header, config.get("phrase_column"))
    frequency_column = config.get("frequency_column")
    if frequency_column not in header.columns:
        frequency_column = None
    batch_size = int(config.get("embedding_batch_size", config["clustering"].get("batch_size", 256)))
    model, execution_device, model_path = load_local_embedding_model(config)
    embedding_prefix = str(config.get("embedding_prefix", "query: "))
    cache_dir = output_dir / ".embedding-cache"
    if cache_dir.exists():
        import shutil
        shutil.rmtree(cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)
    classifier_bundle = fit_large_classifier(load_model_labels(labels_path), model, batch_size, config)
    policy = config.get("intent_policy", {}) if isinstance(config.get("intent_policy", {}), dict) else {}
    commercial_prototypes = [
        str(value)
        for key in ("commercial_prototypes", "implicit_commercial_prototypes")
        for value in policy.get(key, [])
        if str(value).strip()
    ]
    informational_prototypes = [str(value) for value in policy.get("informational_prototypes", []) if str(value).strip()]
    policy_commercial_vectors = encode_queries(model, commercial_prototypes, batch_size) if commercial_prototypes else None
    policy_informational_vectors = encode_queries(model, informational_prototypes, batch_size) if informational_prototypes else None
    total_rows = 0
    representatives: dict[str, list[pd.DataFrame]] = {"commercial": [], "informational": []}
    use_columns = [phrase_column] + ([frequency_column] if frequency_column else [])
    reader = iter_delimited_chunks(input_path, chunksize=chunk_size, usecols=use_columns)
    for part_number, chunk in enumerate(reader, start=1):
        raw = chunk[phrase_column].dropna().astype(str).str.strip()
        raw = raw[raw.ne("")]
        table = pd.DataFrame({"Phrase": raw, "Source Row": raw.index + 2})
        table["Normalized"] = table["Phrase"].map(normalize)
        if frequency_column:
            table["Search Volume"] = pd.to_numeric(chunk.loc[raw.index, frequency_column], errors="coerce").fillna(0).to_numpy()
        else:
            table["Search Volume"] = 0.0
        table = table.drop_duplicates("Normalized", keep="first")
        if table.empty:
            continue
        vectors = encode_queries(model, table["Normalized"], batch_size, embedding_prefix)
        np.save(cache_dir / f"part-{part_number:06d}.npy", vectors)
        prediction_data = predict_two_stage(
            classifier_bundle,
            table["Normalized"],
            vectors,
            config,
            policy_commercial_vectors,
            policy_informational_vectors,
        )
        probabilities = prediction_data["probabilities"]
        table["Intent"] = prediction_data["prediction"]
        table["Classification confidence"] = prediction_data["confidence"]
        for column_index, label in enumerate(("commercial", "informational", "garbage")):
            table[f"P({label})"] = probabilities[:, column_index]
        table["Topic relevance"] = 1.0 - table["P(garbage)"]
        table["Relevance confidence"] = np.maximum(table["P(garbage)"], table["Topic relevance"])
        table["Policy adjustment"] = prediction_data["policy_adjustment"]
        table["Informational marker hit"] = prediction_data["informational_marker_hit"]
        table["Commercial marker hit"] = prediction_data["commercial_marker_hit"]
        table["Weak question marker hit"] = prediction_data["weak_question_marker_hit"]
        table["Weak question without transaction"] = prediction_data[
            "weak_question_without_transaction"
        ]
        table["Informational family hit"] = prediction_data["informational_family_hit"]
        table["Commercial family hit"] = prediction_data["commercial_family_hit"]
        table["Pre-family intent"] = prediction_data["pre_family_intent"]
        table["Intent family conflict"] = prediction_data["family_conflict"]
        table["Intent family override"] = prediction_data["family_override"]
        table["Decision basis"] = np.where(
            prediction_data["family_override"],
            "Main-model intent family rule",
            "Classifier trained on model labels",
        )
        table["Relevance quarantine"] = prediction_data["relevance_quarantine"]
        table["Garbage threshold"] = prediction_data["garbage_threshold"]
        table["Quarantine threshold"] = prediction_data["quarantine_threshold"]
        table["Part"] = part_number
        table.to_csv(large_part_path(parts_dir, part_number), index=False, compression="gzip", encoding="utf-8")
        total_rows += len(table)
        mask = stable_sample_mask(table["Normalized"])
        for intent in representatives:
            chosen = table.loc[mask & table["Intent"].eq(intent), ["Phrase", "Normalized", "Search Volume", "Intent"]]
            if not chosen.empty:
                representatives[intent].append(chosen.head(1000))
        print(json.dumps({"event": "part_completed", "part": part_number, "rows": int(len(table))}, ensure_ascii=False))

    centroids: dict[str, tuple[np.ndarray, list[str]]] = {}
    cluster_rows: list[pd.DataFrame] = []
    for intent, samples in representatives.items():
        sample = pd.concat(samples, ignore_index=True).drop_duplicates("Normalized").head(10000) if samples else pd.DataFrame()
        if len(sample) < 8:
            continue
        vectors = encode_queries(model, sample["Normalized"], batch_size, embedding_prefix)
        cluster_count = min(max(2, int(np.sqrt(len(sample) / 2))), 80, len(sample))
        fitted = MiniBatchKMeans(n_clusters=cluster_count, random_state=42, batch_size=min(2048, len(sample)), n_init="auto").fit(vectors)
        names = cluster_labels(sample["Phrase"].tolist(), fitted.labels_, config.get("cluster_stop_words", []))
        centroid_names = []
        for cluster_id in range(cluster_count):
            first = next((names[index] for index, value in enumerate(fitted.labels_) if value == cluster_id), f"Cluster {cluster_id}")
            centroid_names.append(first)
        centroids[intent] = (np.asarray(fitted.cluster_centers_), centroid_names)
        sample["Cluster"] = names
        sample["Cluster probability"] = np.max(vectors @ fitted.cluster_centers_.T, axis=1)
        cluster_rows.append(sample)

    review_candidates: dict[str, list[pd.DataFrame]] = {
        "uncertain": [],
        "garbage_boundary": [],
        "relevance_boundary": [],
        "intent_boundary": [],
        "high_value": [],
        "cluster_representative": [],
    }
    summaries: list[dict[str, Any]] = []
    for part in sorted(parts_dir.glob("part-*.csv.gz")):
        frame = pd.read_csv(part, compression="gzip")
        cache_path = cache_dir / f"{part.stem.replace('.csv', '')}.npy"
        vectors = np.load(cache_path, mmap_mode="r") if cache_path.is_file() else None
        for intent, (centroid_vectors, names) in centroids.items():
            mask = frame["Intent"].eq(intent)
            if not mask.any():
                continue
            if vectors is None:
                vectors = encode_queries(model, frame["Normalized"], batch_size, embedding_prefix)
            similarity = vectors[mask.to_numpy()] @ centroid_vectors.T
            indexes = np.argmax(similarity, axis=1)
            frame.loc[mask, "Cluster"] = [names[index] for index in indexes]
            frame.loc[mask, "Cluster probability"] = np.max(similarity, axis=1)
        if "Cluster" not in frame.columns:
            frame["Cluster"] = "Unclustered"
        else:
            frame["Cluster"] = frame["Cluster"].fillna("Unclustered")
        if "Cluster probability" not in frame.columns:
            frame["Cluster probability"] = 0.0
        else:
            frame["Cluster probability"] = pd.to_numeric(frame["Cluster probability"], errors="coerce").fillna(0.0)
        if "Review reason" not in frame.columns:
            frame["Review reason"] = ""
        frame.loc[
            frame["Intent"].eq("human_review"), "Review reason"
        ] = "Relevance quarantine: probability is between review and garbage thresholds"
        for correction_column in ("Correct Intent", "Correct Cluster", "Reviewer Notes"):
            if correction_column not in frame.columns:
                frame[correction_column] = ""
        classification_threshold = float(config.get("classification_review_threshold", 0.65))
        cluster_threshold = float(config.get("clustering", {}).get("review_threshold", 0.35))
        low = frame[
            frame["Classification confidence"].lt(classification_threshold)
            | frame["Cluster"].eq("Unclustered")
            | frame["Cluster probability"].lt(cluster_threshold)
        ].copy()
        if not low.empty:
            low["Review reason"] = "Low confidence, unclustered, or weak cluster"
            review_candidates["uncertain"].append(
                low.sort_values(["Classification confidence", "Search Volume"], ascending=[True, False]).head(200)
            )
        garbage_boundary = frame[frame["Intent"].eq("garbage")].nlargest(100, "Topic relevance").copy()
        if not garbage_boundary.empty:
            garbage_boundary["Review reason"] = "Garbage near the topic-relevance boundary"
            review_candidates["garbage_boundary"].append(garbage_boundary)
        if "P(garbage)" in frame.columns:
            relevance_boundary = frame.copy()
            relevance_boundary["_relevance_margin"] = (
                pd.to_numeric(relevance_boundary["P(garbage)"], errors="coerce").fillna(0)
                - classifier_bundle["settings"]["garbage_threshold"]
            ).abs()
            relevance_boundary = relevance_boundary.nsmallest(100, "_relevance_margin")
            relevance_boundary["Review reason"] = "Topic relevance probability boundary"
            review_candidates["relevance_boundary"].append(relevance_boundary)
        if {"P(commercial)", "P(informational)"}.issubset(frame.columns):
            intent_boundary = frame.copy()
            intent_boundary["_intent_margin"] = (intent_boundary["P(commercial)"] - intent_boundary["P(informational)"]).abs()
            intent_boundary = intent_boundary.nsmallest(75, "_intent_margin")
            intent_boundary["Review reason"] = "Commercial/informational probability boundary"
            review_candidates["intent_boundary"].append(intent_boundary)
        high_value = frame.nlargest(75, "Search Volume").copy()
        if not high_value.empty:
            high_value["Review reason"] = "High-priority phrase"
            review_candidates["high_value"].append(high_value)
        clustered = frame[frame["Cluster"].ne("Unclustered")]
        if not clustered.empty:
            cluster_rows = (
                clustered.sort_values("Search Volume", ascending=False)
                .groupby(["Intent", "Cluster"], dropna=False, group_keys=False)
                .head(1)
                .head(75)
                .copy()
            )
            cluster_rows["Review reason"] = "Representative high-priority phrase in cluster"
            review_candidates["cluster_representative"].append(cluster_rows)
        frame.to_csv(part, index=False, compression="gzip", encoding="utf-8")
        grouped = frame.groupby(["Intent", "Cluster"], dropna=False).agg(
            Phrases=("Phrase", "size"),
            Total_Search_Volume=("Search Volume", "sum"),
            Average_Confidence=("Classification confidence", "mean"),
            Average_Garbage_Probability=("P(garbage)", "mean"),
        ).reset_index()
        summaries.extend(grouped.to_dict("records"))

    summary = pd.DataFrame(summaries)
    if not summary.empty:
        summary = summary.groupby(["Intent", "Cluster"], dropna=False).agg(
            Phrases=("Phrases", "sum"),
            Total_Search_Volume=("Total_Search_Volume", "sum"),
            Average_Confidence=("Average_Confidence", "mean"),
            Average_Garbage_Probability=("Average_Garbage_Probability", "mean"),
        ).reset_index().sort_values("Total_Search_Volume", ascending=False)
    review_parts: list[pd.DataFrame] = []
    queue_limits = {
        "uncertain": 175,
        "garbage_boundary": 100,
        "relevance_boundary": 100,
        "intent_boundary": 75,
        "high_value": 75,
        "cluster_representative": 75,
    }
    for kind, limit in queue_limits.items():
        batches = review_candidates[kind]
        if not batches:
            continue
        candidates = pd.concat(batches, ignore_index=True)
        if kind == "uncertain":
            candidates = candidates.sort_values(["Classification confidence", "Search Volume"], ascending=[True, False])
        elif kind == "intent_boundary":
            candidates = candidates.sort_values(["_intent_margin", "Search Volume"], ascending=[True, False])
        elif kind == "garbage_boundary":
            candidates = candidates.sort_values(["Topic relevance", "Search Volume"], ascending=[False, False])
        elif kind == "relevance_boundary":
            candidates = candidates.sort_values(["_relevance_margin", "Search Volume"], ascending=[True, False])
        else:
            candidates = candidates.sort_values("Search Volume", ascending=False)
        review_parts.append(candidates.head(limit))
    uncertain_frame = pd.concat(review_parts, ignore_index=True, sort=False) if review_parts else pd.DataFrame()
    if not uncertain_frame.empty:
        review_key = uncertain_frame.get("Normalized", uncertain_frame["Phrase"]).astype(str)
        reason_map = (
            pd.DataFrame({"key": review_key, "reason": uncertain_frame["Review reason"]})
            .groupby("key", sort=False)["reason"]
            .agg(lambda values: " | ".join(dict.fromkeys(values)))
        )
        uncertain_frame["_review_key"] = review_key
        uncertain_frame = uncertain_frame.drop_duplicates("_review_key", keep="first").head(300).copy()
        uncertain_frame["Review reason"] = uncertain_frame["_review_key"].map(reason_map)
        uncertain_frame = uncertain_frame.drop(columns=[column for column in ("_review_key", "_intent_margin") if column in uncertain_frame])
    uncertain_path = output_dir / "uncertain_review.csv"
    # These two exports are intended for people as well as scripts.  The BOM
    # lets desktop Excel on Windows recognise Russian UTF-8 text correctly.
    uncertain_frame.to_csv(uncertain_path, index=False, encoding="utf-8-sig")
    summary_path = output_dir / "cluster_summary.csv"
    summary.to_csv(summary_path, index=False, encoding="utf-8-sig")
    source_stem = Path(source_name).stem if source_name else input_path.stem
    report_path = output_dir / f"{source_stem}_clustered.xlsx"
    audit_info = build_large_cluster_relevance_audit(
        parts_dir,
        summary,
        output_dir / "cluster_relevance_audit.json",
        str(config.get("workflow_job_id", "")),
    )
    manifest = {
        "status": "awaiting_cluster_relevance_review",
        "workflow_job_id": config.get("workflow_job_id", ""),
        "input": source_name or str(input_path.resolve()),
        "format": "partitioned_csv_gzip",
        "parts_directory": str(parts_dir.resolve()),
        "parts": len(list(parts_dir.glob("part-*.csv.gz"))),
        "classified_rows_per_chunk_deduplicated": total_rows,
        "chunk_size": chunk_size,
        "embedding_model": str(model_path),
        "execution_device": execution_device,
        "embedding_batch_size": batch_size,
        "garbage_threshold": classifier_bundle["settings"]["garbage_threshold"],
        "quarantine_threshold": classifier_bundle["settings"].get("quarantine_threshold"),
        "clustering": "MiniBatchKMeans on deterministic representative samples; all rows assigned to nearest sample centroid.",
        "uncertain_review": str(uncertain_path.resolve()),
        "cluster_summary": str(summary_path.resolve()),
        "excel_result": str(report_path.resolve()),
        # Kept for compatibility with existing workflow state readers.
        "excel_summary": str(report_path.resolve()),
        "cluster_relevance_audit": audit_info,
    }
    (output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    import shutil
    shutil.rmtree(cache_dir, ignore_errors=True)
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Run supervised SEO classification and clustering.")
    parser.add_argument("input")
    parser.add_argument("--config", type=Path, required=True)
    parser.add_argument("--labels", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    result = run_pipeline(args.input, args.config, args.labels, args.output)
    print(f"Output workbook: {result.resolve()}")


if __name__ == "__main__":
    main()
